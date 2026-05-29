"""
库房当前库存 Excel 解析，供 ``POST /tl/import_warehouse_inventory_excel`` 使用。
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Tuple

_WAREHOUSE_NAME_CANDIDATES = ("库房名称", "仓库名称", "名称", "库房名")
_INVENTORY_CANDIDATES = ("当前库存", "库存", "库存(吨)", "库存吨数")
_DATE_CANDIDATES = ("库存日期", "日期", "inventory_date")


class WarehouseInventoryExcelError(ValueError):
    """表头无法识别、工作簿为空等。"""


@dataclass
class WarehouseInventoryImportRow:
    excel_row: int
    warehouse_name: str
    inventory_ton: Decimal
    inventory_date: Optional[date] = None


def _norm_header(s: object) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", "", str(s).strip())


def _cell_str(v: object) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v != v:
        return ""
    return str(v).strip()


def _resolve_col_index(headers: List[str], candidates: Tuple[str, ...]) -> Optional[int]:
    norm = {_norm_header(h): i for i, h in enumerate(headers) if h}
    for cand in candidates:
        idx = norm.get(_norm_header(cand))
        if idx is not None:
            return idx
    return None


def _cell_decimal(v: object) -> Optional[Decimal]:
    s = _cell_str(v)
    if not s:
        return None
    if s in ("-", "—", "无", "暂无", "NA", "N/A", "待定", "null", "None"):
        return None
    s = s.replace(",", "").replace("，", "")
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _parse_date_cell(v: object) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel

            dt = from_excel(v)
            if isinstance(dt, datetime):
                return dt.date()
            if isinstance(dt, date):
                return dt
        except Exception:
            pass
    s = _cell_str(v)
    if not s:
        return None
    s = s[:10]
    try:
        return date.fromisoformat(s)
    except ValueError:
        return None


def _find_header_row(all_rows: List[tuple], *, max_scan_rows: int = 30) -> Tuple[int, List[str]]:
    scan = min(len(all_rows), max_scan_rows)
    for ri in range(scan):
        row = all_rows[ri]
        if not row:
            continue
        headers = [_cell_str(c) for c in row]
        has_wh = _resolve_col_index(headers, _WAREHOUSE_NAME_CANDIDATES) is not None
        has_inv = _resolve_col_index(headers, _INVENTORY_CANDIDATES) is not None
        if has_wh and has_inv:
            return ri, headers
    raise WarehouseInventoryExcelError(
        "未识别到表头：需包含「库房名称」与「当前库存」列"
    )


def parse_warehouse_inventory_workbook(content: bytes) -> Tuple[List[WarehouseInventoryImportRow], Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise WarehouseInventoryExcelError("缺少 openpyxl 依赖") from e

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = wb["导入数据"] if "导入数据" in wb.sheetnames else wb.active
    if sheet is None:
        raise WarehouseInventoryExcelError("工作簿为空")

    all_rows = list(sheet.iter_rows(values_only=True))
    if not all_rows:
        raise WarehouseInventoryExcelError("工作表无数据")

    header_row_idx, headers = _find_header_row(all_rows)
    wh_col = _resolve_col_index(headers, _WAREHOUSE_NAME_CANDIDATES)
    inv_col = _resolve_col_index(headers, _INVENTORY_CANDIDATES)
    date_col = _resolve_col_index(headers, _DATE_CANDIDATES)
    if wh_col is None or inv_col is None:
        raise WarehouseInventoryExcelError("表头缺少库房名称或当前库存列")

    parsed: List[WarehouseInventoryImportRow] = []
    skipped_empty = 0
    for ri, row in enumerate(all_rows[header_row_idx + 1 :], start=header_row_idx + 2):
        if not row:
            skipped_empty += 1
            continue
        wh_name = _cell_str(row[wh_col] if wh_col < len(row) else None)
        inv_val = _cell_decimal(row[inv_col] if inv_col < len(row) else None)
        if not wh_name and inv_val is None:
            skipped_empty += 1
            continue
        if not wh_name:
            continue
        if inv_val is None:
            continue
        inv_date: Optional[date] = None
        if date_col is not None and date_col < len(row):
            inv_date = _parse_date_cell(row[date_col])
        parsed.append(
            WarehouseInventoryImportRow(
                excel_row=ri,
                warehouse_name=wh_name,
                inventory_ton=inv_val,
                inventory_date=inv_date,
            )
        )

    meta = {
        "sheet": sheet.title,
        "header_row": header_row_idx + 1,
        "parsed_rows": len(parsed),
        "skipped_empty": skipped_empty,
    }
    return parsed, meta
