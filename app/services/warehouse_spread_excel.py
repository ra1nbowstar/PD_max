"""
「库房分析」类 Excel：解析库房名称、对标城市、差额与毛利，供差额配置导入接口使用。
"""

from __future__ import annotations

import io
import re
import unicodedata
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Tuple

_WAREHOUSE_NAME_CANDIDATES = ("库房名称", "仓库名称", "名称", "库房名")
_GROSS_MARGIN_CANDIDATES = ("毛利", "保底毛利", "实际毛利")
_WAREHOUSE_PRICE_CANDIDATES = ("定价", "库房报价")
_REGION_CANDIDATES = ("地区", "省份", "省")
_SKIP_WAREHOUSE_NAMES = frozenset({"战略库房", "合计"})


class WarehouseSpreadExcelError(ValueError):
    """表头无法识别、工作簿为空等。"""


@dataclass
class SpreadImportRow:
    warehouse_name: str
    sheet_name: str
    excel_row: int
    benchmark_city: Optional[str] = None
    city_spread: Optional[Decimal] = None
    gross_margin_config: Optional[Decimal] = None
    warehouse_price: Optional[Decimal] = None
    region: Optional[str] = None


@dataclass
class SheetParseSummary:
    sheet_name: str
    parsed_rows: int = 0
    skipped_rows: int = 0
    columns: Dict[str, Optional[str]] = field(default_factory=dict)


def _norm_header(s: object) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", "", str(s).strip())


def _cell_str(v: object) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v != v:
        return ""
    return str(v).strip()


def _cell_decimal(v: object) -> Optional[Decimal]:
    s = _cell_str(v)
    if not s:
        return None
    if s in ("-", "—", "无", "暂无", "NA", "N/A", "待定", "null", "None"):
        return None
    s = s.replace(",", "").replace("，", "")
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _resolve_col_index(headers: List[str], candidates: Tuple[str, ...]) -> Optional[int]:
    norm = {_norm_header(h): i for i, h in enumerate(headers) if h}
    for cand in candidates:
        idx = norm.get(_norm_header(cand))
        if idx is not None:
            return idx
    return None


def _find_spread_col(headers: List[str]) -> Optional[int]:
    for i, h in enumerate(headers):
        nh = _norm_header(h)
        if not nh:
            continue
        if "对比" in nh and ("差额" in nh or "差价" in nh):
            return i
        if "对标" in nh and ("差额" in nh or "差价" in nh):
            return i
    return None


def _benchmark_city_from_header(header: str) -> str:
    h = _cell_str(header)
    if not h:
        return ""
    m = re.search(r"对比(.+?)(?:差额|差价)", h)
    if m:
        return m.group(1).strip()
    m = re.search(r"对标(.+?)(?:\d|[\s]*(?:差额|差价))", h)
    if m:
        return m.group(1).strip()
    return ""


def _parse_spread_cell(
    v: object, *, header_city: str
) -> Tuple[Optional[str], Optional[Decimal]]:
    if v is None:
        return (header_city or None), None
    if isinstance(v, bool):
        return (header_city or None), None
    if isinstance(v, (int, float)):
        return (header_city or None), Decimal(str(v))

    s = _cell_str(v)
    if not s:
        return (header_city or None), None

    m = re.search(
        r"对比(.+?)(\d+(?:\.\d+)?).*?差价\s*([+-]?\d+(?:\.\d+)?)",
        s,
    )
    if m:
        return m.group(1).strip(), Decimal(m.group(3))

    m = re.search(r"以([\u4e00-\u9fff]{2,4})为(?:价格)?基准", s)
    if m:
        return m.group(1).strip(), None

    m = re.search(r"([+-]?\d+(?:\.\d+)?)\s*$", s)
    if m and header_city:
        return header_city, Decimal(m.group(1))

    if header_city:
        return header_city, None

    return None, None


def _region_benchmark_city(region: str) -> str:
    s = _cell_str(region)
    if not s:
        return ""
    s = re.sub(r"\s+", "", s)
    for prefix in (
        "黑龙江",
        "内蒙古",
        "河北",
        "山西",
        "辽宁",
        "吉林",
        "江苏",
        "浙江",
        "安徽",
        "福建",
        "江西",
        "山东",
        "河南",
        "湖北",
        "湖南",
        "广东",
        "海南",
        "四川",
        "贵州",
        "云南",
        "陕西",
        "甘肃",
        "青海",
        "台湾",
        "广西",
        "西藏",
        "宁夏",
        "新疆",
        "北京",
        "天津",
        "上海",
        "重庆",
    ):
        if s.startswith(prefix) and len(s) > len(prefix):
            rest = s[len(prefix) :].strip("省市县区")
            if rest:
                return rest
    return s


def _merge_row(existing: SpreadImportRow, incoming: SpreadImportRow) -> SpreadImportRow:
    return SpreadImportRow(
        warehouse_name=existing.warehouse_name,
        sheet_name=incoming.sheet_name,
        excel_row=incoming.excel_row,
        benchmark_city=incoming.benchmark_city or existing.benchmark_city,
        city_spread=incoming.city_spread if incoming.city_spread is not None else existing.city_spread,
        gross_margin_config=(
            incoming.gross_margin_config
            if incoming.gross_margin_config is not None
            else existing.gross_margin_config
        ),
        warehouse_price=(
            incoming.warehouse_price if incoming.warehouse_price is not None else existing.warehouse_price
        ),
        region=incoming.region or existing.region,
    )


def canonical_warehouse_key(name: str) -> str:
    s = unicodedata.normalize("NFKC", name.strip())
    s = s.replace("\u2ee2", "\u9a6c").replace("\u2f8f", "\u9a6c")
    return re.sub(r"\s+", "", s)


def parse_warehouse_spread_workbook(content: bytes) -> Tuple[Dict[str, SpreadImportRow], List[SheetParseSummary]]:
    if not content:
        raise WarehouseSpreadExcelError("文件内容为空")
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise WarehouseSpreadExcelError("服务端未安装 openpyxl，无法导入 Excel") from e

    merged: Dict[str, SpreadImportRow] = {}
    summaries: List[SheetParseSummary] = []

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            summary = SheetParseSummary(sheet_name=sheet_name)
            if not rows:
                summaries.append(summary)
                continue

            headers = [_cell_str(c) for c in rows[0]]
            wh_idx = _resolve_col_index(headers, _WAREHOUSE_NAME_CANDIDATES)
            if wh_idx is None:
                summary.skipped_rows = max(0, len(rows) - 1)
                summaries.append(summary)
                continue

            gm_idx = _resolve_col_index(headers, _GROSS_MARGIN_CANDIDATES)
            price_idx = _resolve_col_index(headers, _WAREHOUSE_PRICE_CANDIDATES)
            region_idx = _resolve_col_index(headers, _REGION_CANDIDATES)
            spread_idx = _find_spread_col(headers)
            header_city = _benchmark_city_from_header(headers[spread_idx]) if spread_idx is not None else ""

            summary.columns = {
                "库房名称": headers[wh_idx],
                "毛利": headers[gm_idx] if gm_idx is not None else None,
                "定价": headers[price_idx] if price_idx is not None else None,
                "地区": headers[region_idx] if region_idx is not None else None,
                "差额列": headers[spread_idx] if spread_idx is not None else None,
            }

            carry_region = ""
            for offset, row in enumerate(rows[1:], start=2):
                wh_raw = row[wh_idx] if wh_idx < len(row) else None
                wh_name = _cell_str(wh_raw)
                if wh_name:
                    if wh_name in _SKIP_WAREHOUSE_NAMES:
                        summary.skipped_rows += 1
                        continue
                    carry_region = ""
                elif not wh_name:
                    summary.skipped_rows += 1
                    continue

                region_val = _cell_str(row[region_idx]) if region_idx is not None and region_idx < len(row) else ""
                if region_val:
                    carry_region = region_val
                region = region_val or carry_region

                gm = _cell_decimal(row[gm_idx]) if gm_idx is not None and gm_idx < len(row) else None
                wh_price = (
                    _cell_decimal(row[price_idx]) if price_idx is not None and price_idx < len(row) else None
                )
                spread_raw = row[spread_idx] if spread_idx is not None and spread_idx < len(row) else None
                bench_city, city_spread = _parse_spread_cell(spread_raw, header_city=header_city)
                if not bench_city and header_city:
                    bench_city = header_city
                if not bench_city and region:
                    bench_city = _region_benchmark_city(region) or None

                if (
                    gm is None
                    and wh_price is None
                    and city_spread is None
                    and not bench_city
                ):
                    summary.skipped_rows += 1
                    continue

                parsed = SpreadImportRow(
                    warehouse_name=wh_name,
                    sheet_name=sheet_name,
                    excel_row=offset,
                    benchmark_city=bench_city,
                    city_spread=city_spread,
                    gross_margin_config=gm,
                    warehouse_price=wh_price,
                    region=region or None,
                )
                prev = merged.get(wh_name)
                merged[wh_name] = _merge_row(prev, parsed) if prev else parsed
                summary.parsed_rows += 1

            summaries.append(summary)
    finally:
        wb.close()

    return merged, summaries
