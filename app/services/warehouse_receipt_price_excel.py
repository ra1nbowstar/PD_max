"""
库房按品种收货价格 Excel 解析，供 ``POST /tl/import_warehouse_receipt_prices_excel`` 使用。
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Tuple

_WAREHOUSE_NAME_CANDIDATES = ("库房名称", "仓库名称", "名称", "库房名")
_CATEGORY_CANDIDATES = ("回收品种", "品种", "品类", "品类名称", "category")
_PRICE_CANDIDATES = ("价格", "收货价格", "回收单价", "元每吨", "单价", "元/吨")


class WarehouseReceiptPriceExcelError(ValueError):
    """表头无法识别、工作簿为空等。"""


@dataclass
class WarehouseReceiptPriceImportRow:
    excel_row: int
    warehouse_name: str
    category_name: str
    price_per_ton: Decimal


def _norm_header(s: object) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", "", str(s).strip())


def _cell_str(v: object) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v != v:
        return ""
    return str(v).strip()


def _resolve_col_index(headers: List[str], candidates: Tuple[str, ...]) -> Optional[int]:
    norm = {_norm_header(h): i for i, h in enumerate(headers) if h}
    for cand in candidates:
        idx = norm.get(_norm_header(cand))
        if idx is not None:
            return idx
    return None


def _cell_decimal(v: object) -> Optional[Decimal]:
    s = _cell_str(v)
    if not s:
        return None
    if s in ("-", "—", "无", "暂无", "NA", "N/A", "待定", "null", "None"):
        return None
    s = s.replace(",", "").replace("，", "")
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _find_header_row(all_rows: List[tuple], *, max_scan_rows: int = 30) -> Tuple[int, List[str]]:
    scan = min(len(all_rows), max_scan_rows)
    for ri in range(scan):
        row = all_rows[ri]
        if not row:
            continue
        headers = [_cell_str(c) for c in row]
        has_wh = _resolve_col_index(headers, _WAREHOUSE_NAME_CANDIDATES) is not None
        has_cat = _resolve_col_index(headers, _CATEGORY_CANDIDATES) is not None
        has_price = _resolve_col_index(headers, _PRICE_CANDIDATES) is not None
        if has_wh and has_cat and has_price:
            return ri, headers
    raise WarehouseReceiptPriceExcelError(
        "未识别到表头：需包含「库房名称」「回收品种」「价格」列"
    )


def parse_warehouse_receipt_price_workbook(
    content: bytes,
) -> Tuple[List[WarehouseReceiptPriceImportRow], Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise WarehouseReceiptPriceExcelError("缺少 openpyxl 依赖") from e

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = wb["导入数据"] if "导入数据" in wb.sheetnames else wb.active
    if sheet is None:
        raise WarehouseReceiptPriceExcelError("工作簿为空")

    all_rows = list(sheet.iter_rows(values_only=True))
    if not all_rows:
        raise WarehouseReceiptPriceExcelError("工作表无数据")

    header_row_idx, headers = _find_header_row(all_rows)
    wh_col = _resolve_col_index(headers, _WAREHOUSE_NAME_CANDIDATES)
    cat_col = _resolve_col_index(headers, _CATEGORY_CANDIDATES)
    price_col = _resolve_col_index(headers, _PRICE_CANDIDATES)
    if wh_col is None or cat_col is None or price_col is None:
        raise WarehouseReceiptPriceExcelError("表头缺少库房名称、回收品种或价格列")

    parsed: List[WarehouseReceiptPriceImportRow] = []
    skipped_empty = 0
    for ri, row in enumerate(all_rows[header_row_idx + 1 :], start=header_row_idx + 2):
        if not row:
            skipped_empty += 1
            continue
        wh_name = _cell_str(row[wh_col] if wh_col < len(row) else None)
        cat_name = _cell_str(row[cat_col] if cat_col < len(row) else None)
        price_val = _cell_decimal(row[price_col] if price_col < len(row) else None)
        if not wh_name and not cat_name and price_val is None:
            skipped_empty += 1
            continue
        if not wh_name or not cat_name:
            continue
        if price_val is None:
            continue
        parsed.append(
            WarehouseReceiptPriceImportRow(
                excel_row=ri,
                warehouse_name=wh_name,
                category_name=cat_name,
                price_per_ton=price_val,
            )
        )

    meta = {
        "sheet": sheet.title,
        "header_row": header_row_idx + 1,
        "parsed_rows": len(parsed),
        "skipped_empty": skipped_empty,
    }
    return parsed, meta
