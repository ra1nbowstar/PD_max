"""
冶炼厂标定价格 Excel 解析，供 ``POST /tl/import_smelter_calibration_excel`` 使用。
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Tuple


class SmelterCalibrationExcelError(ValueError):
    """表头无法识别、工作簿为空等。"""


_FACTORY_ID_CANDIDATES = ("冶炼厂id", "冶炼厂ID", "工厂id", "工厂ID", "factory_id", "factoryid")
_FACTORY_NAME_CANDIDATES = ("冶炼厂", "冶炼厂名称", "冶炼厂名", "工厂", "工厂名称", "厂家")
_PRICE_CANDIDATES = ("标定价格", "标定价", "标定单价", "价格", "单价")
_DATE_CANDIDATES = ("定价日期", "日期", "价格日期", "price_date")


@dataclass
class SmelterCalibrationImportRow:
    excel_row: int
    factory_id: Optional[int]
    factory_name: Optional[str]
    calibration_price: Decimal
    price_date: Optional[date] = None


def _norm_header(s: object) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", "", str(s).strip())


def _cell_str(v: object) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v != v:
        return ""
    return str(v).replace("\u3000", " ").strip()


def _resolve_col_index(headers: List[str], candidates: Tuple[str, ...]) -> Optional[int]:
    norm = {_norm_header(h): i for i, h in enumerate(headers) if h}
    for cand in candidates:
        idx = norm.get(_norm_header(cand))
        if idx is not None:
            return idx
    return None


def _cell_decimal(v: object) -> Optional[Decimal]:
    s = _cell_str(v)
    if not s:
        return None
    if s in ("-", "—", "无", "暂无", "NA", "N/A", "待定", "null", "None"):
        return None
    s = s.replace(",", "").replace("，", "")
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _parse_date_cell(v: object) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel

            dt = from_excel(v)
            if isinstance(dt, datetime):
                return dt.date()
            if isinstance(dt, date):
                return dt
        except Exception:
            pass
    s = _cell_str(v)
    if not s:
        return None
    s = s[:10]
    try:
        return date.fromisoformat(s)
    except ValueError:
        return None


def _parse_factory_id_cell(v: object) -> Optional[int]:
    s = _cell_str(v)
    if not s:
        return None
    try:
        fid = int(float(s))
    except (TypeError, ValueError):
        return None
    return fid if fid >= 1 else None


def _find_header_row(all_rows: List[tuple], *, max_scan_rows: int = 30) -> Tuple[int, List[str]]:
    scan = min(len(all_rows), max_scan_rows)
    for ri in range(scan):
        row = all_rows[ri]
        if not row:
            continue
        headers = [_cell_str(c) for c in row]
        has_price = _resolve_col_index(headers, _PRICE_CANDIDATES) is not None
        has_factory = (
            _resolve_col_index(headers, _FACTORY_ID_CANDIDATES) is not None
            or _resolve_col_index(headers, _FACTORY_NAME_CANDIDATES) is not None
        )
        if has_price and has_factory:
            return ri, headers
    raise SmelterCalibrationExcelError(
        "未识别表头：请在表头行提供「冶炼厂/冶炼厂id」与「标定价格」列（可含「定价日期」）。"
    )


def parse_smelter_calibration_workbook(content: bytes) -> Tuple[List[SmelterCalibrationImportRow], Dict[str, Any]]:
    """解析 xlsx 首工作表（或名为「导入数据」的工作表）为标定价格导入行。"""
    if not content:
        raise SmelterCalibrationExcelError("文件内容为空")
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise SmelterCalibrationExcelError("服务端未安装 openpyxl，无法解析 Excel") from e

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb["导入数据"] if "导入数据" in wb.sheetnames else wb.active
        assert ws is not None
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            raise SmelterCalibrationExcelError("工作表为空")

        hdr_idx, headers = _find_header_row(all_rows)
        factory_id_col = _resolve_col_index(headers, _FACTORY_ID_CANDIDATES)
        factory_name_col = _resolve_col_index(headers, _FACTORY_NAME_CANDIDATES)
        price_col = _resolve_col_index(headers, _PRICE_CANDIDATES)
        date_col = _resolve_col_index(headers, _DATE_CANDIDATES)
        if price_col is None:
            raise SmelterCalibrationExcelError("表头须包含「标定价格」或同义列")

        rows_out: List[SmelterCalibrationImportRow] = []
        skipped_empty = 0
        for offset, row in enumerate(all_rows[hdr_idx + 1 :]):
            excel_row = hdr_idx + 2 + offset
            if not row:
                skipped_empty += 1
                continue

            factory_id: Optional[int] = None
            factory_name: Optional[str] = None
            if factory_id_col is not None and factory_id_col < len(row):
                factory_id = _parse_factory_id_cell(row[factory_id_col])
            if factory_name_col is not None and factory_name_col < len(row):
                name = _cell_str(row[factory_name_col])
                factory_name = name or None

            price = _cell_decimal(row[price_col] if price_col < len(row) else None)
            if factory_id is None and not factory_name:
                if price is None:
                    skipped_empty += 1
                    continue
                raise SmelterCalibrationExcelError(
                    f"第 {excel_row} 行：缺少「冶炼厂」或「冶炼厂id」"
                )
            if price is None:
                if factory_id is None and not factory_name:
                    skipped_empty += 1
                    continue
                raise SmelterCalibrationExcelError(f"第 {excel_row} 行：「标定价格」无效或为空")

            price_date: Optional[date] = None
            if date_col is not None and date_col < len(row):
                price_date = _parse_date_cell(row[date_col])
                if row[date_col] is not None and _cell_str(row[date_col]) and price_date is None:
                    raise SmelterCalibrationExcelError(
                        f"第 {excel_row} 行：「定价日期」格式无效，应为 YYYY-MM-DD"
                    )

            rows_out.append(
                SmelterCalibrationImportRow(
                    excel_row=excel_row,
                    factory_id=factory_id,
                    factory_name=factory_name,
                    calibration_price=price,
                    price_date=price_date,
                )
            )

        meta = {
            "sheet": ws.title,
            "header_row": hdr_idx + 1,
            "columns": {
                "冶炼厂id": headers[factory_id_col] if factory_id_col is not None else None,
                "冶炼厂": headers[factory_name_col] if factory_name_col is not None else None,
                "标定价格": headers[price_col],
                "定价日期": headers[date_col] if date_col is not None else None,
            },
            "parsed_rows": len(rows_out),
            "skipped_empty": skipped_empty,
        }
        return rows_out, meta
    finally:
        wb.close()
