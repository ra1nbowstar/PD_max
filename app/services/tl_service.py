"""
TL比价模块服务层
负责仓库、冶炼厂、品类、比价、运费、价格表、品类映射等数据库操作
"""
import hashlib
import io
import json
import logging
import os
import re
import uuid
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import AbstractSet, Any, Dict, List, Optional, Set, Tuple

from pymysql.err import IntegrityError as PyMySQLIntegrityError

from app.config import UPLOAD_DIR, XUNRONGBAO_SHIPPING_PREMIUM_PER_TON
from app.database import get_conn
from app.finance_log import log_finance_event
from app.models.tl import OPTIMAL_PRICE_BASIS_ALLOWED, UpdateQuoteDetailRequest
from app.quote_price_sources import (
    API_KEY_TO_DB,
    merge_sources_after_fill,
    normalize_client_sources,
    SOURCE_DERIVED,
    SOURCE_ORIGINAL,
)
from app.price_tax_utils import (
    apply_per_ton_premium_to_quote_row,
    derive_net_and_vat_from_quote_row,
    derive_vat_prices_from_stated_price,
    fill_vat_from_exclusive_net,
    inclusive_from_net,
    merge_factory_rates,
    net_from_inclusive,
    parse_price_basis_from_remark,
)
from app.services.tl_dict_geo_crud import (
    CODE_DB as SA_CODE_DB,
    CODE_DUP_LINK as SA_CODE_DUP_LINK,
    CODE_DUP_NAME as SA_CODE_DUP_NAME,
    CODE_INTERNAL as SA_CODE_INTERNAL,
    CODE_NOT_FOUND as SA_CODE_NOT_FOUND,
    CODE_OK as SA_CODE_OK,
    CODE_VALIDATION as SA_CODE_VALIDATION,
    smelter_create as sa_smelter_create,
    smelter_get as sa_smelter_get,
    smelter_list as sa_smelter_list,
    smelter_update as sa_smelter_update,
    warehouse_create as sa_wh_create,
    warehouse_link_bind as sa_wh_link_bind,
    warehouse_link_unbind as sa_wh_link_unbind,
    warehouse_link_update_tier_price_spread as sa_wh_link_update_tier,
    warehouse_links_inbound as sa_wh_links_inbound,
    warehouse_links_outbound as sa_wh_links_outbound,
    warehouse_links_replace_outbound as sa_wh_links_replace_outbound,
    warehouse_links_batch_bind as sa_wh_links_batch_bind,
    warehouse_links_batch_unbind as sa_wh_links_batch_unbind,
    warehouse_links_list_all as sa_wh_links_list_all,
    warehouse_list as sa_wh_list,
    warehouse_update as sa_wh_update,
)
from app.services.vlm_extractor_service import QwenVLFullExtractor, VLMConfig

logger = logging.getLogger(__name__)

_MARKER_HEX_RE = re.compile(r"^#[0-9A-Fa-f]{6}$")


def _comparison_quote_calendar_date() -> date:
    """
    未传「报价日期」时，在 quote_details 中取与**该日历日**距离最近的一条报价。
    默认按 Asia/Shanghai；可通过环境变量 QUOTE_COMPARISON_TZ 设为其它 IANA 时区（如 UTC）。
    """
    tz_name = (os.getenv("QUOTE_COMPARISON_TZ") or "Asia/Shanghai").strip() or "Asia/Shanghai"
    try:
        from zoneinfo import ZoneInfo

        return datetime.now(ZoneInfo(tz_name)).date()
    except Exception:
        logger.warning(
            "QUOTE_COMPARISON_TZ=%r 无效，比价基准日回退为服务器本地当天", tz_name
        )
        return date.today()


def _unit_for_optimal_price_basis(
    basis: str,
    breakdown: Optional[Tuple[float, float, float, float]],
    qrow: Optional[Dict[str, Optional[float]]],
) -> Optional[float]:
    """
    最优价用的单价（元/吨）：base/1pct/3pct/13pct 来自统一反推 breakdown；
    普票、反向发票取表中对应列。
    """
    if basis == "base":
        return breakdown[0] if breakdown else None
    if basis == "1pct":
        return breakdown[1] if breakdown else None
    if basis == "3pct":
        return breakdown[2] if breakdown else None
    if basis == "13pct":
        return breakdown[3] if breakdown else None
    if basis == "normal_invoice":
        if not qrow:
            return None
        v = qrow.get("price_normal_invoice")
        return float(v) if v is not None else None
    if basis == "reverse_invoice":
        if not qrow:
            return None
        v = qrow.get("price_reverse_invoice")
        return float(v) if v is not None else None
    return None


def _build_comparison_price_metrics(
    price: Optional[float],
    source: str,
    qrow: Optional[Dict[str, Optional[float]]],
    merged: Dict[str, float],
    target_tax: Optional[str],
    t: float,
    fr: float,
    bases: List[str],
) -> Dict[str, Any]:
    """
    由 resolve 得到的报价与运费、吨数生成比价行中的价/税/利润字段块
    （不含仓库、冶炼厂、品类等展示维度）。
    """
    if price is not None and target_tax and target_tax in merged:
        p_net = round(net_from_inclusive(float(price), merged[target_tax]), 2)
    elif price is not None:
        p_net = float(price)
    else:
        p_net = None

    freight_cost_total = round(float(fr) * t, 2)

    quote_amount: Optional[float] = (
        round(float(p_net) * t, 2) if p_net is not None else None
    )
    profit = (
        round(quote_amount - freight_cost_total, 2)
        if quote_amount is not None
        else round(-freight_cost_total, 2)
    )

    breakdown = (
        derive_net_and_vat_from_quote_row(qrow, merged) if qrow else None
    )
    if breakdown:
        base_net, p1_vat, p3_vat, _p13 = breakdown
        profit_base = round(base_net * t - freight_cost_total, 2)
        profit_3 = round(p3_vat * t - freight_cost_total, 2)
    else:
        base_net = None
        p1_vat = None
        p3_vat = None
        profit_base = None
        profit_3 = None

    optimal_profits: Dict[str, Optional[float]] = {}
    for b in bases:
        u = _unit_for_optimal_price_basis(b, breakdown, qrow)
        optimal_profits[b] = (
            round(u * t - freight_cost_total, 2) if u is not None else None
        )

    return {
        "单价": p_net if source != "unavailable" else None,
        "总价": quote_amount,
        "运费单价": fr,
        "运费": freight_cost_total,
        "总运费": freight_cost_total,
        "报价": p_net if source != "unavailable" else None,
        "报价金额": quote_amount,
        "报价来源": source,
        "基准价": base_net,
        "含1%税价": p1_vat,
        "含3%税价": p3_vat,
        "利润": profit,
        "利润_基准": profit_base,
        "利润_含3%": profit_3,
        "最优价各口径利润": optimal_profits,
    }


class PurchaseSuggestionLLMError(Exception):
    """采购建议接口调用上游大模型失败（由路由映射为 HTTP 502）。"""


PRICE_TABLE_UPLOAD_DIR = Path(UPLOAD_DIR) / "price_tables"
PRICE_TABLE_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


def _cell_json(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, date):
        return v.isoformat()
    return v


def _json_cell_to_dict(val: Any) -> Optional[Dict[str, Any]]:
    """解析库表 JSON 列（或已解析的 dict）为字典。"""
    if val is None:
        return None
    if isinstance(val, dict):
        return val
    if isinstance(val, (bytes, bytearray)):
        val = val.decode("utf-8")
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        return json.loads(s)
    return None


def _strip_optional_str(v: Optional[str]) -> Optional[str]:
    if v is None:
        return None
    t = str(v).strip()
    return t if t else None


def _split_category_alias_names(raw: Any) -> List[str]:
    """将「大白、大白货车、白壳电池」这类合并别名拆成独立品类别名。"""
    if raw is None:
        return []
    parts = [
        x.strip()
        for x in re.split(r"[、,，]+", str(raw).replace("\u3000", " "))
        if x.strip()
    ]
    return list(dict.fromkeys(parts))


def _color_config_to_json_str(val: Any) -> Optional[str]:
    if val is None:
        return None
    return json.dumps(val, ensure_ascii=False)


def _color_config_from_db(val: Any) -> Any:
    if val is None:
        return None
    if isinstance(val, (dict, list)):
        return val
    if isinstance(val, (bytes, bytearray)):
        val = val.decode("utf-8")
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        return json.loads(s)
    return val


def _strip_nonempty(v: Optional[str]) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _full_cn_site_address(
    province: Optional[str],
    city: Optional[str],
    district: Optional[str],
    street: Optional[str],
) -> bool:
    return bool(
        _strip_nonempty(province)
        and _strip_nonempty(city)
        and _strip_nonempty(district)
        and _strip_nonempty(street)
    )


def _marker_hex_from_wh_color_config(cc: Any) -> Optional[str]:
    """从 TL 仓库颜色配置 JSON/dict 中提取六位 #RRGGBB（供 tl_dict_geo_crud 落库）。"""
    if cc is None:
        return None
    if isinstance(cc, dict):
        d = cc
    else:
        d = _color_config_from_db(cc)
    if isinstance(d, dict):
        h = d.get("marker") or d.get("hex")
        if isinstance(h, str):
            hs = h.strip()
            if _MARKER_HEX_RE.match(hs):
                return hs
    return None


def _raise_tl_geo_crud_result(res: Dict[str, Any]) -> Dict[str, Any]:
    c = int(res.get("code", SA_CODE_INTERNAL))
    if c == SA_CODE_OK:
        return res
    msg = str(res.get("msg") or "操作失败")
    if c in (SA_CODE_VALIDATION, SA_CODE_NOT_FOUND, SA_CODE_DUP_NAME, SA_CODE_DUP_LINK):
        raise ValueError(msg)
    if c in (SA_CODE_DB, SA_CODE_INTERNAL):
        raise RuntimeError(msg)
    raise RuntimeError(msg)


QUOTE_PRICE_ANCHOR_ORDER = (
    "价格",
    "价格_13pct增值税",
    "价格_3pct增值税",
    "价格_1pct增值税",
    "普通发票价格",
    "反向发票价格",
)


def _chinese_item_to_prices_en(
    item: Dict[str, Any],
    touched_cn: Optional[AbstractSet[str]] = None,
) -> Dict[str, Any]:
    """
    将中文键报价列转为 derive_net_and_vat_from_quote_row 所需英文键。
    touched_cn 非空时：仅取「本次请求中改动的列」里、按 QUOTE_PRICE_ANCHOR_ORDER 优先级最高的那一档，
    避免部分更新时旧库里的「基准价」盖住用户新改的含税价。
    """
    if touched_cn:
        for c in QUOTE_PRICE_ANCHOR_ORDER:
            if c in touched_cn and item.get(c) is not None:
                dbk = API_KEY_TO_DB.get(c, c)
                return {dbk: item.get(c)}
        return {}
    out: Dict[str, Any] = {}
    for c in QUOTE_PRICE_ANCHOR_ORDER:
        v = item.get(c)
        if v is not None:
            dbk = API_KEY_TO_DB.get(c, c)
            out[dbk] = v
    return out


def _apply_factory_tax_rates_to_quote_item(
    item: Dict[str, Any],
    tax_by_fid: Dict[int, Dict[str, float]],
    touched_price_keys_cn: Optional[AbstractSet[str]] = None,
) -> bool:
    """
    确认写入 / 修改前：按冶炼厂税率（factory_tax_rates ∪ 默认）统一写入「价格」与含1%/3%/13%列。

    - 全量条目（touched_price_keys_cn=None）：与 derive_net_and_vat_from_quote_row 一致（基准 → 正算含税；
      或仅有含税列 → 反算基准再正算；普票/反向发票列按不含税理解）。
    - 部分更新：仅根据本次改动的价格锚点列重算，避免未改动的旧基准价干扰。
    """
    fid = item.get("冶炼厂id")
    if fid is None:
        return False
    merged = merge_factory_rates(tax_by_fid.get(int(fid)))

    prices_en = _chinese_item_to_prices_en(item, touched_price_keys_cn)
    derived = derive_net_and_vat_from_quote_row(prices_en, merged)
    if derived is None:
        return False

    net, p1, p3, p13 = derived
    item["价格"] = round(float(net), 2)
    item["价格_1pct增值税"] = p1
    item["价格_3pct增值税"] = p3
    item["价格_13pct增值税"] = p13
    return True


class TLService:

    # ==================== 接口0：添加仓库 ====================

    def _resolve_warehouse_type_name_by_id(self, type_id: int) -> Optional[str]:
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT name FROM dict_warehouse_types WHERE id = %s AND is_active = 1",
                        (int(type_id),),
                    )
                    row = cur.fetchone()
                    return row[0] if row else None
        except Exception:
            return None

    def add_warehouse(
        self,
        name: str,
        address: Optional[str] = None,
        warehouse_type_id: Optional[int] = None,
        warehouse_color_config: Optional[Any] = None,
        province: Optional[str] = None,
        city: Optional[str] = None,
        district: Optional[str] = None,
        longitude: Optional[float] = None,
        latitude: Optional[float] = None,
        warehouse_type_name: Optional[str] = None,
        contact_name: Optional[str] = None,
        contact_phone: Optional[str] = None,
        hazardous_waste_license_qty: Optional[float] = None,
        monthly_avg_receipt_ton: Optional[float] = None,
        freight_amount: Optional[float] = None,
    ) -> Dict[str, Any]:
        """极简录入仅 name/地址；含省市区详址时走 tl_dict_geo_crud，经纬度默认天地图（未同时传经度+纬度时）。"""
        addr = _strip_optional_str(address)
        if _full_cn_site_address(province, city, district, addr):
            type_name = _strip_nonempty(warehouse_type_name)
            if not type_name and warehouse_type_id is not None:
                type_name = self._resolve_warehouse_type_name_by_id(int(warehouse_type_id))
            if not type_name:
                raise ValueError(
                    "完整地址模式下需提供「库房类型名」或有效的「仓库类型id」（已在系统中启用）"
                )
            hex_color = _marker_hex_from_wh_color_config(warehouse_color_config)
            payload = {
                "name": str(name).strip(),
                "type": type_name,
                "province": _strip_nonempty(province),
                "city": _strip_nonempty(city),
                "district": _strip_nonempty(district),
                "address": addr,
                "color": hex_color,
                "longitude": longitude,
                "latitude": latitude,
                "status": 1,
            }
            if contact_name is not None:
                payload["contact_name"] = str(contact_name).strip() or None
            if contact_phone is not None:
                payload["contact_phone"] = str(contact_phone).strip() or None
            if hazardous_waste_license_qty is not None:
                payload["hazardous_waste_license_qty"] = hazardous_waste_license_qty
            if monthly_avg_receipt_ton is not None:
                payload["monthly_avg_receipt_ton"] = monthly_avg_receipt_ton
            if freight_amount is not None:
                payload["freight_amount"] = freight_amount
            try:
                res = _raise_tl_geo_crud_result(sa_wh_create(payload))
                data = res.get("data") or {}
                wid = int(data.get("id", 0))
                return {"code": 200, "msg": "仓库新建成功", "仓库id": wid, "新建": True}
            except ValueError:
                raise
            except RuntimeError as e:
                logger.error(f"添加仓库失败(地理落库): {e}")
                raise

        wh_cc_json = (
            _color_config_to_json_str(warehouse_color_config)
            if warehouse_color_config is not None
            else None
        )
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    if warehouse_type_id is not None:
                        cur.execute(
                            "SELECT id FROM dict_warehouse_types "
                            "WHERE id = %s AND is_active = 1",
                            (warehouse_type_id,),
                        )
                        if not cur.fetchone():
                            raise ValueError(
                                f"库房类型 id={warehouse_type_id} 不存在或未启用，"
                                f"请先用 add_warehouse_type 维护类型"
                            )
                    cur.execute(
                        "SELECT id FROM dict_warehouses WHERE name = %s",
                        (name,),
                    )
                    row = cur.fetchone()
                    if row:
                        return {"code": 200, "msg": "仓库已存在", "仓库id": row[0], "新建": False}
                    cur.execute(
                        "INSERT INTO dict_warehouses "
                        "(name, address, warehouse_type_id, color_config, is_active) "
                        "VALUES (%s, %s, %s, %s, 1)",
                        (name, addr, warehouse_type_id, wh_cc_json),
                    )
                    return {"code": 200, "msg": "仓库新建成功", "仓库id": cur.lastrowid, "新建": True}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"添加仓库失败: {e}")
            raise

    # ==================== 接口0b：新建冶炼厂 ====================

    def add_smelter(
        self,
        name: str,
        address: Optional[str] = None,
        province: Optional[str] = None,
        city: Optional[str] = None,
        district: Optional[str] = None,
        longitude: Optional[float] = None,
        latitude: Optional[float] = None,
    ) -> Dict[str, Any]:
        """全地址落库时经纬度由天地图解析；未传经度/纬度或只传其一由 maybe_geocode 处理（可回退为 NULL）。"""
        addr = _strip_optional_str(address)
        if _full_cn_site_address(province, city, district, addr):
            payload = {
                "name": str(name).strip(),
                "province": _strip_nonempty(province),
                "city": _strip_nonempty(city),
                "district": _strip_nonempty(district),
                "address": addr,
                "longitude": longitude,
                "latitude": latitude,
                "status": 1,
            }
            try:
                res = _raise_tl_geo_crud_result(sa_smelter_create(payload))
                data = res.get("data") or {}
                fid = int(data.get("id", 0))
                return {"code": 200, "msg": "冶炼厂新建成功", "冶炼厂id": fid, "新建": True}
            except ValueError:
                raise
            except RuntimeError as e:
                logger.error(f"新建冶炼厂失败(地理落库): {e}")
                raise

        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT id, is_active FROM dict_factories WHERE name = %s",
                        (name,),
                    )
                    row = cur.fetchone()
                    if row:
                        smelter_id, is_active = row
                        if is_active == 1:
                            return {"code": 200, "msg": "冶炼厂已存在", "冶炼厂id": smelter_id, "新建": False}
                        cur.execute(
                            "UPDATE dict_factories SET is_active = 1 WHERE id = %s",
                            (smelter_id,),
                        )
                        return {"code": 200, "msg": "冶炼厂已恢复启用", "冶炼厂id": smelter_id, "新建": False}

                    cur.execute(
                        "INSERT INTO dict_factories (name, address, is_active) VALUES (%s, %s, 1)",
                        (name, addr),
                    )
                    return {"code": 200, "msg": "冶炼厂新建成功", "冶炼厂id": cur.lastrowid, "新建": True}
        except Exception as e:
            logger.error(f"新建冶炼厂失败: {e}")
            raise

    # ==================== 接口1：获取仓库列表 ====================

    def _warehouse_type_ids_by_names(self, names: List[str]) -> Dict[str, int]:
        names = [n.strip() for n in names if n and str(n).strip()]
        if not names:
            return {}
        uniq = list(dict.fromkeys(names))
        out: Dict[str, int] = {}
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    ph = ",".join(["%s"] * len(uniq))
                    cur.execute(
                        f"SELECT name, id FROM dict_warehouse_types "
                        f"WHERE is_active = 1 AND name IN ({ph})",
                        tuple(uniq),
                    )
                    for name, tid in cur.fetchall():
                        out[str(name)] = int(tid)
            return out
        except Exception as e:
            logger.warning(f"按名称解析库房类型 id 失败: {e}")
            return {}

    def _site_wh_item_to_tl_row(
        self,
        item: Dict[str, Any],
        type_id_by_name: Dict[str, int],
    ) -> Dict[str, Any]:
        tname = str(item.get("type") or "").strip()
        wt_id = type_id_by_name.get(tname) if tname else None
        hex_c = item.get("color")
        wh_cc = {"marker": hex_c} if hex_c else None
        rec: Dict[str, Any] = {
            "仓库id": int(item["id"]),
            "仓库名": item["name"],
            "地址": item.get("address") or "",
            "省": item.get("province") or "",
            "市": item.get("city") or "",
            "区": item.get("district") or "",
            "经度": item.get("longitude"),
            "纬度": item.get("latitude"),
            "库房联系人": item.get("contactName") or "",
            "电话": item.get("contactPhone") or "",
            "危废经营许可数量": item.get("hazardousWasteLicenseQty"),
            "月均收货": item.get("monthlyAvgReceiptTon"),
            "运费": item.get("freightAmount"),
            "仓库类型id": wt_id,
            "类型": tname,
            "库房类型颜色配置": None,
            "仓库颜色配置": wh_cc,
            "颜色配置": None,
        }
        return rec

    def _batch_warehouse_type_colors(self, ids: List[int]) -> Dict[int, Any]:
        ids = [int(x) for x in ids if x is not None]
        if not ids:
            return {}
        uniq = list(dict.fromkeys(ids))
        out: Dict[int, Any] = {}
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    ph = ",".join(["%s"] * len(uniq))
                    cur.execute(
                        f"SELECT id, color_config FROM dict_warehouse_types "
                        f"WHERE id IN ({ph})",
                        tuple(uniq),
                    )
                    for tid, raw_cc in cur.fetchall():
                        out[int(tid)] = _color_config_from_db(raw_cc)
        except Exception as e:
            logger.warning(f"批量加载库房类型颜色失败: {e}")
        return out

    def get_warehouses(
        self,
        keyword: Optional[str] = None,
        page: Optional[int] = None,
        size: Optional[int] = None,
        province: Optional[str] = None,
        city: Optional[str] = None,
        district: Optional[str] = None,
        status: Optional[int] = None,
    ) -> Any:
        if page is not None:
            try:
                pg = max(1, int(page))
                sz = min(200, max(1, int(size or 20)))
                kw = (
                    str(keyword).strip()
                    if keyword is not None and str(keyword).strip()
                    else None
                )
                eff_status = status if status is not None else 1
                res = _raise_tl_geo_crud_result(
                    sa_wh_list(pg, sz, kw, None, province, city, district, eff_status)
                )
                payload = res["data"] or {}
                items_raw = payload.get("list") or []
                tnames = [
                    str(x.get("type") or "").strip()
                    for x in items_raw
                    if x.get("type")
                ]
                tid_by_name = self._warehouse_type_ids_by_names(tnames)
                tids = [tid_by_name[t] for t in tnames if t in tid_by_name]
                tcol = self._batch_warehouse_type_colors(tids)
                out_rows: List[Dict[str, Any]] = []
                for x in items_raw:
                    tname = str(x.get("type") or "").strip()
                    wt_id = tid_by_name.get(tname) if tname else None
                    t_cc = tcol.get(int(wt_id)) if wt_id is not None else None
                    row = self._site_wh_item_to_tl_row(x, tid_by_name)
                    row["库房类型颜色配置"] = t_cc
                    row["颜色配置"] = t_cc
                    out_rows.append(row)
                return {
                    "list": out_rows,
                    "total": int(payload.get("total") or 0),
                    "page": int(payload.get("page") or pg),
                    "size": int(payload.get("size") or sz),
                }
            except ValueError:
                raise
            except RuntimeError as e:
                logger.error(f"分页获取仓库列表失败: {e}")
                raise
        try:
            conditions = ["dw.is_active = 1"]
            params: List[Any] = []
            kw = str(keyword).strip() if keyword is not None and str(keyword).strip() else ""
            if kw:
                conditions.append("(dw.name LIKE %s OR IFNULL(wt.name, '') LIKE %s)")
                params.extend([f"%{kw}%", f"%{kw}%"])
            where_sql = " AND ".join(conditions)
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        f"SELECT dw.id AS `仓库id`, dw.name AS `仓库名`, "
                        f"dw.address AS `地址`, "
                        f"dw.province AS `省`, dw.city AS `市`, dw.district AS `区`, "
                        f"dw.longitude AS `经度`, dw.latitude AS `纬度`, "
                        f"dw.contact_name AS `库房联系人`, "
                        f"dw.contact_phone AS `电话`, "
                        f"dw.hazardous_waste_license_qty AS `危废经营许可数量`, "
                        f"dw.monthly_avg_receipt_ton AS `月均收货`, "
                        f"dw.freight_amount AS `运费`, "
                        f"dw.warehouse_type_id AS `仓库类型id`, "
                        f"wt.name AS `类型`, wt.color_config AS `库房类型颜色配置`, "
                        f"dw.color_config AS `仓库颜色配置` "
                        f"FROM dict_warehouses dw "
                        f"LEFT JOIN dict_warehouse_types wt ON dw.warehouse_type_id = wt.id "
                        f"WHERE {where_sql} "
                        "ORDER BY dw.id",
                        tuple(params),
                    )
                    columns = [desc[0] for desc in cur.description]
                    rows = cur.fetchall()
                    out: List[Dict[str, Any]] = []
                    for row in rows:
                        rec = dict(zip(columns, row))
                        hw = rec.get("危废经营许可数量")
                        mar = rec.get("月均收货")
                        fa = rec.get("运费")
                        rec["库房联系人"] = rec.get("库房联系人") or ""
                        rec["电话"] = rec.get("电话") or ""
                        rec["危废经营许可数量"] = (
                            float(hw) if hw is not None else None
                        )
                        rec["月均收货"] = float(mar) if mar is not None else None
                        rec["运费"] = float(fa) if fa is not None else None
                        type_cc = _color_config_from_db(rec.get("库房类型颜色配置"))
                        wh_cc = _color_config_from_db(rec.get("仓库颜色配置"))
                        rec["库房类型颜色配置"] = type_cc
                        rec["仓库颜色配置"] = wh_cc
                        rec["颜色配置"] = type_cc
                        out.append(rec)
                    return out
        except Exception as e:
            logger.error(f"获取仓库列表失败: {e}")
            raise

    # ==================== 接口1b：修改仓库 ====================

    _WH_SITE_PATCH_KEYS = frozenset({"省", "市", "区", "经度", "纬度", "库房类型名"})
    _WH_BUSINESS_PATCH_KEYS = frozenset(
        {"库房联系人", "电话", "危废经营许可数量", "月均收货", "运费"}
    )

    def _business_warehouse_patch_cn_to_en(self, patch: Dict[str, Any]) -> Dict[str, Any]:
        """库房扩展字段：中文请求键 → tl_dict_geo_crud 英文 patch。"""
        out: Dict[str, Any] = {}
        if "库房联系人" in patch:
            v = patch.get("库房联系人")
            out["contact_name"] = None if v is None else str(v).strip()
        if "电话" in patch:
            v = patch.get("电话")
            out["contact_phone"] = None if v is None else str(v).strip()
        if "危废经营许可数量" in patch:
            out["hazardous_waste_license_qty"] = patch.get("危废经营许可数量")
        if "月均收货" in patch:
            out["monthly_avg_receipt_ton"] = patch.get("月均收货")
        if "运费" in patch:
            out["freight_amount"] = patch.get("运费")
        return out

    def _build_site_warehouse_update_patch(self, patch: Dict[str, Any]) -> Dict[str, Any]:
        out: Dict[str, Any] = {}
        if "仓库名" in patch:
            raw = patch["仓库名"]
            if raw is None or str(raw).strip() == "":
                raise ValueError("仓库名不能为空")
            out["name"] = str(raw).strip()
        if "地址" in patch:
            addr = patch["地址"]
            out["address"] = _strip_optional_str(addr) if addr is not None else ""
        if "省" in patch:
            v = patch["省"]
            out["province"] = _strip_nonempty(str(v)) if v is not None else ""
        if "市" in patch:
            v = patch["市"]
            out["city"] = _strip_nonempty(str(v)) if v is not None else ""
        if "区" in patch:
            v = patch["区"]
            out["district"] = _strip_nonempty(str(v)) if v is not None else ""
        if "经度" in patch:
            out["longitude"] = patch["经度"]
        if "纬度" in patch:
            out["latitude"] = patch["纬度"]
        if "库房类型名" in patch:
            tn = patch["库房类型名"]
            out["type"] = "" if tn is None or str(tn).strip() == "" else str(tn).strip()
        elif "仓库类型id" in patch:
            wtid = patch["仓库类型id"]
            if wtid is None:
                out["type"] = ""
            else:
                nm = self._resolve_warehouse_type_name_by_id(int(wtid))
                if not nm:
                    raise ValueError(f"库房类型 id={wtid} 不存在或未启用")
                out["type"] = nm
        if "仓库颜色配置" in patch:
            cc = patch["仓库颜色配置"]
            if cc is None:
                out["color"] = ""
            else:
                hx = _marker_hex_from_wh_color_config(cc)
                if not hx:
                    raise ValueError(
                        "仓库颜色配置须为 JSON，且包含 marker（或 hex）字段为六位十六进制色值，如 #3388FF"
                    )
                out["color"] = hx
        if "is_active" in patch and patch["is_active"] is not None:
            out["status"] = 1 if patch["is_active"] else 0
        return out

    def update_warehouse(self, warehouse_id: int, patch: Dict[str, Any]) -> Dict[str, Any]:
        allowed = (
            {"仓库名", "is_active", "地址", "仓库类型id", "仓库颜色配置"}
            | self._WH_SITE_PATCH_KEYS
            | self._WH_BUSINESS_PATCH_KEYS
        )
        keys = set(patch.keys()) & allowed
        if not keys:
            raise ValueError(
                "至少需要提供一个待修改字段：仓库名、is_active、地址、仓库类型id、仓库颜色配置、"
                "省、市、区、经度、纬度、库房类型名、"
                "库房联系人、电话、危废经营许可数量、月均收货、运费 之一"
            )

        biz_patch = self._business_warehouse_patch_cn_to_en(patch)
        use_site = bool(keys & self._WH_SITE_PATCH_KEYS)
        if use_site:
            try:
                site_patch = self._build_site_warehouse_update_patch(patch)
                merged = {**biz_patch, **site_patch}
                if not merged:
                    raise ValueError("没有有效的修改项")
                _raise_tl_geo_crud_result(sa_wh_update(warehouse_id, merged))
                return {"code": 200, "msg": "仓库信息修改成功"}
            except ValueError:
                raise
            except RuntimeError as e:
                logger.error(f"修改仓库失败(地理落库): {e}")
                raise

        if biz_patch:
            try:
                _raise_tl_geo_crud_result(sa_wh_update(warehouse_id, biz_patch))
            except ValueError:
                raise
            except RuntimeError as e:
                logger.error(f"修改仓库扩展字段失败(地理落库): {e}")
                raise
            if not (keys - self._WH_BUSINESS_PATCH_KEYS):
                return {"code": 200, "msg": "仓库信息修改成功"}

        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT id FROM dict_warehouses WHERE id = %s", (warehouse_id,))
                    if not cur.fetchone():
                        raise ValueError(f"仓库 id={warehouse_id} 不存在")

                    updates: List[str] = []
                    params: List[Any] = []

                    if "仓库名" in patch:
                        name = patch["仓库名"]
                        if name is None or str(name).strip() == "":
                            raise ValueError("仓库名不能为空")
                        name = str(name).strip()
                        cur.execute(
                            "SELECT id FROM dict_warehouses WHERE name = %s AND id <> %s",
                            (name, warehouse_id),
                        )
                        if cur.fetchone():
                            raise ValueError(f"仓库名 '{name}' 已存在")
                        updates.append("name = %s")
                        params.append(name)

                    if "is_active" in patch and patch["is_active"] is not None:
                        updates.append("is_active = %s")
                        params.append(1 if patch["is_active"] else 0)

                    if "地址" in patch:
                        addr = patch["地址"]
                        updates.append("address = %s")
                        params.append(_strip_optional_str(addr) if addr is not None else None)

                    if "仓库类型id" in patch:
                        wtid = patch["仓库类型id"]
                        if wtid is not None:
                            if int(wtid) < 1:
                                raise ValueError("仓库类型id 无效")
                            cur.execute(
                                "SELECT id FROM dict_warehouse_types WHERE id = %s AND is_active = 1",
                                (int(wtid),),
                            )
                            if not cur.fetchone():
                                raise ValueError(f"库房类型 id={wtid} 不存在或未启用")
                            updates.append("warehouse_type_id = %s")
                            params.append(int(wtid))
                        else:
                            updates.append("warehouse_type_id = NULL")

                    if "仓库颜色配置" in patch:
                        cc = patch["仓库颜色配置"]
                        if cc is not None:
                            updates.append("color_config = %s")
                            params.append(_color_config_to_json_str(cc))
                        else:
                            updates.append("color_config = NULL")

                    if not updates:
                        raise ValueError("没有有效的修改项")

                    params.append(warehouse_id)
                    cur.execute(
                        f"UPDATE dict_warehouses SET {', '.join(updates)} WHERE id = %s",
                        tuple(params),
                    )

            return {"code": 200, "msg": "仓库信息修改成功"}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"修改仓库失败: {e}")
            raise

    # ==================== 接口1c：删除仓库（软删除） ====================

    def delete_warehouse(self, warehouse_id: int) -> Dict[str, Any]:
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT id FROM dict_warehouses WHERE id = %s AND is_active = 1",
                        (warehouse_id,),
                    )
                    if not cur.fetchone():
                        raise ValueError(f"仓库 id={warehouse_id} 不存在或已删除")

                    cur.execute(
                        "UPDATE dict_warehouses SET is_active = 0 WHERE id = %s",
                        (warehouse_id,),
                    )
                    cur.execute(
                        "DELETE FROM dict_warehouse_links WHERE from_warehouse_id = %s "
                        "OR to_warehouse_id = %s",
                        (warehouse_id, warehouse_id),
                    )
            return {"code": 200, "msg": "仓库已删除"}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"删除仓库失败: {e}")
            raise

    def purge_warehouse(self, warehouse_id: int) -> Dict[str, Any]:
        """从 dict_warehouses 物理删除；存在外键引用时失败。"""
        if warehouse_id < 1:
            raise ValueError("仓库 id 无效")
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT id FROM dict_warehouses WHERE id = %s",
                        (warehouse_id,),
                    )
                    if not cur.fetchone():
                        raise ValueError(f"仓库 id={warehouse_id} 不存在")
                    try:
                        cur.execute(
                            "DELETE FROM dict_warehouses WHERE id = %s",
                            (warehouse_id,),
                        )
                    except PyMySQLIntegrityError as e:
                        logger.warning("硬删除仓库触发外键约束: %s", e)
                        raise ValueError(
                            "该仓库仍被运费或其它业务数据引用，无法物理删除；"
                            "请先解除关联或使用软删除接口"
                        ) from e
            return {"code": 200, "msg": "仓库已永久删除"}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"硬删除仓库失败: {e}")
            raise

    # ==================== 库房单向关联（有向图边）====================

    def bind_warehouse_link(
        self,
        from_wh_id: int,
        to_wh_id: int,
        tier_price_spread: Any = None,
    ) -> Dict[str, Any]:
        """新增出边：源库房 -> 对标库房。"""
        res = _raise_tl_geo_crud_result(
            sa_wh_link_bind(from_wh_id, to_wh_id, tier_price_spread=tier_price_spread)
        )
        data = res.get("data") or {}
        return {
            "code": 200,
            "msg": str(res.get("msg") or "绑定成功"),
            "data": {
                "关联id": data.get("linkId"),
                "源库房id": data.get("fromWarehouseId"),
                "对标库房id": data.get("toWarehouseId"),
                "创建时间": data.get("createTime"),
                "阶梯价差": data.get("tierPriceSpread"),
            },
        }

    def update_warehouse_link_tier_price_spread(
        self,
        from_wh_id: int,
        to_wh_id: int,
        tier_price_spread: Any,
    ) -> Dict[str, Any]:
        """修改源库房→对标库房边上的阶梯价差（传 null 清空）。"""
        res = _raise_tl_geo_crud_result(
            sa_wh_link_update_tier(from_wh_id, to_wh_id, tier_price_spread)
        )
        data = res.get("data") or {}
        return {
            "code": 200,
            "msg": str(res.get("msg") or "修改成功"),
            "data": {
                "源库房id": data.get("fromWarehouseId"),
                "对标库房id": data.get("toWarehouseId"),
                "阶梯价差": data.get("tierPriceSpread"),
            },
        }

    def unbind_warehouse_link(self, from_wh_id: int, to_wh_id: int) -> Dict[str, Any]:
        """删除出边。"""
        _raise_tl_geo_crud_result(sa_wh_link_unbind(from_wh_id, to_wh_id))
        return {"code": 200, "msg": "解绑成功"}

    def get_warehouse_links_outbound(
        self,
        warehouse_id: int,
        page: int = 1,
        size: int = 50,
    ) -> Dict[str, Any]:
        """分页列出某库房指向哪些库房（出边）。"""
        res = _raise_tl_geo_crud_result(
            sa_wh_links_outbound(warehouse_id, page=page, size=size)
        )
        payload = res.get("data") or {}
        raw_list = payload.get("list") or []
        all_types: List[str] = []
        for it in raw_list:
            tgt = (it.get("target") or {}) if isinstance(it.get("target"), dict) else {}
            tn = str(tgt.get("type") or "").strip()
            if tn:
                all_types.append(tn)
        tid_by_name = self._warehouse_type_ids_by_names(all_types)
        tcol = self._batch_warehouse_type_colors(
            [tid_by_name[t] for t in all_types if t in tid_by_name]
        )
        out_rows: List[Dict[str, Any]] = []
        for it in raw_list:
            tgt = it.get("target") or {}
            tname = str(tgt.get("type") or "").strip()
            wt_id = tid_by_name.get(tname) if tname else None
            row = self._site_wh_item_to_tl_row(tgt, tid_by_name)
            if wt_id is not None:
                t_cc = tcol.get(int(wt_id))
                row["库房类型颜色配置"] = t_cc
                row["颜色配置"] = t_cc
            out_rows.append(
                {
                    "关联id": it.get("linkId"),
                    "源库房id": it.get("fromWarehouseId"),
                    "对标库房id": it.get("toWarehouseId"),
                    "创建时间": it.get("createTime"),
                    "距离千米": it.get("distanceKm"),
                    "阶梯价差": it.get("tierPriceSpread"),
                    "对标库房": row,
                }
            )
        return {
            "code": 200,
            "msg": "查询成功",
            "list": out_rows,
            "total": int(payload.get("total") or 0),
            "page": int(payload.get("page") or page),
            "size": int(payload.get("size") or size),
        }

    def get_warehouse_links_inbound(
        self,
        warehouse_id: int,
        page: int = 1,
        size: int = 50,
    ) -> Dict[str, Any]:
        """分页列出哪些库房指向该库房（入边）。"""
        res = _raise_tl_geo_crud_result(
            sa_wh_links_inbound(warehouse_id, page=page, size=size)
        )
        payload = res.get("data") or {}
        raw_list = payload.get("list") or []
        all_types: List[str] = []
        for it in raw_list:
            src = (it.get("source") or {}) if isinstance(it.get("source"), dict) else {}
            tn = str(src.get("type") or "").strip()
            if tn:
                all_types.append(tn)
        tid_by_name = self._warehouse_type_ids_by_names(all_types)
        tcol = self._batch_warehouse_type_colors(
            [tid_by_name[t] for t in all_types if t in tid_by_name]
        )
        out_rows: List[Dict[str, Any]] = []
        for it in raw_list:
            src = it.get("source") or {}
            tname = str(src.get("type") or "").strip()
            wt_id = tid_by_name.get(tname) if tname else None
            row = self._site_wh_item_to_tl_row(src, tid_by_name)
            if wt_id is not None:
                t_cc = tcol.get(int(wt_id))
                row["库房类型颜色配置"] = t_cc
                row["颜色配置"] = t_cc
            out_rows.append(
                {
                    "关联id": it.get("linkId"),
                    "源库房id": it.get("fromWarehouseId"),
                    "对标库房id": it.get("toWarehouseId"),
                    "创建时间": it.get("createTime"),
                    "距离千米": it.get("distanceKm"),
                    "阶梯价差": it.get("tierPriceSpread"),
                    "源库房": row,
                }
            )
        return {
            "code": 200,
            "msg": "查询成功",
            "list": out_rows,
            "total": int(payload.get("total") or 0),
            "page": int(payload.get("page") or page),
            "size": int(payload.get("size") or size),
        }

    def get_warehouse_links_list(
        self,
        page: int = 1,
        size: int = 50,
        warehouse_id: Optional[int] = None,
        from_warehouse_id: Optional[int] = None,
        to_warehouse_id: Optional[int] = None,
        keyword: Optional[str] = None,
        has_tier_price_spread: Optional[bool] = None,
    ) -> Dict[str, Any]:
        """库房关联总列表（每条含源、对标库房摘要及距离、阶梯价差）；可选筛选。"""
        res = _raise_tl_geo_crud_result(
            sa_wh_links_list_all(
                page=page,
                size=size,
                warehouse_id=warehouse_id,
                from_warehouse_id=from_warehouse_id,
                to_warehouse_id=to_warehouse_id,
                keyword=keyword,
                has_tier_price_spread=has_tier_price_spread,
            )
        )
        payload = res.get("data") or {}
        raw_list = payload.get("list") or []
        all_types: List[str] = []
        for it in raw_list:
            for side in ("source", "target"):
                w = (it.get(side) or {}) if isinstance(it.get(side), dict) else {}
                tn = str(w.get("type") or "").strip()
                if tn:
                    all_types.append(tn)
        tid_by_name = self._warehouse_type_ids_by_names(all_types)
        tcol = self._batch_warehouse_type_colors(
            [tid_by_name[t] for t in all_types if t in tid_by_name]
        )
        out_rows: List[Dict[str, Any]] = []
        for it in raw_list:
            src = it.get("source") or {}
            tgt = it.get("target") or {}
            stname = str(src.get("type") or "").strip()
            ttname = str(tgt.get("type") or "").strip()
            s_wt_id = tid_by_name.get(stname) if stname else None
            t_wt_id = tid_by_name.get(ttname) if ttname else None
            src_tl = self._site_wh_item_to_tl_row(src, tid_by_name)
            tgt_tl = self._site_wh_item_to_tl_row(tgt, tid_by_name)
            if s_wt_id is not None:
                c = tcol.get(int(s_wt_id))
                src_tl["库房类型颜色配置"] = c
                src_tl["颜色配置"] = c
            if t_wt_id is not None:
                c = tcol.get(int(t_wt_id))
                tgt_tl["库房类型颜色配置"] = c
                tgt_tl["颜色配置"] = c
            out_rows.append(
                {
                    "关联id": it.get("linkId"),
                    "源库房id": it.get("fromWarehouseId"),
                    "对标库房id": it.get("toWarehouseId"),
                    "创建时间": it.get("createTime"),
                    "距离千米": it.get("distanceKm"),
                    "阶梯价差": it.get("tierPriceSpread"),
                    "源库房": src_tl,
                    "对标库房": tgt_tl,
                }
            )
        return {
            "code": 200,
            "msg": "查询成功",
            "list": out_rows,
            "total": int(payload.get("total") or 0),
            "page": int(payload.get("page") or page),
            "size": int(payload.get("size") or size),
        }

    def get_tier_price_spread_list(
        self,
        page: int = 1,
        size: int = 50,
        warehouse_id: Optional[int] = None,
        from_warehouse_id: Optional[int] = None,
        to_warehouse_id: Optional[int] = None,
        keyword: Optional[str] = None,
        has_tier_price_spread: Optional[bool] = None,
    ) -> Dict[str, Any]:
        """阶梯差价列表：结构与 get_warehouse_links_list 相同（含距离千米、阶梯价差）。"""
        return self.get_warehouse_links_list(
            page=page,
            size=size,
            warehouse_id=warehouse_id,
            from_warehouse_id=from_warehouse_id,
            to_warehouse_id=to_warehouse_id,
            keyword=keyword,
            has_tier_price_spread=has_tier_price_spread,
        )

    def replace_warehouse_links_outbound(
        self,
        from_wh_id: int,
        to_wh_ids: List[int],
    ) -> Dict[str, Any]:
        """将源库房出边整体替换为目标 id 列表（改）。"""
        res = _raise_tl_geo_crud_result(
            sa_wh_links_replace_outbound(from_wh_id, to_wh_ids)
        )
        data = res.get("data") or {}
        return {
            "code": 200,
            "msg": str(res.get("msg") or "替换成功"),
            "data": {
                "源库房id": data.get("fromWarehouseId"),
                "目标库房id列表": data.get("toWarehouseIds") or [],
            },
        }

    def batch_bind_warehouse_links(
        self,
        from_wh_id: int,
        to_wh_ids: List[int],
    ) -> Dict[str, Any]:
        """同一源库房一次性绑定多条出边（已存在的边跳过）。"""
        res = _raise_tl_geo_crud_result(
            sa_wh_links_batch_bind(from_wh_id, to_wh_ids)
        )
        data = res.get("data") or {}
        return {
            "code": 200,
            "msg": str(res.get("msg") or "绑定完成"),
            "data": {
                "源库房id": data.get("fromWarehouseId"),
                "新增边数": int(data.get("inserted") or 0),
                "跳过已存在边数": int(data.get("skippedDuplicate") or 0),
            },
        }

    def batch_unbind_warehouse_links(
        self,
        from_wh_id: int,
        to_wh_ids: List[int],
    ) -> Dict[str, Any]:
        """同一源库房一次性解绑多条出边。"""
        res = _raise_tl_geo_crud_result(
            sa_wh_links_batch_unbind(from_wh_id, to_wh_ids)
        )
        data = res.get("data") or {}
        return {
            "code": 200,
            "msg": str(res.get("msg") or "解绑完成"),
            "data": {
                "源库房id": data.get("fromWarehouseId"),
                "删除边数": int(data.get("deleted") or 0),
            },
        }

    # ==================== 库房类型维护（类型与颜色一对一）====================

    def get_warehouse_types(
        self,
        keyword: Optional[str] = None,
        include_inactive: bool = False,
    ) -> List[Dict[str, Any]]:
        try:
            conditions: List[str] = []
            params: List[Any] = []
            if not include_inactive:
                conditions.append("is_active = 1")
            if keyword is not None and str(keyword).strip():
                conditions.append("name LIKE %s")
                params.append(f"%{str(keyword).strip()}%")
            where_sql = " AND ".join(conditions) if conditions else "1=1"
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        f"SELECT id AS `类型id`, name AS `类型名`, color_config AS `颜色配置`, "
                        f"is_active AS `is_active` "
                        f"FROM dict_warehouse_types WHERE {where_sql} "
                        "ORDER BY id",
                        tuple(params),
                    )
                    columns = [d[0] for d in cur.description]
                    out: List[Dict[str, Any]] = []
                    for row in cur.fetchall():
                        rec = dict(zip(columns, row))
                        rec["颜色配置"] = _color_config_from_db(rec.get("颜色配置"))
                        out.append(rec)
                    return out
        except Exception as e:
            logger.error(f"获取库房类型列表失败: {e}")
            raise

    def add_warehouse_type(
        self, name: str, color_config: Optional[Any] = None
    ) -> Dict[str, Any]:
        n = str(name).strip()
        if not n:
            raise ValueError("类型名不能为空")
        if len(n) > 50:
            raise ValueError("类型名长度不能超过 50 字符")
        cc_json = _color_config_to_json_str(color_config) if color_config is not None else None
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT id, is_active FROM dict_warehouse_types WHERE name = %s",
                        (n,),
                    )
                    row = cur.fetchone()
                    if row:
                        tid, act = int(row[0]), row[1]
                        if act == 1:
                            return {
                                "code": 200,
                                "msg": "类型已存在",
                                "类型id": tid,
                                "新建": False,
                            }
                        if cc_json is None:
                            cur.execute(
                                "UPDATE dict_warehouse_types SET is_active = 1 WHERE id = %s",
                                (tid,),
                            )
                        else:
                            cur.execute(
                                "UPDATE dict_warehouse_types SET is_active = 1, "
                                "color_config = CAST(%s AS JSON) WHERE id = %s",
                                (cc_json, tid),
                            )
                        return {
                            "code": 200,
                            "msg": "类型已恢复启用",
                            "类型id": tid,
                            "新建": False,
                        }
                    if cc_json is None:
                        cur.execute(
                            "INSERT INTO dict_warehouse_types (name, color_config, is_active) "
                            "VALUES (%s, NULL, 1)",
                            (n,),
                        )
                    else:
                        cur.execute(
                            "INSERT INTO dict_warehouse_types (name, color_config, is_active) "
                            "VALUES (%s, CAST(%s AS JSON), 1)",
                            (n, cc_json),
                        )
                    return {
                        "code": 200,
                        "msg": "库房类型新建成功",
                        "类型id": cur.lastrowid,
                        "新建": True,
                    }
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"新建库房类型失败: {e}")
            raise

    def update_warehouse_type(self, type_id: int, patch: Dict[str, Any]) -> Dict[str, Any]:
        allowed = {"类型名", "颜色配置", "is_active"}
        if not (set(patch.keys()) & allowed):
            raise ValueError("至少需要提供一个待修改字段：类型名、颜色配置、is_active 之一")
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT id FROM dict_warehouse_types WHERE id = %s",
                        (type_id,),
                    )
                    if not cur.fetchone():
                        raise ValueError(f"库房类型 id={type_id} 不存在")

                    updates: List[str] = []
                    params: List[Any] = []

                    if "类型名" in patch:
                        nn = patch["类型名"]
                        if nn is None or str(nn).strip() == "":
                            raise ValueError("类型名不能为空")
                        nn = str(nn).strip()
                        cur.execute(
                            "SELECT id FROM dict_warehouse_types WHERE name = %s AND id <> %s",
                            (nn, type_id),
                        )
                        if cur.fetchone():
                            raise ValueError(f"类型名「{nn}」已存在")
                        updates.append("name = %s")
                        params.append(nn)

                    if "is_active" in patch and patch["is_active"] is not None:
                        updates.append("is_active = %s")
                        params.append(1 if patch["is_active"] else 0)

                    if "颜色配置" in patch:
                        cc = patch["颜色配置"]
                        if cc is None:
                            updates.append("color_config = NULL")
                        else:
                            updates.append("color_config = CAST(%s AS JSON)")
                            params.append(_color_config_to_json_str(cc))

                    if not updates:
                        raise ValueError("没有有效的修改项")

                    params.append(type_id)
                    cur.execute(
                        f"UPDATE dict_warehouse_types SET {', '.join(updates)} WHERE id = %s",
                        tuple(params),
                    )

            return {"code": 200, "msg": "库房类型已更新"}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"更新库房类型失败: {e}")
            raise

    def delete_warehouse_type(self, type_id: int) -> Dict[str, Any]:
        """软删除类型：相关仓库的 warehouse_type_id 置空（颜色随类型失效）。"""
        if type_id < 1:
            raise ValueError("类型id 无效")
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT id FROM dict_warehouse_types WHERE id = %s AND is_active = 1",
                        (type_id,),
                    )
                    if not cur.fetchone():
                        raise ValueError(f"库房类型 id={type_id} 不存在或已停用")
                    cur.execute(
                        "UPDATE dict_warehouses SET warehouse_type_id = NULL "
                        "WHERE warehouse_type_id = %s",
                        (type_id,),
                    )
                    cur.execute(
                        "UPDATE dict_warehouse_types SET is_active = 0 WHERE id = %s",
                        (type_id,),
                    )
            return {"code": 200, "msg": "库房类型已删除"}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"删除库房类型失败: {e}")
            raise

    # ==================== 接口2：获取冶炼厂列表 ====================

    _SM_SITE_PATCH_KEYS = frozenset({"省", "市", "区", "经度", "纬度"})

    def _site_smelter_item_to_tl_row(self, item: Dict[str, Any]) -> Dict[str, Any]:
        """分页冶炼厂条目 → TL 字段（无颜色）。"""
        return {
            "冶炼厂id": int(item["id"]),
            "冶炼厂": item["name"],
            "地址": item.get("address") or "",
            "省": item.get("province") or "",
            "市": item.get("city") or "",
            "区": item.get("district") or "",
            "经度": item.get("longitude"),
            "纬度": item.get("latitude"),
            "循融宝发货": bool(item.get("循融宝发货")),
        }

    def get_smelters(
        self,
        keyword: Optional[str] = None,
        page: Optional[int] = None,
        size: Optional[int] = None,
        province: Optional[str] = None,
        city: Optional[str] = None,
        district: Optional[str] = None,
        status: Optional[int] = None,
    ) -> Any:
        if page is not None:
            try:
                pg = max(1, int(page))
                sz = min(200, max(1, int(size or 20)))
                kw = (
                    str(keyword).strip()
                    if keyword is not None and str(keyword).strip()
                    else None
                )
                eff_status = status if status is not None else 1
                res = _raise_tl_geo_crud_result(
                    sa_smelter_list(pg, sz, kw, province, city, district, eff_status)
                )
                payload = res["data"] or {}
                items_raw = payload.get("list") or []
                out_rows = [self._site_smelter_item_to_tl_row(x) for x in items_raw]
                return {
                    "list": out_rows,
                    "total": int(payload.get("total") or 0),
                    "page": int(payload.get("page") or pg),
                    "size": int(payload.get("size") or sz),
                }
            except ValueError:
                raise
            except RuntimeError as e:
                logger.error(f"分页获取冶炼厂列表失败: {e}")
                raise
        try:
            conditions = ["is_active = 1"]
            params: List[Any] = []
            if keyword is not None and str(keyword).strip():
                conditions.append("name LIKE %s")
                params.append(f"%{str(keyword).strip()}%")
            where_sql = " AND ".join(conditions)
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        f"SELECT id AS `冶炼厂id`, name AS `冶炼厂`, address AS `地址`, "
                        f"province AS `省`, city AS `市`, district AS `区`, "
                        f"longitude AS `经度`, latitude AS `纬度`, "
                        f"COALESCE(use_xunrongbao, 0) AS `循融宝发货` "
                        f"FROM dict_factories WHERE {where_sql} "
                        "ORDER BY id",
                        tuple(params),
                    )
                    columns = [desc[0] for desc in cur.description]
                    rows = cur.fetchall()
                    out: List[Dict[str, Any]] = []
                    for row in rows:
                        rec = dict(zip(columns, row))
                        out.append(rec)
                    return out
        except Exception as e:
            logger.error(f"获取冶炼厂列表失败: {e}")
            raise

    def get_smelter(self, smelter_id: int) -> Dict[str, Any]:
        """单个冶炼厂详情（含循融宝发货、地址与坐标）；启用状态见 is_active。"""
        res = _raise_tl_geo_crud_result(sa_smelter_get(smelter_id))
        data = res.get("data") or {}
        row = self._site_smelter_item_to_tl_row(data)
        row["is_active"] = bool(int(data.get("status", 1)))
        return row

    def list_smelter_xunrongbao(
        self,
        include_inactive: bool = False,
    ) -> Dict[str, Any]:
        """查询全部冶炼厂的循融宝发货状态及当前系统加价（元/吨）。"""
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    if include_inactive:
                        cur.execute(
                            "SELECT id, name, COALESCE(use_xunrongbao, 0), is_active "
                            "FROM dict_factories ORDER BY id"
                        )
                    else:
                        cur.execute(
                            "SELECT id, name, COALESCE(use_xunrongbao, 0), is_active "
                            "FROM dict_factories WHERE is_active = 1 ORDER BY id"
                        )
                    rows = cur.fetchall()
            return {
                "加价元每吨": XUNRONGBAO_SHIPPING_PREMIUM_PER_TON,
                "list": [
                    {
                        "冶炼厂id": int(r[0]),
                        "冶炼厂": r[1],
                        "循融宝发货": bool(int(r[2])),
                        "is_active": bool(int(r[3])),
                    }
                    for r in rows
                ],
            }
        except Exception as e:
            logger.error(f"列出冶炼厂循融宝状态失败: {e}")
            raise

    def set_smelter_xunrongbao(self, smelter_id: int, enabled: bool) -> Dict[str, Any]:
        """仅修改循融宝发货开关（等价于 update_smelter 仅传 循融宝发货）。"""
        return self.update_smelter(smelter_id, {"循融宝发货": enabled})

    def clear_smelter_xunrongbao(self, smelter_id: int) -> Dict[str, Any]:
        """关闭循融宝发货（不删除冶炼厂，不软删）。"""
        return self.update_smelter(smelter_id, {"循融宝发货": False})

    def get_missing_geo_info(self) -> Dict[str, Any]:
        """返回启用中且缺少经度或纬度的仓库、冶炼厂，供前端集中补全地址坐标。"""
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT dw.id AS `仓库id`, dw.name AS `仓库名`, "
                        "dw.address AS `地址`, dw.province AS `省`, dw.city AS `市`, dw.district AS `区`, "
                        "dw.longitude AS `经度`, dw.latitude AS `纬度`, "
                        "dw.is_active AS `is_active`, wt.name AS `类型` "
                        "FROM dict_warehouses dw "
                        "LEFT JOIN dict_warehouse_types wt ON dw.warehouse_type_id = wt.id "
                        "WHERE (dw.longitude IS NULL OR dw.latitude IS NULL) "
                        "AND dw.is_active = 1 "
                        "ORDER BY dw.id"
                    )
                    wh_columns = [desc[0] for desc in cur.description]
                    warehouses: List[Dict[str, Any]] = []
                    for row in cur.fetchall():
                        rec = dict(zip(wh_columns, row))
                        rec["缺失字段"] = [
                            field
                            for field in ("经度", "纬度")
                            if rec.get(field) is None
                        ]
                        warehouses.append(rec)

                    cur.execute(
                        "SELECT id AS `冶炼厂id`, name AS `冶炼厂`, address AS `地址`, "
                        "province AS `省`, city AS `市`, district AS `区`, "
                        "longitude AS `经度`, latitude AS `纬度`, is_active AS `is_active` "
                        "FROM dict_factories "
                        "WHERE (longitude IS NULL OR latitude IS NULL) "
                        "AND is_active = 1 "
                        "ORDER BY id"
                    )
                    sm_columns = [desc[0] for desc in cur.description]
                    smelters: List[Dict[str, Any]] = []
                    for row in cur.fetchall():
                        rec = dict(zip(sm_columns, row))
                        rec["缺失字段"] = [
                            field
                            for field in ("经度", "纬度")
                            if rec.get(field) is None
                        ]
                        smelters.append(rec)

            return {
                "warehouses": warehouses,
                "smelters": smelters,
                "summary": {
                    "warehouses": len(warehouses),
                    "smelters": len(smelters),
                    "total": len(warehouses) + len(smelters),
                },
            }
        except Exception as e:
            logger.error(f"获取经纬度缺失列表失败: {e}")
            raise

    # ==================== 接口2b：修改冶炼厂 ====================

    def _build_site_smelter_update_patch(self, patch: Dict[str, Any]) -> Dict[str, Any]:
        """冶炼厂省市区变更且未手传经纬度时，由 tl_dict_geo_crud 内 maybe_geocode 重算坐标。"""
        out: Dict[str, Any] = {}
        if "循融宝发货" in patch and patch["循融宝发货"] is not None:
            out["use_xunrongbao"] = bool(patch["循融宝发货"])
        if "冶炼厂名" in patch:
            raw = patch["冶炼厂名"]
            if raw is None or str(raw).strip() == "":
                raise ValueError("冶炼厂名不能为空")
            out["name"] = str(raw).strip()
        if "地址" in patch:
            addr = patch["地址"]
            out["address"] = _strip_optional_str(addr) if addr is not None else ""
        if "省" in patch:
            v = patch["省"]
            out["province"] = _strip_nonempty(str(v)) if v is not None else ""
        if "市" in patch:
            v = patch["市"]
            out["city"] = _strip_nonempty(str(v)) if v is not None else ""
        if "区" in patch:
            v = patch["区"]
            out["district"] = _strip_nonempty(str(v)) if v is not None else ""
        if "经度" in patch:
            out["longitude"] = patch["经度"]
        if "纬度" in patch:
            out["latitude"] = patch["纬度"]
        if "is_active" in patch and patch["is_active"] is not None:
            out["status"] = 1 if patch["is_active"] else 0
        return out

    def update_smelter(self, smelter_id: int, patch: Dict[str, Any]) -> Dict[str, Any]:
        allowed = {"冶炼厂名", "is_active", "地址", "循融宝发货"} | self._SM_SITE_PATCH_KEYS
        keys = set(patch.keys()) & allowed
        if not keys:
            raise ValueError(
                "至少需要提供一个待修改字段：冶炼厂名、is_active、地址、循融宝发货、省、市、区、经度、纬度 之一"
            )

        use_site = bool(keys & self._SM_SITE_PATCH_KEYS)
        if use_site:
            try:
                site_patch = self._build_site_smelter_update_patch(patch)
                if not site_patch:
                    raise ValueError("没有有效的修改项")
                _raise_tl_geo_crud_result(sa_smelter_update(smelter_id, site_patch))
                return {"code": 200, "msg": "冶炼厂信息修改成功"}
            except ValueError:
                raise
            except RuntimeError as e:
                logger.error(f"修改冶炼厂失败(地理落库): {e}")
                raise

        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT id FROM dict_factories WHERE id = %s", (smelter_id,))
                    if not cur.fetchone():
                        raise ValueError(f"冶炼厂 id={smelter_id} 不存在")

                    updates: List[str] = []
                    params: List[Any] = []

                    if "冶炼厂名" in patch:
                        name = patch["冶炼厂名"]
                        if name is None or str(name).strip() == "":
                            raise ValueError("冶炼厂名不能为空")
                        name = str(name).strip()
                        cur.execute(
                            "SELECT id FROM dict_factories WHERE name = %s AND id <> %s",
                            (name, smelter_id),
                        )
                        if cur.fetchone():
                            raise ValueError(f"冶炼厂名 '{name}' 已存在")
                        updates.append("name = %s")
                        params.append(name)

                    if "is_active" in patch and patch["is_active"] is not None:
                        updates.append("is_active = %s")
                        params.append(1 if patch["is_active"] else 0)

                    if "地址" in patch:
                        addr = patch["地址"]
                        updates.append("address = %s")
                        params.append(_strip_optional_str(addr) if addr is not None else None)

                    if "循融宝发货" in patch and patch["循融宝发货"] is not None:
                        updates.append("use_xunrongbao = %s")
                        params.append(1 if patch["循融宝发货"] else 0)

                    if not updates:
                        raise ValueError("没有有效的修改项")

                    params.append(smelter_id)
                    cur.execute(
                        f"UPDATE dict_factories SET {', '.join(updates)} WHERE id = %s",
                        tuple(params),
                    )

            return {"code": 200, "msg": "冶炼厂信息修改成功"}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"修改冶炼厂失败: {e}")
            raise

    def batch_set_smelters_xunrongbao(self, items: List[Dict[str, Any]]) -> Dict[str, Any]:
        """批量设置冶炼厂是否循融宝发货；每项含 冶炼厂id、循融宝发货。"""
        if not items:
            raise ValueError("列表不能为空")
        try:
            with get_conn() as conn:
                conn.autocommit(False)
                try:
                    with conn.cursor() as cur:
                        for it in items:
                            fid = int(it["冶炼厂id"])
                            flag = bool(it["循融宝发货"])
                            cur.execute(
                                "UPDATE dict_factories SET use_xunrongbao = %s WHERE id = %s",
                                (1 if flag else 0, fid),
                            )
                            if cur.rowcount != 1:
                                raise ValueError(f"冶炼厂 id={fid} 不存在")
                    conn.commit()
                except Exception:
                    conn.rollback()
                    raise
            return {
                "code": 200,
                "msg": f"已更新 {len(items)} 个冶炼厂的循融宝发货配置",
            }
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"批量设置循融宝发货失败: {e}")
            raise

    # ==================== 接口2c：删除冶炼厂（软删除） ====================

    def delete_smelter(self, smelter_id: int) -> Dict[str, Any]:
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT id FROM dict_factories WHERE id = %s AND is_active = 1",
                        (smelter_id,),
                    )
                    if not cur.fetchone():
                        raise ValueError(f"冶炼厂 id={smelter_id} 不存在或已删除")

                    cur.execute(
                        "UPDATE dict_factories SET is_active = 0 WHERE id = %s",
                        (smelter_id,),
                    )
            return {"code": 200, "msg": "冶炼厂已删除"}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"删除冶炼厂失败: {e}")
            raise

    def purge_smelter(self, smelter_id: int, *, cascade: bool = True) -> Dict[str, Any]:
        """从 dict_factories 物理删除。

        cascade=True（默认）：同一事务内删除该厂关联的需求/报价/运费等后删厂。
        cascade=False：仅当子表无任何引用时才删厂；否则 409（严格模式，供脚本校验用）。
        """
        if smelter_id < 1:
            raise ValueError("冶炼厂 id 无效")

        _child_count_sql = """
            SELECT
              (SELECT COUNT(*) FROM factory_demand_items fdi
                 INNER JOIN factory_demands fd ON fd.id = fdi.demand_id
                 WHERE fd.factory_id = %s),
              (SELECT COUNT(*) FROM factory_demands WHERE factory_id = %s),
              (SELECT COUNT(*) FROM quote_details WHERE factory_id = %s),
              (SELECT COUNT(*) FROM quote_table_metadata WHERE factory_id = %s),
              (SELECT COUNT(*) FROM freight_rates WHERE factory_id = %s),
              (SELECT COUNT(*) FROM factory_tax_rates WHERE factory_id = %s)
        """

        try:
            if cascade:
                with get_conn() as conn:
                    conn.autocommit(False)
                    try:
                        with conn.cursor() as cur:
                            cur.execute(
                                "SELECT id, name FROM dict_factories WHERE id = %s",
                                (smelter_id,),
                            )
                            row = cur.fetchone()
                            if not row:
                                raise ValueError(f"冶炼厂 id={smelter_id} 不存在")
                            factory_name = str(row[1])
                            cur.execute(_child_count_sql, (smelter_id,) * 6)
                            c_row = cur.fetchone()
                            n_di, n_dm, n_qd, n_qm, n_fr, n_tr = (
                                int(c_row[0] or 0),
                                int(c_row[1] or 0),
                                int(c_row[2] or 0),
                                int(c_row[3] or 0),
                                int(c_row[4] or 0),
                                int(c_row[5] or 0),
                            )

                            cur.execute(
                                """
                                DELETE fdi FROM factory_demand_items fdi
                                INNER JOIN factory_demands fd ON fd.id = fdi.demand_id
                                WHERE fd.factory_id = %s
                                """,
                                (smelter_id,),
                            )
                            cur.execute(
                                "DELETE FROM factory_demands WHERE factory_id = %s",
                                (smelter_id,),
                            )
                            cur.execute(
                                "DELETE FROM quote_details WHERE factory_id = %s",
                                (smelter_id,),
                            )
                            cur.execute(
                                "DELETE FROM quote_table_metadata WHERE factory_id = %s",
                                (smelter_id,),
                            )
                            cur.execute(
                                "DELETE FROM freight_rates WHERE factory_id = %s",
                                (smelter_id,),
                            )
                            cur.execute(
                                "DELETE FROM dict_factories WHERE id = %s",
                                (smelter_id,),
                            )
                            if cur.rowcount == 0:
                                raise ValueError(f"冶炼厂 id={smelter_id} 删除失败")
                        conn.commit()
                    except Exception:
                        conn.rollback()
                        raise

                log_finance_event(
                    "冶炼厂硬删除(级联) | id=%s name=%s | 删 demand_items=%s demands=%s "
                    "quote_details=%s quote_metadata=%s freight_rates=%s tax_rates(删厂前)=%s",
                    smelter_id,
                    factory_name,
                    n_di,
                    n_dm,
                    n_qd,
                    n_qm,
                    n_fr,
                    n_tr,
                )
                return {
                    "code": 200,
                    "msg": (
                        f"已永久删除冶炼厂 id={smelter_id}，并清除关联数据（见 deleted_counts）"
                    ),
                    "cascade": True,
                    "deleted_counts": {
                        "factory_demand_items": n_di,
                        "factory_demands": n_dm,
                        "quote_details": n_qd,
                        "quote_table_metadata": n_qm,
                        "freight_rates": n_fr,
                        "factory_tax_rates": n_tr,
                    },
                }

            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT id FROM dict_factories WHERE id = %s",
                        (smelter_id,),
                    )
                    if not cur.fetchone():
                        raise ValueError(f"冶炼厂 id={smelter_id} 不存在")
                    cur.execute(_child_count_sql, (smelter_id,) * 6)
                    c_row = cur.fetchone()
                    n_di, n_dm, n_qd, n_qm, n_fr, n_tr = (
                        int(c_row[0] or 0),
                        int(c_row[1] or 0),
                        int(c_row[2] or 0),
                        int(c_row[3] or 0),
                        int(c_row[4] or 0),
                        int(c_row[5] or 0),
                    )
                    total_children = n_di + n_dm + n_qd + n_qm + n_fr + n_tr
                    if total_children > 0:
                        raise ValueError(
                            "已指定 cascade=false（仅当无子表引用时才删厂）。"
                            f"冶炼厂 id={smelter_id} 仍存在关联："
                            f"demand_items={n_di}, demands={n_dm}, quote_details={n_qd}, "
                            f"quote_metadata={n_qm}, freight_rates={n_fr}, tax_rates={n_tr}。"
                            "默认删除会级联清空上述数据，请勿传 cascade=false；或先手工清理子表。"
                        )
                    cur.execute(
                        "DELETE FROM dict_factories WHERE id = %s",
                        (smelter_id,),
                    )
            return {
                "code": 200,
                "msg": "冶炼厂已永久删除（cascade=false 且无关联子表）",
                "cascade": False,
            }
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"硬删除冶炼厂失败: {e}")
            raise

    def batch_delete_warehouses(self, warehouse_ids: List[int]) -> Dict[str, Any]:
        """批量软删除仓库（将 is_active 置 0）。"""
        if not warehouse_ids:
            raise ValueError("仓库id列表不能为空")
        ids = list(dict.fromkeys(int(x) for x in warehouse_ids))
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    ph = ",".join(["%s"] * len(ids))
                    cur.execute(
                        f"UPDATE dict_warehouses SET is_active = 0 "
                        f"WHERE id IN ({ph}) AND is_active = 1",
                        tuple(ids),
                    )
                    n = cur.rowcount
                    cur.execute(
                        f"DELETE FROM dict_warehouse_links WHERE from_warehouse_id IN ({ph}) "
                        f"OR to_warehouse_id IN ({ph})",
                        tuple(ids + ids),
                    )
            return {"code": 200, "msg": f"已批量停用 {n} 个仓库", "更新行数": n}
        except Exception as e:
            logger.error(f"批量停用仓库失败: {e}")
            raise

    def batch_delete_smelters(self, smelter_ids: List[int]) -> Dict[str, Any]:
        """批量软删除冶炼厂。"""
        if not smelter_ids:
            raise ValueError("冶炼厂id列表不能为空")
        ids = list(dict.fromkeys(int(x) for x in smelter_ids))
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    ph = ",".join(["%s"] * len(ids))
                    cur.execute(
                        f"UPDATE dict_factories SET is_active = 0 "
                        f"WHERE id IN ({ph}) AND is_active = 1",
                        tuple(ids),
                    )
                    n = cur.rowcount
            return {"code": 200, "msg": f"已批量停用 {n} 个冶炼厂", "更新行数": n}
        except Exception as e:
            logger.error(f"批量停用冶炼厂失败: {e}")
            raise

    # ==================== 接口3：获取品类列表 ====================

    def get_categories(self) -> List[Dict[str, Any]]:
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT category_id AS `品类id`, "
                        "GROUP_CONCAT(name ORDER BY row_id SEPARATOR '、') AS `品类名` "
                        "FROM dict_categories "
                        "WHERE is_active = 1 "
                        "GROUP BY category_id "
                        "ORDER BY category_id"
                    )
                    columns = [desc[0] for desc in cur.description]
                    rows = cur.fetchall()
                    return [dict(zip(columns, row)) for row in rows]
        except Exception as e:
            logger.error(f"获取品类列表失败: {e}")
            raise

    # ==================== 接口3b：上传品种 ====================

    def upload_variety(self, items: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        批量维护品种（dict_categories）：不存在则新建分组并 is_main=1；
        名称已存在且启用则跳过；已存在但停用则恢复启用。
        """
        if not items:
            raise ValueError("品种数据不能为空")

        seen: set[str] = set()
        names: List[str] = []
        for item in items:
            raw = item.get("品种名")
            if raw is None:
                continue
            for n in _split_category_alias_names(raw):
                if len(n) > 50:
                    raise ValueError(f"品种名长度不能超过50字符: {n[:30]}…")
                if n in seen:
                    continue
                seen.add(n)
                names.append(n)

        if not names:
            raise ValueError("无有效的品种名")

        created = existed = reactivated = 0
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    for n in names:
                        cur.execute(
                            "SELECT row_id, category_id, is_active "
                            "FROM dict_categories WHERE name = %s",
                            (n,),
                        )
                        row = cur.fetchone()
                        if row:
                            _rid, _cid, is_active = row
                            if is_active == 1:
                                existed += 1
                            else:
                                cur.execute(
                                    "UPDATE dict_categories SET is_active = 1 WHERE row_id = %s",
                                    (_rid,),
                                )
                                reactivated += 1
                        else:
                            cur.execute(
                                "SELECT COALESCE(MAX(category_id), 0) + 1 FROM dict_categories"
                            )
                            new_cat_id = cur.fetchone()[0]
                            cur.execute(
                                "INSERT INTO dict_categories "
                                "(category_id, name, is_main, is_active) "
                                "VALUES (%s, %s, 1, 1)",
                                (new_cat_id, n),
                            )
                            created += 1

            parts = []
            if created:
                parts.append(f"新建 {created} 个")
            if existed:
                parts.append(f"已存在 {existed} 个")
            if reactivated:
                parts.append(f"恢复启用 {reactivated} 个")
            msg = "、".join(parts) if parts else "无变更"
            return {
                "code": 200,
                "msg": f"品种已处理：{msg}",
                "新建": created,
                "已存在": existed,
                "恢复启用": reactivated,
            }
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"上传品种失败: {e}")
            raise

    # ==================== 接口4：获取比价表 ====================
    def get_comparison(
        self,
        warehouse_ids: List[int],
        smelter_ids: List[int],
        category_ids: List[int],
        price_type: Optional[str] = None,
        tons: float = 1.0,
        tons_by_category: Optional[Dict[int, float]] = None,
        optimal_basis_list: Optional[List[str]] = None,
        optimal_sort_basis: Optional[str] = None,
        quote_date_str: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        price_type: 目标税率类型，None=普通价, 1pct/3pct/13pct/normal_invoice/reverse_invoice
        吨数 t（按品类）：**单价**为展示用「报价」（元/吨，按 price_type 折合不含税）；**总价** = 单价×t = **`报价金额`**；
        **运费单价**（元/吨）来自运费表；**运费** = 运费单价×t = **`总运费`**（全程运费金额，元）；
        若传入 ``tons_by_category``，则每个品类使用对应 t；否则所有品类共用 ``tons``。
        **利润** = 总价 − 运费（与 **报价金额 − 总运费** 一致）。
        明细中同时保留 **`报价`/`报价金额`/`总运费`** 与上述 **`单价`/`总价`/`运费单价`/`运费`** 便于新旧前端兼容。
        前端最终比价、明细排序与 **`冶炼厂利润排行`** 均以该 **`利润`**（及所选最优价口径）为准。
        **最优价各口径利润**=该口径下元/吨单价×t−总运费（与主利润同一套总运费）。
        同时按表中已有列统一反推 `基准价`（不含税）、`含1%税价`、`含3%税价`（与 OCR 按税点入库、再换算一致）；
        `利润_基准`=基准价×t−总运费，`利润_含3%`=含3%税价×t−总运费。
        **最优价各口径利润**：由 optimal_basis_list 指定（如 base、1pct、3pct、13pct、普票列等），每条明细返回 `最优价各口径利润` 字典；
        明细与冶炼厂排行按 optimal_sort_basis（默认列表首项）对应利润从高到低排序。
        取价逻辑（按优先级）：
          1. 报价表中直接有对应 price_type 的价格 → 直接使用
          2. 有不含税 unit_price（基准价）+ 税率表 → 目标含税价 = unit_price × (1+税率)
          3. 目标是普通价(unit_price) 但列空 → 由已知含1%/3%/13%价反算不含税基准
          4. 仅有某一档含税价 → 先反算不含税，再换算到目标税率
          5. 以上均无 → None，返回 price_source="unavailable"

        **报价日期**：
        - 若传入 `quote_date_str`（YYYY-MM-DD）：只使用该日的 `quote_details`。
        - 否则：以比价基准日（默认 `Asia/Shanghai` 当天，见 `QUOTE_COMPARISON_TZ`）为参照，
          对每个 (冶炼厂, 品种名) 在 `quote_details` 中取 **与基准日日历距离最近** 的一条；
          距离相同时按 **`created_at` 最新**（视为最近上传/写入）优先，再以 `id` 较大者优先。

        **冶炼厂与明细范围**：
        - `冶炼厂id列表` 仅过滤停用厂后 **保持请求中的顺序与去重结果**；
        - 对每个「选中仓库 × 上述冶炼厂 × 选中品类」均返回明细行；若该组合在 `freight_rates` 中尚无记录，
          **运费单价按 0** 补行，避免「库里勾选了厂但响应里完全没有该厂」的错位感。

        **循融宝（dict_factories.use_xunrongbao=1）**：
        - 库内 `quote_details` 按**不含**循融宝吨加价；加价版按 `XUNRONGBAO_SHIPPING_PREMIUM_PER_TON`（默认 80 元/吨）
          在不含税基准上加价后重算各含税列（与 `apply_per_ton_premium_to_quote_row` 一致）。
        - 明细**顶层**的单价/总价/利润/最优价等仍为**含循融宝**口径（与历史接口一致，便于沿用排序与冶炼厂排行）；
          另返回 **`不含循融宝`**、**`含循融宝`** 两个字典，结构相同、分别标注两种计价；非循融宝厂该两字为 `null`，
          **`冶炼厂循融宝发货`** 为 0，**`循融宝加价元每吨`** 为 `null`。
        """
        if not warehouse_ids or not smelter_ids or not category_ids:
            return {
                "明细": [],
                "冶炼厂利润排行": [],
                "最优价排序口径": (optimal_sort_basis or (optimal_basis_list or ["3pct"])[0]),
            }

        bases = list(optimal_basis_list or ["3pct"])
        sort_basis = optimal_sort_basis if optimal_sort_basis is not None else bases[0]
        for b in bases:
            if b not in OPTIMAL_PRICE_BASIS_ALLOWED:
                raise ValueError(
                    f"不支持的最优价计税口径: {b!r}，允许：{sorted(OPTIMAL_PRICE_BASIS_ALLOWED)}"
                )
        if sort_basis not in bases:
            raise ValueError(
                f"最优价排序口径 {sort_basis!r} 须在最优价计税口径列表中，当前为 {bases}"
            )

        # price_type → (quote_details列名, 展示名)
        PRICE_COL_MAP = {
            None:             ("unit_price",            "普通价"),
            "1pct":           ("price_1pct_vat",        "1%增值税"),
            "3pct":           ("price_3pct_vat",        "3%增值税"),
            "13pct":          ("price_13pct_vat",       "13%增值税"),
            "normal_invoice": ("price_normal_invoice",  "普通发票"),
            "reverse_invoice":("price_reverse_invoice", "反向发票"),
        }
        # 仅以下三种有税率换算意义
        VAT_TAX_TYPE_MAP = {"1pct": "1pct", "3pct": "3pct", "13pct": "13pct"}

        if price_type not in PRICE_COL_MAP:
            raise ValueError(f"不支持的 price_type: {price_type}")

        target_col, price_type_name = PRICE_COL_MAP[price_type]
        target_tax = VAT_TAX_TYPE_MAP.get(price_type)  # None 表示不需要税率换算

        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    wh_ph = ",".join(["%s"] * len(warehouse_ids))
                    sm_ph = ",".join(["%s"] * len(smelter_ids))
                    cat_ph = ",".join(["%s"] * len(category_ids))

                    req_smelter_ids = [int(x) for x in smelter_ids]
                    cur.execute(
                        f"SELECT id FROM dict_factories WHERE id IN ({sm_ph}) AND is_active = 1",
                        tuple(req_smelter_ids),
                    )
                    active_sm = {int(row[0]) for row in cur.fetchall()}
                    # 与请求顺序一致（仅去掉停用/无效 id、去重），避免 fetchall 顺序与前端所选冶炼厂错位
                    smelter_ids = []
                    _seen_sm: Set[int] = set()
                    for sid in req_smelter_ids:
                        if sid in active_sm and sid not in _seen_sm:
                            _seen_sm.add(sid)
                            smelter_ids.append(sid)
                    if not smelter_ids:
                        return {
                            "明细": [],
                            "冶炼厂利润排行": [],
                            "最优价排序口径": sort_basis,
                        }
                    sm_ph = ",".join(["%s"] * len(smelter_ids))

                    # 品类主名称（用于展示）
                    cur.execute(
                        f"SELECT DISTINCT category_id, "
                        f"COALESCE(MAX(CASE WHEN is_main=1 THEN name END), MAX(name)) AS cat_name "
                        f"FROM dict_categories "
                        f"WHERE category_id IN ({cat_ph}) AND is_active = 1 "
                        f"GROUP BY category_id",
                        tuple(category_ids),
                    )
                    cat_map: Dict[int, str] = {row[0]: row[1] for row in cur.fetchall()}

                    # 最新运费
                    cur.execute(
                        f"""
                        SELECT dw.id, dw.name, df.id, df.name, fr.price_per_ton
                        FROM freight_rates fr
                        JOIN dict_warehouses dw ON fr.warehouse_id = dw.id
                        JOIN dict_factories  df ON fr.factory_id  = df.id
                        WHERE dw.id IN ({wh_ph})
                          AND df.id IN ({sm_ph})
                          AND fr.effective_date = (
                              SELECT MAX(fr2.effective_date)
                              FROM freight_rates fr2
                              WHERE fr2.factory_id  = fr.factory_id
                                AND fr2.warehouse_id = fr.warehouse_id
                          )
                        """,
                        tuple(warehouse_ids) + tuple(smelter_ids),
                    )
                    freight_map: Dict[tuple, tuple] = {}
                    for wid, wname, fid, fname, freight in cur.fetchall():
                        freight_map[(int(wid), int(fid))] = (wname, fname, freight)

                    # 请求中的每个「仓库×启用冶炼厂」都应有明细行；无运费记录时按 0 元/吨补全，避免与所选冶炼厂列表对不齐
                    cur.execute(
                        f"SELECT id, name, contact_name, contact_phone, "
                        f"hazardous_waste_license_qty, monthly_avg_receipt_ton, freight_amount "
                        f"FROM dict_warehouses WHERE id IN ({wh_ph})",
                        tuple(warehouse_ids),
                    )
                    wid_to_name: Dict[int, str] = {}
                    wid_to_wh_ext: Dict[int, Dict[str, Any]] = {}
                    for r in cur.fetchall():
                        wid_i = int(r[0])
                        wid_to_name[wid_i] = str(r[1])
                        hw_v, mar_v, fa_v = r[4], r[5], r[6]
                        wid_to_wh_ext[wid_i] = {
                            "库房联系人": (r[2] or "") if r[2] is not None else "",
                            "电话": (r[3] or "") if r[3] is not None else "",
                            "危废经营许可数量": (
                                float(hw_v) if hw_v is not None else None
                            ),
                            "月均收货": (
                                float(mar_v) if mar_v is not None else None
                            ),
                            # 与 get_warehouses 中 dict_warehouses.freight_amount 一致；
                            # 明细顶层已有「运费」表示全程运费(元)，此处用「运费参考」避免键冲突
                            "运费参考": (
                                float(fa_v) if fa_v is not None else None
                            ),
                        }
                    cur.execute(
                        f"SELECT id, name FROM dict_factories WHERE id IN ({sm_ph})",
                        tuple(smelter_ids),
                    )
                    fid_to_name: Dict[int, str] = {
                        int(r[0]): str(r[1]) for r in cur.fetchall()
                    }
                    for wid in warehouse_ids:
                        w_int = int(wid)
                        for fid in smelter_ids:
                            key = (w_int, int(fid))
                            if key not in freight_map:
                                wn = wid_to_name.get(w_int, f"仓库{w_int}")
                                fn = fid_to_name.get(int(fid), f"冶炼厂{int(fid)}")
                                freight_map[key] = (wn, fn, 0.0)

                    # category_id → 品类名称列表（用于匹配价格表）
                    cur.execute(
                        f"SELECT category_id, name FROM dict_categories "
                        f"WHERE category_id IN ({cat_ph}) AND is_active = 1",
                        tuple(category_ids),
                    )
                    cat_id_to_names: Dict[int, List[str]] = {}
                    for cat_id, name in cur.fetchall():
                        n = str(name).strip()
                        if not n:
                            continue
                        lst = cat_id_to_names.setdefault(cat_id, [])
                        if n not in lst:
                            lst.append(n)

                    if not cat_id_to_names:
                        return {
                            "明细": [],
                            "冶炼厂利润排行": [],
                            "最优价排序口径": sort_basis,
                        }

                    # 所有品类名称（去重，与 quote_details 用 TRIM 后匹配）
                    all_cat_names: List[str] = []
                    _seen_cn: set = set()
                    for names in cat_id_to_names.values():
                        for n in names:
                            if n not in _seen_cn:
                                _seen_cn.add(n)
                                all_cat_names.append(n)
                    cn_ph = ",".join(["%s"] * len(all_cat_names))

                    # 税率表：{factory_id: {tax_type: rate}}
                    cur.execute(
                        f"SELECT factory_id, tax_type, tax_rate "
                        f"FROM factory_tax_rates "
                        f"WHERE factory_id IN ({sm_ph})",
                        tuple(smelter_ids),
                    )
                    tax_rate_map: Dict[int, Dict[str, float]] = {}
                    for fid, ttype, rate in cur.fetchall():
                        tax_rate_map.setdefault(fid, {})[ttype] = float(rate)

                    if quote_date_str is not None and str(quote_date_str).strip():
                        try:
                            exact_qd = date.fromisoformat(str(quote_date_str).strip())
                        except (ValueError, TypeError):
                            raise ValueError(
                                f"报价日期 格式不正确: {quote_date_str}，应为 YYYY-MM-DD"
                            )
                        cur.execute(
                            f"""
                            SELECT factory_id, TRIM(category_name) AS category_name,
                                   unit_price, price_1pct_vat, price_3pct_vat, price_13pct_vat,
                                   price_normal_invoice, price_reverse_invoice
                            FROM quote_details
                            WHERE factory_id IN ({sm_ph})
                              AND TRIM(category_name) IN ({cn_ph})
                              AND quote_date = %s
                            """,
                            tuple(smelter_ids) + tuple(all_cat_names) + (exact_qd,),
                        )
                    else:
                        ref_day = _comparison_quote_calendar_date()
                        cur.execute(
                            f"""
                            SELECT qd.factory_id, TRIM(qd.category_name) AS category_name,
                                   qd.unit_price, qd.price_1pct_vat, qd.price_3pct_vat,
                                   qd.price_13pct_vat,
                                   qd.price_normal_invoice, qd.price_reverse_invoice
                            FROM quote_details qd
                            INNER JOIN (
                                SELECT id FROM (
                                    SELECT id,
                                           ROW_NUMBER() OVER (
                                               PARTITION BY factory_id, TRIM(category_name)
                                               ORDER BY ABS(DATEDIFF(quote_date, %s)) ASC,
                                                        created_at DESC,
                                                        id DESC
                                           ) AS rn
                                    FROM quote_details
                                    WHERE factory_id IN ({sm_ph})
                                      AND TRIM(category_name) IN ({cn_ph})
                                ) ranked
                                WHERE ranked.rn = 1
                            ) pick ON pick.id = qd.id
                            """,
                            (ref_day,)
                            + tuple(smelter_ids)
                            + tuple(all_cat_names),
                        )
                    # raw_price_map: {(factory_id, category_name): {col: value}}
                    col_names = ["unit_price", "price_1pct_vat", "price_3pct_vat",
                                 "price_13pct_vat", "price_normal_invoice", "price_reverse_invoice"]
                    raw_price_map: Dict[tuple, Dict[str, Optional[float]]] = {}
                    name_to_cat_id: Dict[str, int] = {}
                    for row in cur.fetchall():
                        fid_r, cat_name = row[0], str(row[1]).strip() if row[1] is not None else ""
                        if not cat_name:
                            continue
                        raw_price_map[(fid_r, cat_name)] = {
                            col: (float(v) if v is not None else None)
                            for col, v in zip(col_names, row[2:])
                        }
                        for cat_id, names in cat_id_to_names.items():
                            if cat_name in names:
                                name_to_cat_id[cat_name] = cat_id
                                break

                    cur.execute(
                        f"SELECT id, COALESCE(use_xunrongbao, 0) FROM dict_factories "
                        f"WHERE id IN ({sm_ph})",
                        tuple(smelter_ids),
                    )
                    xrb_fids = {int(r[0]) for r in cur.fetchall() if int(r[1]) == 1}
                    # 循融宝厂：库内报价为「不含加价」；加价版单独建表，与 raw_price_map 并行取价
                    raw_price_map_xrb: Dict[tuple, Dict[str, Optional[float]]] = {}
                    for map_key, prow in raw_price_map.items():
                        fid_k, _cname = map_key
                        if fid_k not in xrb_fids:
                            continue
                        merged_x = merge_factory_rates(tax_rate_map.get(fid_k, {}))
                        raw_price_map_xrb[map_key] = apply_per_ton_premium_to_quote_row(
                            dict(prow),
                            merged_x,
                            XUNRONGBAO_SHIPPING_PREMIUM_PER_TON,
                        )

            # 换算逻辑（纯 Python，连接已关闭）
            # col → tax_type 的对应关系，用于反算不含税价
            COL_TO_TAX: Dict[str, str] = {
                "price_1pct_vat": "1pct",
                "price_3pct_vat": "3pct",
                "price_13pct_vat": "13pct",
            }

            def resolve_price(
                fid: int,
                cat_id: int,
                raw_map: Dict[tuple, Dict[str, Optional[float]]],
            ) -> Tuple[Optional[float], str, Optional[Dict[str, Optional[float]]]]:
                """
                返回 (price, source, prices_row)。
                prices_row 为本次取价所依据的 quote 列字典，须与 derive_net_and_vat_from_quote_row
                使用同一份数据；否则在品类多别名时，若首个别名在库中为空壳行（键全为 None），
                会出现主行「利润」正常但「利润_含3%」「最优价各口径利润」及冶炼厂汇总合计为 0 的错误。
                """
                cat_names = cat_id_to_names.get(cat_id, [])
                for cat_name in cat_names:
                    prices = raw_map.get((fid, cat_name), {})
                    if not prices or not any(
                        v is not None for v in prices.values()
                    ):
                        continue

                    rates = tax_rate_map.get(fid, {})
                    merged = merge_factory_rates(rates)

                    # 1. 直接有目标列
                    direct = prices.get(target_col)
                    if direct is not None:
                        return direct, "direct", prices

                    # 2. 不含税 unit_price → 目标税率含税价
                    if target_tax and prices.get("unit_price") is not None and target_tax in merged:
                        up = float(prices["unit_price"])
                        calc = inclusive_from_net(up, merged[target_tax])
                        return calc, "calc_from_base", prices

                    # 3. 目标为不含税基准，由已知含税价反算
                    if target_col == "unit_price":
                        for col, src_tax in COL_TO_TAX.items():
                            known_price = prices.get(col)
                            if known_price is not None and src_tax in merged:
                                net = net_from_inclusive(float(known_price), merged[src_tax])
                                return round(net, 2), f"calc_from_{src_tax}", prices
                        # 与 derive_net_and_vat_from_quote_row 一致：仅有普票/反向发票列时按不含税理解
                        for col in ("price_normal_invoice", "price_reverse_invoice"):
                            v = prices.get(col)
                            if v is not None:
                                return float(v), f"direct_{col}", prices

                    # 4. 从某一档含税价反算不含税，再换算到目标税率
                    if target_tax and target_tax in merged:
                        for col, src_tax in COL_TO_TAX.items():
                            known_price = prices.get(col)
                            if known_price is not None and src_tax in merged:
                                net = net_from_inclusive(float(known_price), merged[src_tax])
                                calc = inclusive_from_net(net, merged[target_tax])
                                return calc, f"calc_from_{src_tax}", prices

                return None, "unavailable", None

            # 组合结果；总运费 = 运费单价（元/吨）× 该品类吨数；总价 = 单价（元/吨）× 该品类吨数
            def tons_for_category(cid: int) -> float:
                if tons_by_category is not None:
                    return float(tons_by_category[int(cid)])
                return float(tons)

            _xrb_nested_keys = (
                "单价",
                "总价",
                "运费单价",
                "运费",
                "总运费",
                "报价",
                "报价金额",
                "报价来源",
                "基准价",
                "含1%税价",
                "含3%税价",
                "利润",
                "利润_基准",
                "利润_含3%",
                "最优价各口径利润",
            )

            def _xrb_branch_snapshot(metrics: Dict[str, Any]) -> Dict[str, Any]:
                snap: Dict[str, Any] = {}
                for k in _xrb_nested_keys:
                    v = metrics[k]
                    if k == "最优价各口径利润" and isinstance(v, dict):
                        snap[k] = dict(v)
                    else:
                        snap[k] = v
                return snap

            result: List[Dict[str, Any]] = []
            for (wid, fid), (wname, fname, freight) in freight_map.items():
                for cid in category_ids:
                    t = tons_for_category(cid)
                    cat_name = cat_map.get(cid)
                    if cat_name is None:
                        continue
                    fr = float(freight) if freight is not None else 0.0
                    merged = merge_factory_rates(tax_rate_map.get(fid, {}))
                    xrb_on = fid in xrb_fids

                    price_exc, source_exc, qrow_exc = resolve_price(
                        fid, cid, raw_price_map
                    )
                    if xrb_on:
                        price_inc, source_inc, qrow_inc = resolve_price(
                            fid, cid, raw_price_map_xrb
                        )
                        metrics_exc = _build_comparison_price_metrics(
                            price_exc,
                            source_exc,
                            qrow_exc,
                            merged,
                            target_tax,
                            t,
                            fr,
                            bases,
                        )
                        metrics_inc = _build_comparison_price_metrics(
                            price_inc,
                            source_inc,
                            qrow_inc,
                            merged,
                            target_tax,
                            t,
                            fr,
                            bases,
                        )
                        top = metrics_inc
                    else:
                        metrics_inc = _build_comparison_price_metrics(
                            price_exc,
                            source_exc,
                            qrow_exc,
                            merged,
                            target_tax,
                            t,
                            fr,
                            bases,
                        )
                        metrics_exc = metrics_inc
                        top = metrics_inc

                    wext = wid_to_wh_ext.get(
                        int(wid),
                        {
                            "库房联系人": "",
                            "电话": "",
                            "危废经营许可数量": None,
                            "月均收货": None,
                            "运费参考": None,
                        },
                    )
                    rec: Dict[str, Any] = {
                        "仓库id": wid,
                        "冶炼厂id": fid,
                        "品类id": cid,
                        "仓库": wname,
                        "冶炼厂": fname,
                        "品类": cat_name,
                        "price_type": price_type_name,
                        "吨数": t,
                        "运费计价方式": "per_ton",
                        **wext,
                        **top,
                        "冶炼厂循融宝发货": 1 if xrb_on else 0,
                        "循融宝加价元每吨": (
                            float(XUNRONGBAO_SHIPPING_PREMIUM_PER_TON)
                            if xrb_on
                            else None
                        ),
                        "不含循融宝": (
                            _xrb_branch_snapshot(metrics_exc) if xrb_on else None
                        ),
                        "含循融宝": (
                            _xrb_branch_snapshot(metrics_inc) if xrb_on else None
                        ),
                    }
                    result.append(rec)

            result.sort(
                key=lambda r: (
                    r["最优价各口径利润"][sort_basis]
                    if r["最优价各口径利润"].get(sort_basis) is not None
                    else float("-inf")
                ),
                reverse=True,
            )

            # 按冶炼厂汇总；排行按「最优价排序口径」对应利润合计从高到低
            agg: Dict[int, Dict[str, Any]] = {}
            for row in result:
                sfid = int(row["冶炼厂id"])
                if sfid not in agg:
                    agg[sfid] = {
                        "冶炼厂id": sfid,
                        "冶炼厂": row["冶炼厂"],
                        "利润": 0.0,
                        "利润_含3%合计": 0.0,
                        "利润_基准合计": 0.0,
                        "最优价口径合计": {b: 0.0 for b in bases},
                    }
                agg[sfid]["利润"] += float(row["利润"])
                if row["利润_含3%"] is not None:
                    agg[sfid]["利润_含3%合计"] += float(row["利润_含3%"])
                if row["利润_基准"] is not None:
                    agg[sfid]["利润_基准合计"] += float(row["利润_基准"])
                op = row["最优价各口径利润"]
                for b in bases:
                    pv = op.get(b)
                    if pv is not None:
                        agg[sfid]["最优价口径合计"][b] += float(pv)

            ranking = sorted(
                (
                    {
                        **v,
                        "利润": round(v["利润"], 2),
                        "利润_含3%合计": round(v["利润_含3%合计"], 2),
                        "利润_基准合计": round(v["利润_基准合计"], 2),
                        "最优价口径合计": {
                            b: round(v["最优价口径合计"][b], 2) for b in bases
                        },
                    }
                    for v in agg.values()
                ),
                key=lambda x: x["最优价口径合计"][sort_basis],
                reverse=True,
            )
            return {
                "明细": result,
                "冶炼厂利润排行": ranking,
                "最优价排序口径": sort_basis,
            }

        except Exception as e:
            logger.error(f"获取比价表失败: {e}")
            raise

    # ==================== 税率表 CRUD ====================

    def get_tax_rates(self, factory_ids: Optional[List[int]] = None) -> List[Dict[str, Any]]:
        """获取税率表，可按冶炼厂过滤"""
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    if factory_ids:
                        ph = ",".join(["%s"] * len(factory_ids))
                        cur.execute(
                            f"SELECT ftr.id, ftr.factory_id, df.name AS factory_name, "
                            f"ftr.tax_type, ftr.tax_rate "
                            f"FROM factory_tax_rates ftr "
                            f"JOIN dict_factories df ON ftr.factory_id = df.id "
                            f"WHERE ftr.factory_id IN ({ph}) "
                            f"ORDER BY ftr.factory_id, ftr.tax_type",
                            tuple(factory_ids),
                        )
                    else:
                        cur.execute(
                            "SELECT ftr.id, ftr.factory_id, df.name AS factory_name, "
                            "ftr.tax_type, ftr.tax_rate "
                            "FROM factory_tax_rates ftr "
                            "JOIN dict_factories df ON ftr.factory_id = df.id "
                            "ORDER BY ftr.factory_id, ftr.tax_type"
                        )
                    cols = [d[0] for d in cur.description]
                    return [dict(zip(cols, row)) for row in cur.fetchall()]
        except Exception as e:
            logger.error(f"获取税率表失败: {e}")
            raise

    def upsert_tax_rates(self, items: List[Dict[str, Any]]) -> Dict[str, Any]:
        """批量设置税率（存在则更新，不存在则插入）"""
        from app.models.tl import VALID_TAX_TYPES
        for item in items:
            if item["tax_type"] not in VALID_TAX_TYPES:
                raise ValueError(f"不支持的 tax_type: {item['tax_type']}，有效值：{VALID_TAX_TYPES}")
            if not (0 <= item["tax_rate"] <= 1):
                raise ValueError(f"tax_rate 必须在 0~1 之间，收到：{item['tax_rate']}")
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    for item in items:
                        # 验证冶炼厂是否存在
                        cur.execute("SELECT id FROM dict_factories WHERE id = %s", (item["factory_id"],))
                        if not cur.fetchone():
                            raise ValueError(f"冶炼厂 ID {item['factory_id']} 不存在")

                        cur.execute(
                            "INSERT INTO factory_tax_rates (factory_id, tax_type, tax_rate) "
                            "VALUES (%s, %s, %s) "
                            "ON DUPLICATE KEY UPDATE tax_rate = VALUES(tax_rate), "
                            "updated_at = CURRENT_TIMESTAMP",
                            (item["factory_id"], item["tax_type"], item["tax_rate"]),
                        )
            log_finance_event(
                "税率变更 | 保存条数=%s | 冶炼厂ids=%s | 明细=%s",
                len(items),
                sorted({int(i["factory_id"]) for i in items}),
                json.dumps(
                    [
                        {
                            "factory_id": i["factory_id"],
                            "tax_type": i["tax_type"],
                            "tax_rate": i["tax_rate"],
                        }
                        for i in items
                    ],
                    ensure_ascii=False,
                ),
            )
            return {"code": 200, "msg": f"已保存 {len(items)} 条税率记录"}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"设置税率失败: {e}")
            raise

    def delete_tax_rate(self, factory_id: int, tax_type: str) -> Dict[str, Any]:
        """删除某冶炼厂的某税率记录"""
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "DELETE FROM factory_tax_rates WHERE factory_id = %s AND tax_type = %s",
                        (factory_id, tax_type),
                    )
                    if cur.rowcount == 0:
                        raise ValueError(f"未找到 factory_id={factory_id}, tax_type={tax_type} 的记录")
            log_finance_event(
                "税率删除 | factory_id=%s tax_type=%s",
                factory_id,
                tax_type,
            )
            return {"code": 200, "msg": "删除成功"}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"删除税率失败: {e}")
            raise

    # ==================== 接口5：上传价格表（OCR解析） ====================

    def _match_factory(
        self, ocr_name: str, factory_list: List[Tuple[int, str]]
    ) -> Optional[int]:
        """将 OCR 识别出的工厂名匹配到 dict_factories 中的冶炼厂，返回 factory_id"""
        if not ocr_name or ocr_name == "未知工厂":
            return None
        for fid, fname in factory_list:
            # 双向包含匹配
            if fname in ocr_name or ocr_name in fname:
                return fid
        return None

    def _match_category(
        self, ocr_cat: str, category_list: List[Tuple[int, int, str]]
    ) -> Optional[Tuple[int, int]]:
        """将 OCR 识别出的品类名匹配到 dict_categories，返回 (category_id, row_id)"""
        if not ocr_cat:
            return None
        for row_id, cat_id, cname in category_list:
            if cname in ocr_cat or ocr_cat in cname:
                return (cat_id, row_id)
        return None

    def _resolve_quote_category_main_name(
        self,
        cur: Any,
        raw_name: Any,
        *,
        allow_create: bool,
    ) -> Tuple[str, int]:
        """
        报价落库前将识别/录入的品类名归一到启用主名称。

        - 命中启用别名：返回同 category_id 下的启用主名称；
        - 命中停用别名：拒绝写入；
        - 未命中：按现有确认报价逻辑可新建为主名称（allow_create=True）。
        """
        cat_name = str(raw_name or "").strip()
        if not cat_name:
            raise ValueError("品类名不能为空")

        cur.execute(
            "SELECT row_id, category_id, is_active FROM dict_categories WHERE name = %s",
            (cat_name,),
        )
        row = cur.fetchone()
        if row:
            _row_id, category_id, is_active = row
            if is_active != 1:
                raise ValueError(
                    f"品类「{cat_name}」已停用，请先在品类管理中启用后再写入报价。"
                )
            cur.execute(
                "SELECT name FROM dict_categories "
                "WHERE category_id = %s AND is_active = 1 "
                "ORDER BY is_main DESC, row_id ASC LIMIT 1",
                (category_id,),
            )
            main_row = cur.fetchone()
            if not main_row or not str(main_row[0]).strip():
                raise ValueError(f"品类「{cat_name}」所在分组没有启用名称，请先维护品类。")
            return str(main_row[0]).strip(), int(category_id)

        if not allow_create:
            raise ValueError(f"品类不存在或未启用: {cat_name}")

        cur.execute("SELECT COALESCE(MAX(category_id), 0) + 1 FROM dict_categories")
        new_cat_id = int(cur.fetchone()[0])
        cur.execute(
            "INSERT INTO dict_categories "
            "(category_id, name, is_main, is_active) "
            "VALUES (%s, %s, 1, 1)",
            (new_cat_id, cat_name),
        )
        return cat_name, new_cat_id

    def upload_price_table(self, files: List[Any]) -> Dict[str, Any]:
        saved_paths: List[Tuple[str, str, str]] = []
        try:
            # 1. 保存图片到磁盘
            for upload_file in files:
                content = upload_file.file.read()
                md5 = hashlib.md5(content).hexdigest()
                suffix = Path(upload_file.filename).suffix or ".jpg"
                filename = f"{uuid.uuid4().hex}{suffix}"
                save_path = PRICE_TABLE_UPLOAD_DIR / filename

                with open(save_path, "wb") as f:
                    f.write(content)
                saved_paths.append((str(save_path), md5, upload_file.filename))

            # 2. VLM识别
            from app import config as app_config
            if not app_config.VLM_API_KEY:
                raise ValueError("未配置 VLM_API_KEY，请在环境变量中设置 VLM_API_KEY")
            vlm_config = VLMConfig(
                api_key=app_config.VLM_API_KEY,
                base_url=app_config.VLM_BASE_URL,
                model=app_config.VLM_MODEL,
                max_tokens=app_config.VLM_MAX_TOKENS,
                image_max_edge=app_config.VLM_IMAGE_MAX_EDGE,
                jpeg_quality=app_config.VLM_JPEG_QUALITY,
                request_timeout=app_config.VLM_REQUEST_TIMEOUT,
                save_individual=False,
            )

            details = []
            with QwenVLFullExtractor(vlm_config) as extractor:
                for image_path, md5, orig_name in saved_paths:
                    result = extractor.recognize(image_path, save_output=False)

                    if not result.success:
                        details.append({
                            "image": orig_name,
                            "success": False,
                            "error": result.error_message,
                        })
                        continue

                    # 3. 构建 full_data（VlmFullData格式，供前端保留并回传）
                    full_data = {
                        "image_path": result.image_path,
                        "file_name": result.file_name,
                        "source_image": orig_name,
                        "company_name": result.company_name,
                        "doc_title": result.doc_title,
                        "subtitle": result.subtitle,
                        "quote_date": result.quote_date,
                        "execution_date": result.execution_date,
                        "valid_period": result.valid_period,
                        "price_unit": result.price_unit,
                        "headers": result.headers,
                        "rows": [row.model_dump() for row in result.rows],
                        "policies": result.policies,
                        "footer_notes": result.footer_notes,
                        "footer_notes_raw": result.footer_notes_raw,
                        "brand_specifications": result.brand_specifications,
                        "raw_full_text": result.raw_full_text,
                        "elapsed_time": result.elapsed_time,
                    }

                    # 4. 映射为前端可编辑的 items（ConfirmPriceTableItem格式）
                    items = self._map_vlm_to_confirm_items(result)

                    details.append({
                        "image": orig_name,
                        "success": True,
                        "full_data": full_data,
                        "items": items,
                    })

            return {"code": 200, "data": {"details": details}}

        except Exception as e:
            logger.error(f"上传价格表失败: {e}")
            for path, _, _ in saved_paths:
                try:
                    os.remove(path)
                except OSError:
                    pass
            raise

    @staticmethod
    def _normalize_excel_header_cell(s: Any) -> str:
        if s is None or (isinstance(s, float) and str(s) == "nan"):
            return ""
        return str(s).replace("\u3000", " ").strip()

    @classmethod
    def _classify_quote_excel_column(cls, col_name: Any) -> Optional[str]:
        """
        将表头映射为逻辑字段：smelter, category, quote_date, net, p1, p3, p13,
        normal_inv, reverse_inv, remark, basis。
        """
        x = cls._normalize_excel_header_cell(col_name)
        if not x:
            return None
        low = x.casefold()
        exact: Dict[str, str] = {
            "冶炼厂": "smelter",
            "冶炼厂名": "smelter",
            "厂家": "smelter",
            "工厂": "smelter",
            "smelter": "smelter",
            "factory": "smelter",
            "品种": "category",
            "品类": "category",
            "品类名": "category",
            "variety": "category",
            "category": "category",
            "日期": "quote_date",
            "报价日期": "quote_date",
            "基准价": "net",
            "普通价": "net",
            "不含税价": "net",
            "不含税基准价": "net",
            "价格": "net",
            "单价": "net",
            "unit_price": "net",
            "3%含税价": "p3",
            "价格_3pct增值税": "p3",
            "13%含税价": "p13",
            "价格_13pct增值税": "p13",
            "1%含税价": "p1",
            "价格_1pct增值税": "p1",
            "普通发票价格": "normal_inv",
            "反向发票价格": "reverse_inv",
            "备注": "remark",
            "价格口径": "basis",
        }
        if x in exact:
            return exact[x]
        if low in exact:
            return exact[low]
        if "含3%" in x or "3%专" in x or "3%增值税" in x:
            return "p3"
        if "含13%" in x or "13%专" in x or "13%增值税" in x:
            return "p13"
        if "含1%" in x or "1%普" in x or "1%增值税" in x:
            return "p1"
        if "反向发票" in x:
            return "reverse_inv"
        if "普通发票" in x and "价格" in x:
            return "normal_inv"
        return None

    @staticmethod
    def _coerce_excel_price(v: Any) -> Optional[float]:
        if v is None:
            return None
        if isinstance(v, bool):
            return None
        if isinstance(v, str):
            s = v.strip()
            if not s:
                return None
            v = s.replace(",", "").replace("，", "")
        try:
            x = float(v)
        except (TypeError, ValueError):
            return None
        if x != x or x in (float("inf"), float("-inf")):
            return None
        return round(x, 4)

    @classmethod
    def _excel_row_dict_to_confirm_item(cls, row_fields: Dict[str, Any]) -> Dict[str, Any]:
        """将一行解析后的逻辑字段转为 confirm_price_table 所需的 item。"""
        sm = cls._normalize_excel_header_cell(row_fields.get("smelter"))
        cat = cls._normalize_excel_header_cell(row_fields.get("category"))
        if not sm or not cat:
            raise ValueError("行缺少冶炼厂或品种（品类）")

        net = cls._coerce_excel_price(row_fields.get("net"))
        p1 = cls._coerce_excel_price(row_fields.get("p1"))
        p3 = cls._coerce_excel_price(row_fields.get("p3"))
        p13 = cls._coerce_excel_price(row_fields.get("p13"))
        pn = cls._coerce_excel_price(row_fields.get("normal_inv"))
        pr = cls._coerce_excel_price(row_fields.get("reverse_inv"))
        remark_raw = row_fields.get("remark")
        remark = cls._normalize_excel_header_cell(remark_raw) or None
        basis_raw = row_fields.get("basis")
        basis_s = cls._normalize_excel_header_cell(basis_raw) or None
        _basis_ok = frozenset({"ex_vat", "incl_1pct", "incl_3pct", "incl_13pct"})
        if basis_s and basis_s in _basis_ok:
            basis = basis_s
        else:
            basis = parse_price_basis_from_remark(
                ((remark or "") + " " + (basis_s or "")).strip()
            )

        src: Dict[str, str] = {}
        if net is not None:
            src["unit_price"] = SOURCE_ORIGINAL
        if p1 is not None:
            src["price_1pct_vat"] = SOURCE_ORIGINAL
        if p3 is not None:
            src["price_3pct_vat"] = SOURCE_ORIGINAL
        if p13 is not None:
            src["price_13pct_vat"] = SOURCE_ORIGINAL
        if pn is not None:
            src["price_normal_invoice"] = SOURCE_ORIGINAL
        if pr is not None:
            src["price_reverse_invoice"] = SOURCE_ORIGINAL

        return {
            "冶炼厂名": sm,
            "冶炼厂id": None,
            "品类名": cat,
            "品类id": None,
            "价格": net,
            "价格口径": basis,
            "备注": remark,
            "价格_1pct增值税": p1,
            "价格_3pct增值税": p3,
            "价格_13pct增值税": p13,
            "普通发票价格": pn,
            "反向发票价格": pr,
            "价格字段来源": src,
        }

    def _parse_quote_excel_workbook(self, content: bytes, filename: str) -> Tuple[List[Dict[str, Any]], Dict[str, Any], Optional[str]]:
        """解析 xlsx 首工作表为 items + full_data；若「日期」列存在且全日相同则返回 suggested_quote_date。"""
        if not content:
            raise ValueError("文件内容为空")
        try:
            import pandas as pd
        except ImportError as e:
            raise ValueError("服务端未安装 pandas，无法解析 Excel") from e

        try:
            df = pd.read_excel(io.BytesIO(content), engine="openpyxl", sheet_name=0)
        except Exception as e:
            raise ValueError(f"无法读取 Excel（须为 .xlsx）：{e}") from e

        if df.empty:
            raise ValueError("表格无数据行")

        logical_to_col: Dict[str, Any] = {}
        for c in df.columns:
            logical = self._classify_quote_excel_column(c)
            if not logical:
                continue
            if logical not in logical_to_col:
                logical_to_col[logical] = c

        if "smelter" not in logical_to_col:
            raise ValueError("表头须包含「冶炼厂」或同义列（如冶炼厂名）")
        if "category" not in logical_to_col:
            raise ValueError("表头须包含「品种」或「品类名」等同义列")

        items: List[Dict[str, Any]] = []
        date_samples: List[str] = []
        preview_rows: List[Dict[str, Any]] = []

        for idx, row in df.iterrows():
            fields: Dict[str, Any] = {}
            for logical, col_key in logical_to_col.items():
                val = row[col_key]
                if pd.isna(val):
                    continue
                if logical == "quote_date":
                    if hasattr(val, "strftime"):
                        fields[logical] = val.strftime("%Y-%m-%d")
                    else:
                        fields[logical] = str(val).strip()[:10]
                else:
                    fields[logical] = val

            if "smelter" not in fields or "category" not in fields:
                continue
            if not str(fields.get("smelter", "")).strip() or not str(
                fields.get("category", "")
            ).strip():
                continue

            if fields.get("quote_date"):
                date_samples.append(str(fields["quote_date"]))

            try:
                item = self._excel_row_dict_to_confirm_item(fields)
            except ValueError:
                continue
            if not any(
                item.get(k) is not None
                for k in (
                    "价格",
                    "价格_1pct增值税",
                    "价格_3pct增值税",
                    "价格_13pct增值税",
                    "普通发票价格",
                    "反向发票价格",
                )
            ):
                continue
            items.append(item)
            preview_rows.append(
                {
                    "row_index": int(idx) + 2,
                    "冶炼厂": item["冶炼厂名"],
                    "品类": item["品类名"],
                }
            )

        if not items:
            raise ValueError(
                "未解析到有效数据行：请确认每行同时有冶炼厂、品种，且至少填写一项价格列"
            )

        suggested: Optional[str] = None
        if date_samples:
            uniq = {d[:10] for d in date_samples if d}
            if len(uniq) == 1:
                suggested = next(iter(uniq))

        full_data: Dict[str, Any] = {
            "source_image": filename,
            "file_name": filename,
            "company_name": "",
            "doc_title": "Excel报价列表",
            "quote_date": suggested or "",
            "execution_date": "",
            "valid_period": "",
            "price_unit": "元/吨",
            "headers": [str(c) for c in df.columns.tolist()],
            "rows": preview_rows[:500],
            "policies": {},
            "footer_notes": [],
            "footer_notes_raw": "",
            "brand_specifications": "",
            "raw_full_text": "",
            "elapsed_time": 0.0,
        }
        return items, full_data, suggested

    def build_quote_list_import_template_excel(self) -> bytes:
        """
        生成报价列表 xlsx 导入模板：首表「导入数据」表头与 export_quote_details_excel / upload_price_table_excel
        对齐；附「填写说明」工作表。
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError as e:
            raise ValueError("服务端未安装 openpyxl，无法生成 Excel") from e

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "导入数据"
        headers = [
            "日期",
            "冶炼厂",
            "品种",
            "基准价",
            "3%含税价",
            "13%含税价",
            "备注",
            "价格口径",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
        for col_letter, width in (
            ("A", 12),
            ("B", 18),
            ("C", 18),
            ("D", 12),
            ("E", 12),
            ("F", 12),
            ("G", 28),
            ("H", 14),
        ):
            ws.column_dimensions[col_letter].width = width

        ws2 = wb.create_sheet("填写说明", 1)
        hints = [
            "1. 首行表头请勿修改；数据从第 2 行填写。填好后使用 POST /tl/upload_price_table_excel 上传本表。",
            "2. 「冶炼厂」须与系统「冶炼厂」字典中的名称完全一致；「品种」为品类名称（可新建）。",
            "3. 「日期」格式 YYYY-MM-DD；若多行日期相同，上传接口可返回 suggested_quote_date 供确认写入时选用。",
            "4. 至少填写一项价格列（如基准价或 3%/13% 含税价）。",
            "5. 「价格口径」用于说明你填写的数字是含税还是不含税：",
            "   - ex_vat：不含税基准价（税前价，默认）",
            "   - incl_1pct：含 1% 税价",
            "   - incl_3pct：含 3% 税价",
            "   - incl_13pct：含 13% 税价",
            "6. 价格口径可留空：系统会先按「备注」识别（如“含3%专票”），识别不到时按 ex_vat 处理。",
            "7. 示例：若你在「3%含税价」填 9500，可将「价格口径」填 incl_3pct；若在「基准价」填 9200，则填 ex_vat（或留空）。",
            "8. 更多同义表头见 upload_price_table_excel 接口说明（与「报价数据导出」列名兼容）。",
        ]
        for i, line in enumerate(hints, start=1):
            ws2.cell(row=i, column=1, value=line)
        ws2.column_dimensions["A"].width = 96

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def upload_price_table_excel(self, files: List[Any]) -> Dict[str, Any]:
        """
        解析 .xlsx 报价列表（首工作表），返回与 upload_price_table 相同结构的 data.details，
        供前端展示并调用 confirm_price_table 写入。表头支持导出模板列名及常见同义词。
        """
        details: List[Dict[str, Any]] = []
        for upload_file in files:
            name = upload_file.filename or "upload.xlsx"
            try:
                content = upload_file.file.read()
                items, full_data, suggested = self._parse_quote_excel_workbook(
                    content, name
                )
                entry: Dict[str, Any] = {
                    "file": name,
                    "success": True,
                    "full_data": full_data,
                    "items": items,
                }
                if suggested:
                    entry["suggested_quote_date"] = suggested
                details.append(entry)
            except ValueError as e:
                details.append({"file": name, "success": False, "error": str(e)})
            except Exception as e:
                logger.error(f"解析报价 Excel 失败: {name} | {e}")
                details.append({"file": name, "success": False, "error": str(e)})

        return {"code": 200, "data": {"details": details}}

    def _map_vlm_to_confirm_items(self, result) -> List[Dict[str, Any]]:
        """将 VLM 结果映射为确认条目：图上可能是基准价或含税价（多列/备注）。此处仅带出 OCR 显式列与预览用不含税/反算（默认税率占位）；确认写入时用冶炼厂系统税率做「基准↔含税」双向统一。"""
        items = []
        doc_factory = (result.company_name or "").strip()
        defaults = merge_factory_rates(None)
        for row in result.rows:
            price_normal = row.price_normal_invoice
            price_reverse = row.price_reverse_invoice
            src: Dict[str, str] = {}

            if row.exclusive_net is not None:
                net_f = round(float(row.exclusive_net), 2)
                basis = getattr(row, "price_basis", None) or parse_price_basis_from_remark(
                    row.remark
                )
                fp1 = float(row.price_1pct_vat) if row.price_1pct_vat is not None else None
                fp3 = float(row.price_3pct_vat) if row.price_3pct_vat is not None else None
                fp13 = float(row.price_13pct_vat) if row.price_13pct_vat is not None else None
                src["unit_price"] = SOURCE_ORIGINAL
                if fp1 is not None:
                    src["price_1pct_vat"] = SOURCE_ORIGINAL
                if fp3 is not None:
                    src["price_3pct_vat"] = SOURCE_ORIGINAL
                if fp13 is not None:
                    src["price_13pct_vat"] = SOURCE_ORIGINAL
            else:
                basis = parse_price_basis_from_remark(row.remark)
                pg = row.price_general
                if pg is not None:
                    net_f, _, _, _ = derive_vat_prices_from_stated_price(
                        float(pg), basis, None
                    )
                    net_f = round(net_f, 2)
                    fp1 = float(row.price_1pct_vat) if row.price_1pct_vat is not None else None
                    fp3 = float(row.price_3pct_vat) if row.price_3pct_vat is not None else None
                    fp13 = float(row.price_13pct_vat) if row.price_13pct_vat is not None else None
                    src["unit_price"] = SOURCE_DERIVED
                    if fp1 is not None:
                        src["price_1pct_vat"] = SOURCE_ORIGINAL
                    if fp3 is not None:
                        src["price_3pct_vat"] = SOURCE_ORIGINAL
                    if fp13 is not None:
                        src["price_13pct_vat"] = SOURCE_ORIGINAL
                elif row.price_3pct_vat is not None:
                    net_f = round(
                        net_from_inclusive(float(row.price_3pct_vat), defaults["3pct"]), 2
                    )
                    fp1 = float(row.price_1pct_vat) if row.price_1pct_vat is not None else None
                    fp3 = float(row.price_3pct_vat)
                    fp13 = float(row.price_13pct_vat) if row.price_13pct_vat is not None else None
                    src["unit_price"] = SOURCE_DERIVED
                    src["price_3pct_vat"] = SOURCE_ORIGINAL
                    if fp1 is not None:
                        src["price_1pct_vat"] = SOURCE_ORIGINAL
                    if fp13 is not None:
                        src["price_13pct_vat"] = SOURCE_ORIGINAL
                elif row.price_13pct_vat is not None:
                    net_f = round(
                        net_from_inclusive(float(row.price_13pct_vat), defaults["13pct"]), 2
                    )
                    fp1 = float(row.price_1pct_vat) if row.price_1pct_vat is not None else None
                    fp3 = float(row.price_3pct_vat) if row.price_3pct_vat is not None else None
                    fp13 = float(row.price_13pct_vat)
                    src["unit_price"] = SOURCE_DERIVED
                    src["price_13pct_vat"] = SOURCE_ORIGINAL
                    if fp1 is not None:
                        src["price_1pct_vat"] = SOURCE_ORIGINAL
                    if fp3 is not None:
                        src["price_3pct_vat"] = SOURCE_ORIGINAL
                elif row.price_1pct_vat is not None:
                    net_f = round(
                        net_from_inclusive(float(row.price_1pct_vat), defaults["1pct"]), 2
                    )
                    fp1 = float(row.price_1pct_vat)
                    fp3 = float(row.price_3pct_vat) if row.price_3pct_vat is not None else None
                    fp13 = float(row.price_13pct_vat) if row.price_13pct_vat is not None else None
                    src["unit_price"] = SOURCE_DERIVED
                    src["price_1pct_vat"] = SOURCE_ORIGINAL
                    if fp3 is not None:
                        src["price_3pct_vat"] = SOURCE_ORIGINAL
                    if fp13 is not None:
                        src["price_13pct_vat"] = SOURCE_ORIGINAL
                else:
                    net_f, fp1, fp3, fp13 = None, None, None, None

            if price_normal is not None:
                src["price_normal_invoice"] = SOURCE_ORIGINAL
            if price_reverse is not None:
                src["price_reverse_invoice"] = SOURCE_ORIGINAL

            row_factory = (getattr(row, "factory_name", None) or "").strip()
            item_factory = row_factory or doc_factory

            items.append({
                "冶炼厂名": item_factory,
                "冶炼厂id": None,
                "品类名": row.category,
                "品类id": None,
                "价格": net_f,
                "价格口径": basis,
                "备注": row.remark or None,
                "价格_1pct增值税": fp1,
                "价格_3pct增值税": fp3,
                "价格_13pct增值税": fp13,
                "普通发票价格": float(price_normal) if price_normal is not None else None,
                "反向发票价格": float(price_reverse) if price_reverse is not None else None,
                "价格字段来源": src,
            })
        return items

    # ==================== 接口5b：确认价格表写入数据库 ====================

    def confirm_price_table(
        self,
        quote_date_str: str,
        items: List[Dict[str, Any]],
        full_data: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        if not items:
            raise ValueError("报价数据不能为空")

        try:
            quote_dt = date.fromisoformat(quote_date_str)
        except (ValueError, TypeError):
            raise ValueError(f"日期格式不正确: {quote_date_str}，应为 YYYY-MM-DD")

        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    inserted, updated = 0, 0

                    for item in items:
                        # 1. 冶炼厂：须已在字典中存在且名称与库中完全一致；禁止确认写入时静默新建
                        if item.get("冶炼厂id") is None:
                            factory_name = str(item.get("冶炼厂名") or "").strip()
                            if not factory_name:
                                raise ValueError("冶炼厂名称不能为空")
                            cur.execute(
                                "SELECT id, is_active FROM dict_factories WHERE name = %s",
                                (factory_name,),
                            )
                            row = cur.fetchone()
                            if not row:
                                raise ValueError(
                                    f"冶炼厂「{factory_name}」在系统中不存在，或与字典中的名称不完全一致"
                                    f"（须与「冶炼厂」管理中的名称一字不差，例如全称「安徽鲁控环保有限公司」）。"
                                    f"请修正名称或先在字典中新增该冶炼厂。"
                                )
                            fid_row, active = int(row[0]), row[1]
                            if active is not None and int(active) != 1:
                                raise ValueError(
                                    f"冶炼厂「{factory_name}」已停用，请先在冶炼厂管理中启用后再写入报价。"
                                )
                            item["冶炼厂id"] = fid_row
                        else:
                            fid = int(item["冶炼厂id"])
                            cur.execute(
                                "SELECT is_active FROM dict_factories WHERE id = %s",
                                (fid,),
                            )
                            row = cur.fetchone()
                            if not row:
                                raise ValueError(f"冶炼厂 id={fid} 不存在，请刷新后重新选择冶炼厂。")
                            active = row[0]
                            if active is not None and int(active) != 1:
                                raise ValueError(
                                    f"冶炼厂 id={fid} 已停用，无法写入报价；请启用后重试。"
                                )

                        # 2. 品类：别名归一到启用主名称后落库；停用品类不在确认时自动恢复
                        original_cat_name = str(item["品类名"]).strip()
                        main_cat_name, category_id = self._resolve_quote_category_main_name(
                            cur,
                            original_cat_name,
                            allow_create=True,
                        )
                        item["品类名"] = main_cat_name
                        item["品类id"] = category_id
                        if original_cat_name and original_cat_name != main_cat_name:
                            item["识别品类名"] = original_cat_name

                    # 3. 存储全量元数据（如果有 full_data）
                    metadata_id = None
                    if full_data:
                        # 取第一条 item 的冶炼厂id作为元数据的 factory_id
                        factory_id_for_meta = items[0].get("冶炼厂id") if items else None
                        if factory_id_for_meta:
                            cur.execute(
                                """
                                INSERT INTO quote_table_metadata
                                (factory_id, quote_date, execution_date, doc_title, subtitle,
                                 valid_period, price_unit, headers, footer_notes, footer_notes_raw,
                                 brand_specifications, policies, raw_full_text, source_image)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                                ON DUPLICATE KEY UPDATE
                                    execution_date = VALUES(execution_date),
                                    doc_title = VALUES(doc_title),
                                    subtitle = VALUES(subtitle),
                                    valid_period = VALUES(valid_period),
                                    price_unit = VALUES(price_unit),
                                    headers = VALUES(headers),
                                    footer_notes = VALUES(footer_notes),
                                    footer_notes_raw = VALUES(footer_notes_raw),
                                    brand_specifications = VALUES(brand_specifications),
                                    policies = VALUES(policies),
                                    raw_full_text = VALUES(raw_full_text),
                                    source_image = VALUES(source_image),
                                    updated_at = CURRENT_TIMESTAMP
                                """,
                                (
                                    factory_id_for_meta,
                                    quote_dt,
                                    full_data.get("execution_date", ""),
                                    full_data.get("doc_title", ""),
                                    full_data.get("subtitle", ""),
                                    full_data.get("valid_period", ""),
                                    full_data.get("price_unit", "元/吨"),
                                    json.dumps(full_data.get("headers", []), ensure_ascii=False),
                                    json.dumps(full_data.get("footer_notes", []), ensure_ascii=False),
                                    full_data.get("footer_notes_raw", ""),
                                    full_data.get("brand_specifications", ""),
                                    json.dumps(full_data.get("policies", {}), ensure_ascii=False),
                                    full_data.get("raw_full_text", ""),
                                    full_data.get("source_image", full_data.get("file_name", "")),
                                ),
                            )
                            # 取 metadata_id（INSERT 或 已存在的）
                            if cur.lastrowid:
                                metadata_id = cur.lastrowid
                            else:
                                cur.execute(
                                    "SELECT id FROM quote_table_metadata WHERE factory_id=%s AND quote_date=%s",
                                    (factory_id_for_meta, quote_dt),
                                )
                                row = cur.fetchone()
                                metadata_id = row[0] if row else None

                    # 3b. 按冶炼厂 factory_tax_rates（与默认合并）统一计算「价格」与含1%/3%/13%价（覆盖上传预览推算）
                    factory_ids = list({item["冶炼厂id"] for item in items})
                    tax_by_fid: Dict[int, Dict[str, float]] = {}
                    if factory_ids:
                        fph = ",".join(["%s"] * len(factory_ids))
                        cur.execute(
                            f"SELECT factory_id, tax_type, tax_rate FROM factory_tax_rates "
                            f"WHERE factory_id IN ({fph})",
                            tuple(factory_ids),
                        )
                        for fid, ttype, tr in cur.fetchall():
                            tax_by_fid.setdefault(int(fid), {})[str(ttype)] = float(tr)
                    snapshots = [{k: it.get(k) for k in API_KEY_TO_DB} for it in items]

                    applied_factory_tax: List[bool] = []
                    for item in items:
                        applied_factory_tax.append(
                            _apply_factory_tax_rates_to_quote_item(item, tax_by_fid)
                        )

                    final_sources_list: List[Dict[str, str]] = []
                    for item, snap, tax_applied in zip(items, snapshots, applied_factory_tax):
                        client_src = normalize_client_sources(item.get("价格字段来源"))
                        merged_src = merge_sources_after_fill(item, snap, client_src)
                        if tax_applied:
                            merged_src["price_1pct_vat"] = SOURCE_DERIVED
                            merged_src["price_3pct_vat"] = SOURCE_DERIVED
                            merged_src["price_13pct_vat"] = SOURCE_DERIVED
                            merged_src["unit_price"] = (
                                SOURCE_ORIGINAL
                                if snap.get("价格") is not None
                                else SOURCE_DERIVED
                            )
                        final_sources_list.append(merged_src)

                    # 4. 写入明细，相同(日期+冶炼厂+品类名)则更新价格
                    written_sources: List[Dict[str, Any]] = []
                    for item, final_src in zip(items, final_sources_list):
                        src_json = json.dumps(final_src, ensure_ascii=False) if final_src else None
                        cur.execute(
                            """
                            INSERT INTO quote_details
                            (quote_date, factory_id, category_name, metadata_id,
                             unit_price, price_1pct_vat, price_3pct_vat, price_13pct_vat,
                             price_normal_invoice, price_reverse_invoice, price_field_sources)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            ON DUPLICATE KEY UPDATE
                                metadata_id = VALUES(metadata_id),
                                unit_price = VALUES(unit_price),
                                price_1pct_vat = VALUES(price_1pct_vat),
                                price_3pct_vat = VALUES(price_3pct_vat),
                                price_13pct_vat = VALUES(price_13pct_vat),
                                price_normal_invoice = VALUES(price_normal_invoice),
                                price_reverse_invoice = VALUES(price_reverse_invoice),
                                price_field_sources = VALUES(price_field_sources),
                                updated_at = CURRENT_TIMESTAMP
                            """,
                            (
                                quote_dt,
                                item["冶炼厂id"],
                                item["品类名"],
                                metadata_id,
                                item.get("价格"),
                                item.get("价格_1pct增值税"),
                                item.get("价格_3pct增值税"),
                                item.get("价格_13pct增值税"),
                                item.get("普通发票价格"),
                                item.get("反向发票价格"),
                                src_json,
                            ),
                        )
                        if cur.rowcount == 1:
                            inserted += 1
                        else:
                            updated += 1
                        written_sources.append(
                            {
                                "冶炼厂id": item["冶炼厂id"],
                                "品类名": item["品类名"],
                                "品类id": item.get("品类id"),
                                "识别品类名": item.get("识别品类名"),
                                "价格字段来源": final_src,
                            }
                        )

            log_finance_event(
                "报价确认写入 | 报价日期=%s | 新增=%s | 更新=%s | 条目数=%s | 冶炼厂ids=%s",
                quote_date_str,
                inserted,
                updated,
                len(items),
                sorted({int(i["冶炼厂id"]) for i in items}),
            )
            return {
                "code": 200,
                "msg": f"写入成功：新增 {inserted} 条，更新 {updated} 条",
                "明细价格来源": written_sources,
            }

        except ValueError:
            raise
        except Exception as e:
            logger.error(f"确认价格表写入失败: {e}")
            raise

    def manual_quote_entry(
        self,
        quote_date_str: str,
        items: List[Dict[str, Any]],
        full_data: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        """手写/表格录入报价，逻辑与 confirm_price_table 相同（可不传 full_data）。"""
        return self.confirm_price_table(quote_date_str, items, full_data)

    def _quote_detail_row_to_item(self, row: Tuple[Any, ...]) -> Dict[str, Any]:
        """SELECT quote_details 一行 → confirm 用的中文键条目（不含冶炼厂名）。"""
        (
            _rid,
            _qd,
            fid,
            cname,
            _meta_id,
            up,
            p1,
            p3,
            p13,
            pn,
            pr,
            _psrc,
        ) = row

        def _f(v: Any) -> Optional[float]:
            if v is None:
                return None
            return float(v)

        return {
            "冶炼厂id": int(fid),
            "品类名": str(cname),
            "价格": _f(up),
            "价格_1pct增值税": _f(p1),
            "价格_3pct增值税": _f(p3),
            "价格_13pct增值税": _f(p13),
            "普通发票价格": _f(pn),
            "反向发票价格": _f(pr),
        }

    def update_quote_detail(self, body: UpdateQuoteDetailRequest) -> Dict[str, Any]:
        """按 id 更新 quote_details；本次请求中出现的价格字段作为锚点，按冶炼厂税率重算各档含税价。"""
        raw = body.model_dump(exclude_unset=True)
        detail_id = int(raw.pop("id"))
        touched_cn: Set[str] = {k for k in raw if k in QUOTE_PRICE_ANCHOR_ORDER}

        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        SELECT id, quote_date, factory_id, category_name, metadata_id,
                               unit_price, price_1pct_vat, price_3pct_vat, price_13pct_vat,
                               price_normal_invoice, price_reverse_invoice, price_field_sources
                        FROM quote_details WHERE id = %s
                        """,
                        (detail_id,),
                    )
                    row = cur.fetchone()
                    if not row:
                        raise ValueError(f"报价明细不存在: id={detail_id}")

                    item = self._quote_detail_row_to_item(row)

                    for k, v in raw.items():
                        if k == "品类名" and v is not None:
                            item["品类名"] = str(v).strip()
                        else:
                            item[k] = v

                    new_qd: Optional[date] = None
                    if raw.get("报价日期") is not None:
                        try:
                            new_qd = date.fromisoformat(str(raw["报价日期"]).strip())
                        except (ValueError, TypeError):
                            raise ValueError(
                                f"报价日期格式不正确: {raw['报价日期']}，应为 YYYY-MM-DD"
                            )
                    item.pop("报价日期", None)

                    if raw.get("冶炼厂id") is not None:
                        nfid = int(raw["冶炼厂id"])
                        cur.execute(
                            "SELECT id FROM dict_factories WHERE id = %s AND is_active = 1",
                            (nfid,),
                        )
                        if not cur.fetchone():
                            raise ValueError(f"冶炼厂不存在或未启用: id={nfid}")
                        item["冶炼厂id"] = nfid

                    fid = int(item["冶炼厂id"])
                    tax_by_fid: Dict[int, Dict[str, float]] = {fid: {}}
                    cur.execute(
                        "SELECT factory_id, tax_type, tax_rate FROM factory_tax_rates "
                        "WHERE factory_id = %s",
                        (fid,),
                    )
                    for _fid, ttype, tr in cur.fetchall():
                        tax_by_fid.setdefault(int(_fid), {})[str(ttype)] = float(tr)

                    snapshot = {k: item.get(k) for k in API_KEY_TO_DB}
                    client_src = normalize_client_sources(item.get("价格字段来源"))

                    anchor_cn: Optional[str] = None
                    if touched_cn:
                        for c in QUOTE_PRICE_ANCHOR_ORDER:
                            if c in touched_cn and item.get(c) is not None:
                                anchor_cn = c
                                break

                    tax_applied = False
                    if touched_cn:
                        tax_applied = _apply_factory_tax_rates_to_quote_item(
                            item, tax_by_fid, touched_cn
                        )
                        if not tax_applied:
                            raise ValueError(
                                "无法根据本次修改的价格推算不含税基准与各档含税价，"
                                "请至少填写：基准价、某一档 1%/3%/13% 含税价，或普票/反向发票价之一（且非空）"
                            )

                    merged_src = merge_sources_after_fill(item, snapshot, client_src)
                    if tax_applied:
                        merged_src["price_1pct_vat"] = SOURCE_DERIVED
                        merged_src["price_3pct_vat"] = SOURCE_DERIVED
                        merged_src["price_13pct_vat"] = SOURCE_DERIVED
                        if anchor_cn == "价格":
                            merged_src["unit_price"] = SOURCE_ORIGINAL
                        else:
                            merged_src["unit_price"] = SOURCE_DERIVED
                    src_json = (
                        json.dumps(merged_src, ensure_ascii=False) if merged_src else None
                    )

                    qd_val = new_qd if new_qd is not None else row[1]
                    cname_val, category_id = self._resolve_quote_category_main_name(
                        cur,
                        item["品类名"],
                        allow_create=False,
                    )
                    item["品类名"] = cname_val
                    item["品类id"] = category_id

                    try:
                        cur.execute(
                            """
                            UPDATE quote_details SET
                                quote_date = %s,
                                factory_id = %s,
                                category_name = %s,
                                unit_price = %s,
                                price_1pct_vat = %s,
                                price_3pct_vat = %s,
                                price_13pct_vat = %s,
                                price_normal_invoice = %s,
                                price_reverse_invoice = %s,
                                price_field_sources = %s,
                                updated_at = CURRENT_TIMESTAMP
                            WHERE id = %s
                            """,
                            (
                                qd_val,
                                item["冶炼厂id"],
                                cname_val,
                                item.get("价格"),
                                item.get("价格_1pct增值税"),
                                item.get("价格_3pct增值税"),
                                item.get("价格_13pct增值税"),
                                item.get("普通发票价格"),
                                item.get("反向发票价格"),
                                src_json,
                                detail_id,
                            ),
                        )
                    except PyMySQLIntegrityError as e:
                        raise ValueError(
                            "更新后与已有报价冲突（同一冶炼厂、品种、日期只能有一条），"
                            f"请调整日期、冶炼厂或品种名。详情: {e}"
                        ) from e

            log_finance_event(
                "报价明细更新 | id=%s | touched=%s | factory_id=%s",
                detail_id,
                sorted(touched_cn),
                fid,
            )
            return {
                "code": 200,
                "msg": "更新成功",
                "data": {
                    "id": detail_id,
                    "价格字段来源": merged_src,
                    "价格": item.get("价格"),
                    "价格_1pct增值税": item.get("价格_1pct增值税"),
                    "价格_3pct增值税": item.get("价格_3pct增值税"),
                    "价格_13pct增值税": item.get("价格_13pct增值税"),
                    "普通发票价格": item.get("普通发票价格"),
                    "反向发票价格": item.get("反向发票价格"),
                },
            }

        except ValueError:
            raise
        except Exception as e:
            logger.error(f"更新报价明细失败: {e}")
            raise

    # ==================== 接口6：上传运费 ====================

    def upload_freight(self, freight_list: List[Dict[str, Any]]) -> Dict[str, Any]:
        try:
            today = date.today().isoformat()
            with get_conn() as conn:
                with conn.cursor() as cur:
                    for item in freight_list:
                        warehouse_name = item["仓库"]
                        smelter_name = item["冶炼厂"]
                        freight = item["运费"]

                        cur.execute(
                            "SELECT id FROM dict_warehouses WHERE name = %s AND is_active = 1",
                            (warehouse_name,),
                        )
                        wh_row = cur.fetchone()
                        if not wh_row:
                            raise ValueError(f"仓库 '{warehouse_name}' 不存在或未启用")

                        cur.execute(
                            "SELECT id FROM dict_factories WHERE name = %s AND is_active = 1",
                            (smelter_name,),
                        )
                        sm_row = cur.fetchone()
                        if not sm_row:
                            raise ValueError(f"冶炼厂 '{smelter_name}' 不存在或未启用")

                        cur.execute(
                            "INSERT INTO freight_rates "
                            "(factory_id, warehouse_id, price_per_ton, effective_date) "
                            "VALUES (%s, %s, %s, %s) "
                            "ON DUPLICATE KEY UPDATE "
                            "price_per_ton = VALUES(price_per_ton), "
                            "updated_at = CURRENT_TIMESTAMP",
                            (sm_row[0], wh_row[0], freight, today),
                        )
            log_finance_event(
                "运费上传(JSON) | 条数=%s | 生效日期=%s",
                len(freight_list),
                today,
            )
            return {"code": 200, "msg": "运费数据已存入数据库"}

        except ValueError:
            raise
        except Exception as e:
            logger.error(f"上传运费失败: {e}")
            raise

    def build_freight_template_excel(self, warehouse_ids: List[int]) -> bytes:
        """首列：所选库房名称（按传入 id 顺序）；表头：库房 + 全部启用冶炼厂；数据格留空。"""
        if not warehouse_ids:
            raise ValueError("库房id列表不能为空")
        seen: set[int] = set()
        ordered_ids: List[int] = []
        for wid in warehouse_ids:
            if wid in seen:
                continue
            seen.add(wid)
            ordered_ids.append(int(wid))
        try:
            from openpyxl import Workbook

            with get_conn() as conn:
                with conn.cursor() as cur:
                    wh_ph = ",".join(["%s"] * len(ordered_ids))
                    cur.execute(
                        f"SELECT id, name FROM dict_warehouses "
                        f"WHERE id IN ({wh_ph}) AND is_active = 1",
                        tuple(ordered_ids),
                    )
                    wh_map: Dict[int, str] = {int(r[0]): str(r[1]) for r in cur.fetchall()}
                    missing = [i for i in ordered_ids if i not in wh_map]
                    if missing:
                        raise ValueError(f"以下库房不存在或未启用: {missing}")

                    cur.execute(
                        "SELECT name FROM dict_factories WHERE is_active = 1 ORDER BY id"
                    )
                    smelter_names = [str(r[0]) for r in cur.fetchall()]
                    if not smelter_names:
                        raise ValueError("没有可用的冶炼厂，请先在冶炼厂字典中维护")

            wb = Workbook()
            ws = wb.active
            ws.title = "运费配置"
            header = ["库房"] + smelter_names
            ws.append(header)
            for wid in ordered_ids:
                ws.append([wh_map[wid]] + [None] * len(smelter_names))
            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"生成运费模板 Excel 失败: {e}")
            raise

    def import_freight_excel(self, content: bytes) -> Dict[str, Any]:
        """
        解析运费矩阵：在表上方若干行内自动定位「表头行」；按表头文字识别「库房/仓库」列与
        各冶炼厂列（不再假定库房固定在第 1 列、表头固定在第 1 行，避免与图表/标题行同表时错位）。
        与 upload_freight 相同写入 freight_rates（当日生效）；空单元格跳过。
        表头或数据行出现库中不存在的冶炼厂、库房名称时，自动写入 dict_factories / dict_warehouses
        （与 add_smelter / add_warehouse 一致；若名称已存在但已停用则恢复启用）。
        """
        if not content:
            raise ValueError("文件内容为空")
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise ValueError("服务端未安装 openpyxl，无法导入 Excel") from e

        def _coerce_freight(v: Any) -> Optional[float]:
            if v is None:
                return None
            if isinstance(v, bool):
                return None
            if isinstance(v, str):
                s = v.strip()
                if not s:
                    return None
                v = s
            try:
                x = float(v)
            except (TypeError, ValueError):
                return None
            if x < 0:
                raise ValueError("运费不能为负数")
            return round(x, 2)

        def _cell_str(v: Any) -> str:
            if v is None:
                return ""
            return str(v).replace("\u3000", " ").strip()

        def _is_warehouse_header_cell(v: Any) -> bool:
            s = _cell_str(v)
            if not s:
                return False
            zh = frozenset(
                {
                    "库房",
                    "仓库",
                    "合作库房",
                    "库房名称",
                    "仓库名称",
                    "收料库房",
                    "收货仓库",
                }
            )
            if s in zh:
                return True
            low = s.casefold()
            return low in {"warehouse", "warehouse name", "depot"}

        def _find_freight_header_layout(
            all_rows: List[tuple],
            *,
            max_scan_rows: int = 50,
        ) -> Tuple[int, int, List[Tuple[int, str]]]:
            """
            返回 (表头行下标0起, 库房列下标0起, [(冶炼厂列下标, 表头名称), ...])。
            冶炼厂列：表头行中除库房列外、文本非空的列；同名列取第一次出现。
            """
            scan = min(len(all_rows), max_scan_rows)
            for ri in range(scan):
                row = all_rows[ri]
                if not row:
                    continue
                wh_col: Optional[int] = None
                for j, cell in enumerate(row):
                    if _is_warehouse_header_cell(cell):
                        wh_col = j
                        break
                if wh_col is None:
                    continue
                factories: List[Tuple[int, str]] = []
                seen: Set[str] = set()
                for j, cell in enumerate(row):
                    if j == wh_col:
                        continue
                    name = _cell_str(cell)
                    if not name:
                        continue
                    dedup = name.casefold() if name.isascii() else name
                    if dedup in seen:
                        continue
                    seen.add(dedup)
                    factories.append((j, name))
                if factories:
                    return ri, wh_col, factories
            raise ValueError(
                "未识别运费表头：请在表上方前 "
                f"{max_scan_rows} 行内提供一行，其中某一列表头为「库房」或「仓库」等，"
                "且其余非空列表头为冶炼厂名称（勿依赖固定列号；有图表/标题时请把矩阵表头写清）。"
            )

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                raise ValueError("工作表为空")

            hdr_idx, wh_col, factory_cols = _find_freight_header_layout(all_rows)

            stats = {
                "new_wh": 0,
                "new_fa": 0,
                "re_wh": 0,
                "re_fa": 0,
            }

            with get_conn() as conn:
                with conn.cursor() as cur:

                    def _ensure_warehouse_id(name: str) -> int:
                        cur.execute(
                            "SELECT id, is_active FROM dict_warehouses WHERE name = %s",
                            (name,),
                        )
                        row = cur.fetchone()
                        if row:
                            wid, act = int(row[0]), row[1]
                            if act != 1:
                                cur.execute(
                                    "UPDATE dict_warehouses SET is_active = 1 WHERE id = %s",
                                    (wid,),
                                )
                                stats["re_wh"] += 1
                            return wid
                        cur.execute(
                            "INSERT INTO dict_warehouses (name, is_active) VALUES (%s, 1)",
                            (name,),
                        )
                        stats["new_wh"] += 1
                        return int(cur.lastrowid)

                    def _ensure_factory_id(name: str) -> int:
                        cur.execute(
                            "SELECT id, is_active FROM dict_factories WHERE name = %s",
                            (name,),
                        )
                        row = cur.fetchone()
                        if row:
                            fid, act = int(row[0]), row[1]
                            if act != 1:
                                cur.execute(
                                    "UPDATE dict_factories SET is_active = 1 WHERE id = %s",
                                    (fid,),
                                )
                                stats["re_fa"] += 1
                            return fid
                        cur.execute(
                            "INSERT INTO dict_factories (name, is_active) VALUES (%s, 1)",
                            (name,),
                        )
                        stats["new_fa"] += 1
                        return int(cur.lastrowid)

                    factory_by_col: List[Tuple[int, str, int]] = []
                    for j, hname in factory_cols:
                        fid = _ensure_factory_id(hname)
                        factory_by_col.append((j, hname, fid))

                    today = date.today().isoformat()
                    written = 0
                    skipped_rows = 0
                    skipped_cells = 0
                    errors: List[str] = []

                    for offset, row in enumerate(all_rows[hdr_idx + 1 :]):
                        excel_row = hdr_idx + 2 + offset
                        if not row:
                            skipped_rows += 1
                            continue
                        wh_cell = row[wh_col] if wh_col < len(row) else None
                        if wh_cell is None or (
                            isinstance(wh_cell, str) and not wh_cell.strip()
                        ):
                            skipped_rows += 1
                            continue
                        wh_name = str(wh_cell).strip()
                        try:
                            wid = _ensure_warehouse_id(wh_name)
                        except Exception as ex:
                            errors.append(
                                f"第{excel_row}行：库房「{wh_name}」未能写入字典：{ex}"
                            )
                            continue

                        for j, _fname, fid in factory_by_col:
                            if j >= len(row):
                                continue
                            cell_v = row[j]
                            freight = _coerce_freight(cell_v)
                            if freight is None:
                                skipped_cells += 1
                                continue
                            cur.execute(
                                "INSERT INTO freight_rates "
                                "(factory_id, warehouse_id, price_per_ton, effective_date) "
                                "VALUES (%s, %s, %s, %s) "
                                "ON DUPLICATE KEY UPDATE "
                                "price_per_ton = VALUES(price_per_ton), "
                                "updated_at = CURRENT_TIMESTAMP",
                                (fid, wid, freight, today),
                            )
                            written += 1

            msg = f"已写入 {written} 条运费（生效日期 {today}）"
            extra = []
            if stats["new_wh"] or stats["re_wh"]:
                extra.append(f"库房新建 {stats['new_wh']}、恢复启用 {stats['re_wh']}")
            if stats["new_fa"] or stats["re_fa"]:
                extra.append(f"冶炼厂新建 {stats['new_fa']}、恢复启用 {stats['re_fa']}")
            if extra:
                msg += "；" + "；".join(extra)
            log_finance_event(
                "运费Excel导入 | 生效日期=%s | 写入条数=%s | 新建库房=%s 恢复库房=%s | 新建冶炼厂=%s 恢复冶炼厂=%s | 错误条数=%s",
                today,
                written,
                stats["new_wh"],
                stats["re_wh"],
                stats["new_fa"],
                stats["re_fa"],
                len(errors),
            )
            return {
                "code": 200,
                "msg": msg,
                "写入条数": written,
                "新建库房数": stats["new_wh"],
                "恢复启用库房数": stats["re_wh"],
                "新建冶炼厂数": stats["new_fa"],
                "恢复启用冶炼厂数": stats["re_fa"],
                "跳过空单元格数": skipped_cells,
                "跳过空行数": skipped_rows,
                "错误明细": errors if errors else None,
            }
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"导入运费 Excel 失败: {e}")
            raise
        finally:
            wb.close()

    def _build_factory_latest_quote_catalog(
        self, factory_ids: List[int]
    ) -> Dict[int, List[Dict[str, Any]]]:
        """
        每个冶炼厂在系统品类下的「最新」报价：同一品类多别名时取 quote_date 最新的一条；无记录则各价为 null。
        与比价接口取价一致（按冶炼厂+品种名称维度的 MAX(quote_date)）。
        """
        if not factory_ids:
            return {}
        input_factory_ids = sorted({int(fid) for fid in factory_ids})
        fac_ph = ",".join(["%s"] * len(input_factory_ids))
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        f"SELECT id FROM dict_factories WHERE id IN ({fac_ph}) AND is_active = 1",
                        tuple(input_factory_ids),
                    )
                    active_factory_ids = [int(row[0]) for row in cur.fetchall()]
                    if not active_factory_ids:
                        return {}
                    fac_ph = ",".join(["%s"] * len(active_factory_ids))

                    cur.execute(
                        "SELECT category_id, name FROM dict_categories "
                        "WHERE is_active = 1 ORDER BY category_id, is_main DESC, row_id"
                    )
                    cat_id_to_names: Dict[int, List[str]] = {}
                    cat_id_main: Dict[int, str] = {}
                    for cid, name in cur.fetchall():
                        cid = int(cid)
                        n = str(name).strip()
                        if not n:
                            continue
                        cat_id_to_names.setdefault(cid, []).append(n)
                        if cid not in cat_id_main:
                            cat_id_main[cid] = n

                    cur.execute(
                        f"""
                        SELECT qd.factory_id, qd.category_name, qd.quote_date,
                               qd.unit_price, qd.price_3pct_vat, qd.price_13pct_vat
                        FROM quote_details qd
                        JOIN (
                            SELECT factory_id, category_name, MAX(quote_date) AS mq
                            FROM quote_details
                            WHERE factory_id IN ({fac_ph})
                            GROUP BY factory_id, category_name
                        ) t ON qd.factory_id = t.factory_id
                           AND qd.category_name = t.category_name
                           AND qd.quote_date = t.mq
                        WHERE qd.factory_id IN ({fac_ph})
                        """,
                        tuple(active_factory_ids) + tuple(active_factory_ids),
                    )
                    # (fid, name) -> (quote_date, unit, p3, p13)
                    latest_by_pair: Dict[Tuple[int, str], Tuple[Any, Any, Any, Any]] = {}
                    for fid, cname, qd_d, up, p3, p13 in cur.fetchall():
                        latest_by_pair[(int(fid), str(cname).strip())] = (
                            qd_d,
                            up,
                            p3,
                            p13,
                        )

            out: Dict[int, List[Dict[str, Any]]] = {
                int(fid): [] for fid in active_factory_ids
            }
            sorted_cids = sorted(cat_id_to_names.keys())
            for fid in active_factory_ids:
                for cid in sorted_cids:
                    display = cat_id_main.get(cid, cat_id_to_names[cid][0])
                    best: Optional[Tuple[Any, Any, Any, Any]] = None
                    best_d: Optional[date] = None
                    for alias in cat_id_to_names[cid]:
                        key = (fid, alias)
                        if key not in latest_by_pair:
                            continue
                        qd_d, up, p3, p13 = latest_by_pair[key]
                        cmp_d = qd_d
                        if isinstance(cmp_d, datetime):
                            cmp_d = cmp_d.date()
                        if best_d is None or (
                            isinstance(cmp_d, date) and cmp_d > best_d
                        ):
                            best_d = cmp_d if isinstance(cmp_d, date) else None
                            best = (qd_d, up, p3, p13)
                    if best is None:
                        out[fid].append(
                            {
                                "品类id": cid,
                                "品种": display,
                                "报价日期": None,
                                "普通价": None,
                                "3%含税价": None,
                                "13%含税价": None,
                            }
                        )
                    else:
                        qd_d, up, p3, p13 = best
                        out[fid].append(
                            {
                                "品类id": cid,
                                "品种": display,
                                "报价日期": _cell_json(qd_d),
                                "普通价": _cell_json(up),
                                "3%含税价": _cell_json(p3),
                                "13%含税价": _cell_json(p13),
                            }
                        )
            return out
        except Exception as e:
            logger.error(f"构建冶炼厂最新报价目录失败: {e}")
            raise

    # ==================== 接口6b：运费列表 ====================

    def get_freight_list(
        self,
        warehouse_id: Optional[int] = None,
        factory_id: Optional[int] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        page: int = 1,
        page_size: int = 50,
        include_latest_quotes: bool = False,
    ) -> Dict[str, Any]:
        if page < 1:
            raise ValueError("page 必须 >= 1")
        page_size = min(max(page_size, 1), 500)
        d_from: Optional[date] = None
        d_to: Optional[date] = None
        if date_from:
            try:
                d_from = date.fromisoformat(date_from)
            except (ValueError, TypeError):
                raise ValueError(f"date_from 格式不正确: {date_from}，应为 YYYY-MM-DD")
        if date_to:
            try:
                d_to = date.fromisoformat(date_to)
            except (ValueError, TypeError):
                raise ValueError(f"date_to 格式不正确: {date_to}，应为 YYYY-MM-DD")
        if d_from and d_to and d_from > d_to:
            raise ValueError("date_from 不能晚于 date_to")

        conditions: List[str] = ["1=1"]
        params: List[Any] = []
        if warehouse_id is not None:
            conditions.append("fr.warehouse_id = %s")
            params.append(warehouse_id)
        if factory_id is not None:
            conditions.append("fr.factory_id = %s")
            params.append(factory_id)
        if d_from is not None:
            conditions.append("fr.effective_date >= %s")
            params.append(d_from)
        if d_to is not None:
            conditions.append("fr.effective_date <= %s")
            params.append(d_to)
        where_sql = " AND ".join(conditions)
        offset = (page - 1) * page_size

        base_from = (
            "FROM freight_rates fr "
            "JOIN dict_warehouses dw ON fr.warehouse_id = dw.id "
            "JOIN dict_factories df ON fr.factory_id = df.id "
            f"WHERE {where_sql}"
        )
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(f"SELECT COUNT(*) {base_from}", tuple(params))
                    total = cur.fetchone()[0]

                    cur.execute(
                        f"""
                        SELECT fr.id,
                               fr.warehouse_id AS `仓库id`,
                               dw.name AS `仓库名`,
                               fr.factory_id AS `冶炼厂id`,
                               df.name AS `冶炼厂`,
                               fr.price_per_ton AS `运费`,
                               fr.effective_date AS `生效日期`,
                               fr.created_at AS `创建时间`,
                               fr.updated_at AS `更新时间`
                        {base_from}
                        ORDER BY fr.effective_date DESC, fr.id DESC
                        LIMIT %s OFFSET %s
                        """,
                        tuple(params) + (page_size, offset),
                    )
                    cols = [d[0] for d in cur.description]
                    rows = [
                        {c: _cell_json(v) for c, v in zip(cols, r)}
                        for r in cur.fetchall()
                    ]
            data: Dict[str, Any] = {"total": total, "list": rows}
            if include_latest_quotes and rows:
                fac_ids = sorted({int(r["冶炼厂id"]) for r in rows})
                data["冶炼厂各品种最新报价"] = self._build_factory_latest_quote_catalog(
                    fac_ids
                )
            return {"code": 200, "data": data}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"获取运费列表失败: {e}")
            raise

    # ==================== 接口6c：编辑运费 ====================

    def update_freight(
        self,
        freight_id: int,
        price_per_ton: float,
        effective_date_str: Optional[str] = None,
    ) -> Dict[str, Any]:
        """按主键更新运费单价；可选修改生效日期（须满足 uk_factory_warehouse_date）。"""
        if freight_id < 1:
            raise ValueError("运费id 无效")
        new_ed: Optional[date] = None
        if effective_date_str is not None and str(effective_date_str).strip() != "":
            try:
                new_ed = date.fromisoformat(str(effective_date_str).strip())
            except (ValueError, TypeError):
                raise ValueError(f"生效日期格式不正确: {effective_date_str}，应为 YYYY-MM-DD")

        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT factory_id, warehouse_id, effective_date, price_per_ton "
                        "FROM freight_rates WHERE id = %s",
                        (freight_id,),
                    )
                    row = cur.fetchone()
                    if not row:
                        raise ValueError(f"运费记录不存在: id={freight_id}")
                    factory_id, warehouse_id, current_ed = int(row[0]), int(row[1]), row[2]
                    old_price_per_ton = float(row[3])
                    if isinstance(current_ed, datetime):
                        current_ed = current_ed.date()

                    target_ed = new_ed if new_ed is not None else current_ed

                    if new_ed is not None and new_ed != current_ed:
                        cur.execute(
                            "SELECT id FROM freight_rates "
                            "WHERE factory_id = %s AND warehouse_id = %s "
                            "AND effective_date = %s AND id <> %s",
                            (factory_id, warehouse_id, new_ed, freight_id),
                        )
                        if cur.fetchone():
                            raise ValueError(
                                "该仓库与冶炼厂在目标生效日期已存在其它运费记录，无法改为该日期"
                            )

                    cur.execute(
                        "UPDATE freight_rates SET price_per_ton = %s, effective_date = %s, "
                        "updated_at = CURRENT_TIMESTAMP WHERE id = %s",
                        (price_per_ton, target_ed, freight_id),
                    )
                    if cur.rowcount == 0:
                        raise ValueError(f"更新失败: id={freight_id}")

            log_finance_event(
                "运费修改 | id=%s factory_id=%s warehouse_id=%s 原单价=%s 新单价=%s 原生效日=%s 新生效日=%s",
                freight_id,
                factory_id,
                warehouse_id,
                old_price_per_ton,
                price_per_ton,
                current_ed.isoformat() if hasattr(current_ed, "isoformat") else current_ed,
                target_ed.isoformat() if hasattr(target_ed, "isoformat") else target_ed,
            )
            return {"code": 200, "msg": "运费已更新"}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"更新运费失败: {e}")
            raise

    # ==================== 接口6d：删除运费 ====================

    def delete_freight(self, freight_id: int) -> Dict[str, Any]:
        """按主键物理删除 `freight_rates` 一条记录（与 6c 使用同一 id）。"""
        if freight_id < 1:
            raise ValueError("运费id 无效")
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT fr.id, fr.factory_id, fr.warehouse_id, fr.price_per_ton, "
                        "fr.effective_date, df.name AS smelter_name, dw.name AS warehouse_name "
                        "FROM freight_rates fr "
                        "INNER JOIN dict_factories df ON df.id = fr.factory_id "
                        "INNER JOIN dict_warehouses dw ON dw.id = fr.warehouse_id "
                        "WHERE fr.id = %s",
                        (freight_id,),
                    )
                    snap = cur.fetchone()
                    if not snap:
                        raise ValueError(f"运费记录不存在: id={freight_id}")
                    _id, _fid, _wid, old_price, eff_d, sm_name, wh_name = snap
                    if isinstance(eff_d, datetime):
                        eff_d = eff_d.date()
                    cur.execute("DELETE FROM freight_rates WHERE id = %s", (freight_id,))
                    if cur.rowcount == 0:
                        raise ValueError(f"运费记录不存在: id={freight_id}")
            log_finance_event(
                "运费删除 | id=%s 冶炼厂=%s(id=%s) 仓库=%s(id=%s) 单价=%s 生效日=%s",
                freight_id,
                sm_name,
                _fid,
                wh_name,
                _wid,
                float(old_price),
                eff_d.isoformat() if hasattr(eff_d, "isoformat") else eff_d,
            )
            return {"code": 200, "msg": "运费已删除"}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"删除运费失败: {e}")
            raise

    def _prepare_quote_details_filter(
        self,
        factory_id: Optional[int],
        category_id: Optional[int],
        quote_date: Optional[str],
        date_from: Optional[str],
        date_to: Optional[str],
        category_name: Optional[str],
        category_exact: bool,
    ) -> Tuple[str, List[Any]]:
        """报价明细列表/导出共用的 WHERE 与参数（含日期与品种校验）。"""
        qd_exact: Optional[date] = None
        d_from: Optional[date] = None
        d_to: Optional[date] = None
        def _parse_filter_date(label: str, raw: Optional[str]) -> Optional[date]:
            if raw is None or str(raw).strip() == "":
                return None
            s = str(raw).strip().replace("/", "-")
            try:
                return date.fromisoformat(s)
            except (ValueError, TypeError):
                raise ValueError(f"{label} 格式不正确: {raw}，应为 YYYY-MM-DD（可用 / 分隔）")

        if quote_date:
            qd_exact = _parse_filter_date("quote_date", quote_date)
        if date_from:
            d_from = _parse_filter_date("date_from", date_from)
        if date_to:
            d_to = _parse_filter_date("date_to", date_to)
        if d_from and d_to and d_from > d_to:
            raise ValueError("date_from 不能晚于 date_to")

        conditions: List[str] = ["1=1"]
        params: List[Any] = []
        if factory_id is not None:
            conditions.append("qd.factory_id = %s")
            params.append(factory_id)
        if category_id is not None:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT name FROM dict_categories WHERE category_id = %s AND is_active = 1",
                        (int(category_id),),
                    )
                    category_names = [str(row[0]).strip() for row in cur.fetchall() if str(row[0]).strip()]
            if category_names:
                placeholders = ", ".join(["%s"] * len(category_names))
                conditions.append(f"qd.category_name IN ({placeholders})")
                params.extend(category_names)
            else:
                conditions.append("1=0")
        if qd_exact is not None:
            conditions.append("qd.quote_date = %s")
            params.append(qd_exact)
        if d_from is not None:
            conditions.append("qd.quote_date >= %s")
            params.append(d_from)
        if d_to is not None:
            conditions.append("qd.quote_date <= %s")
            params.append(d_to)
        if category_name and category_id is None:
            if category_exact:
                conditions.append("qd.category_name = %s")
                params.append(category_name)
            else:
                names = [
                    x.strip()
                    for x in re.split(r"[、,，]+", str(category_name))
                    if x.strip()
                ]
                if not names:
                    names = [str(category_name).strip()]
                conditions.append("(" + " OR ".join(["qd.category_name LIKE %s"] * len(names)) + ")")
                params.extend([f"%{name}%" for name in names])
        where_sql = " AND ".join(conditions)
        return where_sql, params

    # ==================== 接口5c：报价数据列表 ====================

    def get_quote_details_list(
        self,
        factory_id: Optional[int] = None,
        category_id: Optional[int] = None,
        quote_date: Optional[str] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        category_name: Optional[str] = None,
        category_exact: bool = False,
        page: int = 1,
        page_size: int = 50,
        response_format: str = "full",
    ) -> Dict[str, Any]:
        if response_format not in ("full", "table"):
            raise ValueError('response_format 仅支持 "full" 或 "table"')
        if page < 1:
            raise ValueError("page 必须 >= 1")
        page_size = min(max(page_size, 1), 500)
        where_sql, params = self._prepare_quote_details_filter(
            factory_id=factory_id,
            category_id=category_id,
            quote_date=quote_date,
            date_from=date_from,
            date_to=date_to,
            category_name=category_name,
            category_exact=category_exact,
        )
        offset = (page - 1) * page_size

        base_from = (
            "FROM quote_details qd "
            "JOIN dict_factories df ON qd.factory_id = df.id "
            "LEFT JOIN quote_table_metadata qtm ON qd.metadata_id = qtm.id "
            "WHERE df.is_active = 1 "
            "AND EXISTS ("
            "SELECT 1 FROM dict_categories dc "
            "WHERE dc.name = TRIM(qd.category_name) AND dc.is_active = 1"
            f") AND {where_sql}"
        )
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(f"SELECT COUNT(*) {base_from}", tuple(params))
                    total = cur.fetchone()[0]

                    cur.execute(
                        f"""
                        SELECT qd.id,
                               qd.quote_date AS `报价日期`,
                               qd.factory_id AS `冶炼厂id`,
                               df.name AS `冶炼厂`,
                               qd.category_name AS `品类名`,
                               qd.metadata_id,
                               qd.unit_price AS `普通价`,
                               qd.price_1pct_vat AS `价格_1pct增值税`,
                               qd.price_3pct_vat AS `价格_3pct增值税`,
                               qd.price_13pct_vat AS `价格_13pct增值税`,
                               qd.price_normal_invoice AS `普通发票价格`,
                               qd.price_reverse_invoice AS `反向发票价格`,
                               qd.price_field_sources AS `价格字段来源`,
                               qtm.execution_date AS `执行日期`,
                               qtm.doc_title AS `文档标题`,
                               qtm.subtitle AS `副标题`,
                               qtm.valid_period AS `有效期`,
                               qtm.price_unit AS `价格单位`,
                               qtm.headers AS `表头列表`,
                               qtm.footer_notes AS `页脚备注`,
                               qtm.footer_notes_raw AS `页脚备注原文`,
                               qtm.brand_specifications AS `品牌规格说明`,
                               qtm.policies AS `政策信息`,
                               qtm.raw_full_text AS `原始识别文本`,
                               qtm.source_image AS `来源图片`,
                               qtm.source_image AS source_image,
                               qtm.created_at AS `元数据创建时间`,
                               qtm.updated_at AS `元数据更新时间`,
                               qd.created_at AS `创建时间`,
                               qd.updated_at AS `更新时间`
                        {base_from}
                        ORDER BY qd.quote_date DESC, qd.factory_id, qd.category_name, qd.id DESC
                        LIMIT %s OFFSET %s
                        """,
                        tuple(params) + (page_size, offset),
                    )
                    cols = [d[0] for d in cur.description]
                    rows = []
                    for r in cur.fetchall():
                        row: Dict[str, Any] = {}
                        for c, v in zip(cols, r):
                            if c in {"价格字段来源", "表头列表", "页脚备注", "政策信息"}:
                                row[c] = _json_cell_to_dict(v)
                            else:
                                row[c] = _cell_json(v)
                        metadata = {
                            "execution_date": row.get("执行日期"),
                            "doc_title": row.get("文档标题"),
                            "subtitle": row.get("副标题"),
                            "valid_period": row.get("有效期"),
                            "price_unit": row.get("价格单位"),
                            "headers": row.get("表头列表"),
                            "footer_notes": row.get("页脚备注"),
                            "footer_notes_raw": row.get("页脚备注原文"),
                            "brand_specifications": row.get("品牌规格说明"),
                            "policies": row.get("政策信息"),
                            "raw_full_text": row.get("原始识别文本"),
                            "source_image": row.get("source_image"),
                            "created_at": row.get("元数据创建时间"),
                            "updated_at": row.get("元数据更新时间"),
                        }
                        row["报价元数据"] = metadata
                        row["full_data"] = metadata
                        rows.append(row)
            if response_format == "table":
                rows = [
                    {
                        "id": item["id"],
                        "日期": item["报价日期"],
                        "冶炼厂": item["冶炼厂"],
                        "品种": item["品类名"],
                        "基准价": item["普通价"],
                        "3%含税价": item["价格_3pct增值税"],
                        "13%含税价": item["价格_13pct增值税"],
                        "图片": item.get("来源图片"),
                        "source_image": item.get("source_image"),
                    }
                    for item in rows
                ]
            return {"code": 200, "data": {"total": total, "list": rows}}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"获取报价数据列表失败: {e}")
            raise

    # ==================== 接口5d：报价数据导出 Excel ====================

    def export_quote_details_excel(
        self,
        factory_id: Optional[int] = None,
        category_id: Optional[int] = None,
        quote_date: Optional[str] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        category_name: Optional[str] = None,
        category_exact: bool = False,
        max_rows: int = 50000,
    ) -> bytes:
        """与列表接口相同筛选条件，导出与表格列一致的 xlsx（最多 max_rows 行）。"""
        max_rows = min(max(max_rows, 1), 100000)
        where_sql, params = self._prepare_quote_details_filter(
            factory_id=factory_id,
            category_id=category_id,
            quote_date=quote_date,
            date_from=date_from,
            date_to=date_to,
            category_name=category_name,
            category_exact=category_exact,
        )
        base_from = (
            "FROM quote_details qd "
            "JOIN dict_factories df ON qd.factory_id = df.id "
            "WHERE df.is_active = 1 "
            "AND EXISTS ("
            "SELECT 1 FROM dict_categories dc "
            "WHERE dc.name = TRIM(qd.category_name) AND dc.is_active = 1"
            f") AND {where_sql}"
        )
        try:
            from openpyxl import Workbook

            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        f"""
                        SELECT qd.quote_date,
                               df.name,
                               qd.category_name,
                               qd.unit_price,
                               qd.price_3pct_vat,
                               qd.price_13pct_vat
                        {base_from}
                        ORDER BY qd.quote_date DESC, qd.factory_id, qd.category_name, qd.id DESC
                        LIMIT %s
                        """,
                        tuple(params) + (max_rows,),
                    )
                    db_rows = cur.fetchall()

            wb = Workbook()
            ws = wb.active
            ws.title = "报价数据"
            ws.append(["日期", "冶炼厂", "品种", "基准价", "3%含税价", "13%含税价"])
            for row in db_rows:
                qd_d, fname, cname, up, p3, p13 = row
                ws.append(
                    [
                        qd_d.isoformat() if isinstance(qd_d, date) else qd_d,
                        fname,
                        cname,
                        _cell_json(up),
                        _cell_json(p3),
                        _cell_json(p13),
                    ]
                )
            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"导出报价 Excel 失败: {e}")
            raise

    # ==================== 接口7a：获取品类映射表 ====================

    def get_category_mapping(self) -> List[Dict[str, Any]]:
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT row_id, category_id, name, is_main "
                        "FROM dict_categories "
                        "WHERE is_active = 1 "
                        "ORDER BY category_id, is_main DESC, row_id"
                    )
                    rows = cur.fetchall()

            result: Dict[int, Dict[str, Any]] = {}
            for row_id, cat_id, name, is_main in rows:
                if cat_id not in result:
                    result[cat_id] = {
                        "品类id": cat_id,
                        "品类名称": [],
                        "别名行": [],
                    }
                result[cat_id]["别名行"].append(
                    {
                        "行id": row_id,
                        "名称": name,
                        "是否主名称": bool(is_main),
                    }
                )
                if is_main:
                    result[cat_id]["品类名称"].insert(0, name)
                else:
                    result[cat_id]["品类名称"].append(name)

            return list(result.values())
        except Exception as e:
            logger.error(f"获取品类映射表失败: {e}")
            raise

    # ==================== 接口7：更新品类映射表 ====================

    @staticmethod
    def _normalize_category_mapping_names(names: List[str]) -> List[str]:
        norm: List[str] = []
        seen: set = set()
        for raw in names:
            if raw is None:
                continue
            parts = _split_category_alias_names(raw)
            if not parts:
                raise ValueError("品类名称列表中含空名称")
            for n in parts:
                if len(n) > 50:
                    raise ValueError(f"品种名长度不能超过 50: {n!r}")
                if n not in seen:
                    seen.add(n)
                    norm.append(n)
        if not norm:
            raise ValueError("品类名称列表不能为空")
        return norm

    @staticmethod
    def _resolve_replace_batch_name_owners(
        replace_rows: List[Tuple[int, int, List[str]]],
    ) -> Dict[str, int]:
        """
        同一次批量提交里，若同一品种名出现在多个「整组替换」条目中，全局只能落在一个 category_id
        （dict_categories.name UNIQUE）。优先保留在「本条名称数量更多」的分组；条数相同则保留在
        请求中更靠前的条目。返回值：name -> 归属 id；新品类（请求里 品类id<=0）用占位 id -(batch_idx+1)。
        """
        best: Dict[str, Tuple[int, int, int]] = {}
        for batch_idx, cid, norm in replace_rows:
            eff_cid = cid if cid > 0 else -(batch_idx + 1)
            glen = len(norm)
            for n in norm:
                if n not in best:
                    best[n] = (eff_cid, glen, batch_idx)
                else:
                    oc, og, ob = best[n]
                    if glen > og or (glen == og and batch_idx < ob):
                        best[n] = (eff_cid, glen, batch_idx)
        return {n: t[0] for n, t in best.items()}

    def update_category_mapping_batch(
        self,
        items: List[Tuple[int, List[str], bool]],
    ) -> Dict[str, Any]:
        """
        批量更新品类映射：先消解「一名多组」冲突，再逐条写入。
        items: (品类id, 品类名称列表, 仅追加别名)
        """
        normalized: List[Tuple[int, int, List[str], bool]] = []
        for batch_idx, (category_id, names, append_only) in enumerate(items):
            norm = self._normalize_category_mapping_names(names)
            normalized.append((batch_idx, category_id, norm, append_only))

        replace_rows: List[Tuple[int, int, List[str]]] = [
            (bi, cid, norm) for bi, cid, norm, app in normalized if not app
        ]
        owner_by_name = self._resolve_replace_batch_name_owners(replace_rows)

        last_cid: Optional[int] = None
        for batch_idx, category_id, norm, append_only in normalized:
            if append_only:
                r = self.update_category_mapping(
                    category_id=category_id,
                    names=norm,
                    append_only=True,
                )
                last_cid = r.get("品类id")
                continue

            eff_cid = category_id if category_id > 0 else -(batch_idx + 1)
            filtered = [n for n in norm if owner_by_name.get(n, eff_cid) == eff_cid]
            if not filtered:
                if category_id > 0:
                    try:
                        self.delete_category(category_id)
                    except ValueError:
                        pass
                continue
            norm = filtered

            r = self.update_category_mapping(
                category_id=category_id,
                names=norm,
                append_only=False,
            )
            last_cid = r.get("品类id")

        out: Dict[str, Any] = {
            "code": 200,
            "msg": "品类映射表更新成功，数据已存入数据库",
        }
        if last_cid is not None:
            out["品类id"] = last_cid
        return out

    def update_category_mapping(
        self,
        category_id: int,
        names: List[str],
        append_only: bool = False,
    ) -> Dict[str, Any]:
        norm = self._normalize_category_mapping_names(names)

        if append_only and category_id <= 0:
            raise ValueError("仅追加别名时 品类id 须为已有分组（>0）")

        had_active_before = False

        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    if category_id <= 0:
                        cur.execute(
                            "SELECT COALESCE(MAX(category_id), 0) + 1 FROM dict_categories"
                        )
                        category_id = int(cur.fetchone()[0])
                    elif append_only:
                        cur.execute(
                            "SELECT name FROM dict_categories "
                            "WHERE category_id = %s AND is_active = 1 "
                            "ORDER BY is_main DESC, row_id ASC",
                            (category_id,),
                        )
                        existing_order = [row[0] for row in cur.fetchall()]
                        had_active_before = len(existing_order) > 0
                        merged: List[str] = []
                        seen_m: set = set()
                        for n in existing_order + norm:
                            if n not in seen_m:
                                seen_m.add(n)
                                merged.append(n)
                        norm = merged
                    else:
                        # 整组替换：该分组下原启用、且不在本次提交列表中的别名一律软删除
                        ph = ",".join(["%s"] * len(norm))
                        cur.execute(
                            f"UPDATE dict_categories SET is_active = 0 "
                            f"WHERE category_id = %s AND is_active = 1 "
                            f"AND name NOT IN ({ph})",
                            (category_id,) + tuple(norm),
                        )

                    # 将该 category_id 下所有旧记录的 is_main 置为 0
                    cur.execute(
                        "UPDATE dict_categories SET is_main = 0 WHERE category_id = %s",
                        (category_id,),
                    )

                    for i, name in enumerate(norm):
                        is_main = 1 if i == 0 else 0

                        cur.execute(
                            "SELECT row_id, category_id FROM dict_categories WHERE name = %s",
                            (name,),
                        )
                        existing = cur.fetchone()

                        if existing:
                            cur.execute(
                                "UPDATE dict_categories "
                                "SET category_id = %s, is_main = %s, is_active = 1 "
                                "WHERE row_id = %s",
                                (category_id, is_main, existing[0]),
                            )
                        else:
                            # 仅追加且分组原先已有启用别名：新插入的一律为别名，不得成为主名称
                            insert_main = (
                                0
                                if (append_only and had_active_before)
                                else is_main
                            )
                            cur.execute(
                                "INSERT INTO dict_categories "
                                "(category_id, name, is_main, is_active) "
                                "VALUES (%s, %s, %s, 1)",
                                (category_id, name, insert_main),
                            )

            return {
                "code": 200,
                "msg": "品类映射表更新成功，数据已存入数据库",
                "品类id": category_id,
            }

        except ValueError:
            raise
        except Exception as e:
            logger.error(f"更新品类映射失败: {e}")
            raise

    # ==================== 接口7b：按行修改品类别名 ====================

    def update_category_row(
        self,
        row_id: int,
        new_name: Optional[str] = None,
        set_main: Optional[bool] = None,
    ) -> Dict[str, Any]:
        if new_name is None and set_main is None:
            raise ValueError("至少需要提供 品种名 或 设为主名称（true）之一")

        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT row_id, category_id, name FROM dict_categories "
                        "WHERE row_id = %s AND is_active = 1",
                        (row_id,),
                    )
                    found = cur.fetchone()
                    if not found:
                        raise ValueError(f"品类别名不存在或已删除: 行id={row_id}")
                    _rid, cat_id, old_name = found

                    if new_name is not None:
                        nn = str(new_name).strip()
                        if not nn:
                            raise ValueError("品种名不能为空")
                        if len(nn) > 50:
                            raise ValueError("品种名长度不能超过 50")
                        cur.execute(
                            "SELECT row_id FROM dict_categories "
                            "WHERE name = %s AND row_id <> %s AND is_active = 1",
                            (nn, row_id),
                        )
                        if cur.fetchone():
                            raise ValueError(f"品种名「{nn}」已被其它别名使用")
                        cur.execute(
                            "UPDATE dict_categories SET name = %s WHERE row_id = %s",
                            (nn, row_id),
                        )
                        cur.execute(
                            "UPDATE quote_details SET category_name = %s WHERE category_name = %s",
                            (nn, old_name),
                        )

                    if set_main is True:
                        cur.execute(
                            "UPDATE dict_categories SET is_main = 0 WHERE category_id = %s",
                            (cat_id,),
                        )
                        cur.execute(
                            "UPDATE dict_categories SET is_main = 1 WHERE row_id = %s",
                            (row_id,),
                        )

            return {"code": 200, "msg": "品类别名已更新"}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"修改品类别名失败: {e}")
            raise

    # ==================== 接口7c：删除品类分组（软删除） ====================

    def delete_category(self, category_id: int) -> Dict[str, Any]:
        if category_id < 1:
            raise ValueError("品类id 无效")
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "UPDATE dict_categories SET is_active = 0 "
                        "WHERE category_id = %s AND is_active = 1",
                        (category_id,),
                    )
                    n = cur.rowcount
                    if n == 0:
                        raise ValueError(f"品类 id={category_id} 不存在或已删除")
            return {"code": 200, "msg": "品类分组已删除", "影响行数": n}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"删除品类分组失败: {e}")
            raise

    # ==================== 接口7d：删除单条品类别名（软删除） ====================

    def delete_category_row(self, row_id: int) -> Dict[str, Any]:
        if row_id < 1:
            raise ValueError("行id 无效")
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT category_id, is_main FROM dict_categories "
                        "WHERE row_id = %s AND is_active = 1",
                        (row_id,),
                    )
                    row = cur.fetchone()
                    if not row:
                        raise ValueError(f"品类别名不存在或已删除: 行id={row_id}")
                    cat_id, was_main = int(row[0]), int(row[1])

                    cur.execute(
                        "UPDATE dict_categories SET is_active = 0 WHERE row_id = %s",
                        (row_id,),
                    )

                    if was_main:
                        cur.execute(
                            "SELECT row_id FROM dict_categories "
                            "WHERE category_id = %s AND is_active = 1 "
                            "ORDER BY row_id ASC LIMIT 1",
                            (cat_id,),
                        )
                        nxt = cur.fetchone()
                        if nxt:
                            cur.execute(
                                "UPDATE dict_categories SET is_main = 1 WHERE row_id = %s",
                                (nxt[0],),
                            )

            return {"code": 200, "msg": "品类别名已删除"}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"删除品类别名失败: {e}")
            raise

    # ==================== 接口A7：采购建议 ====================

    def get_purchase_suggestion(
        self,
        warehouse_ids: List[int],
        demands: List[Dict[str, Any]],
        price_type: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        根据仓库列表和需求（品类+吨数），查询最新运费和报价（报价选取与 get_comparison 未传「报价日期」一致：按比价基准日取日历距离最近，并列按 created_at 最新）。
        冶炼厂默认取 dict_factories 中全部启用冶炼厂，无需前端传入。
        整理结构化数据后调用 LLM 生成各仓库发车建议表。
        price_type: 目标税率类型，None=普通价, 1pct/3pct/13pct/normal_invoice/reverse_invoice
        """
        if not warehouse_ids or not demands:
            raise ValueError("仓库列表和需求不能为空")

        # price_type → (quote_details列名, 展示名)
        PRICE_COL_MAP = {
            None:             ("unit_price",            "普通价"),
            "1pct":           ("price_1pct_vat",        "1%增值税"),
            "3pct":           ("price_3pct_vat",        "3%增值税"),
            "13pct":          ("price_13pct_vat",       "13%增值税"),
            "normal_invoice": ("price_normal_invoice",  "普通发票"),
            "reverse_invoice":("price_reverse_invoice", "反向发票"),
        }
        VAT_TAX_TYPE_MAP = {"1pct": "1pct", "3pct": "3pct", "13pct": "13pct"}

        if price_type not in PRICE_COL_MAP:
            raise ValueError(f"不支持的 price_type: {price_type}")

        target_col, price_type_name = PRICE_COL_MAP[price_type]
        target_tax = VAT_TAX_TYPE_MAP.get(price_type)

        category_ids = list({d["category_id"] for d in demands})

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id FROM dict_factories WHERE is_active = 1 ORDER BY id"
                )
                smelter_ids = [r[0] for r in cur.fetchall()]
                if not smelter_ids:
                    raise ValueError("没有可用的冶炼厂，请先在 dict_factories 中维护启用冶炼厂")

                wh_ph = ",".join(["%s"] * len(warehouse_ids))
                sm_ph = ",".join(["%s"] * len(smelter_ids))
                cat_ph = ",".join(["%s"] * len(category_ids))

                # 仓库名称
                cur.execute(
                    f"SELECT id, name FROM dict_warehouses WHERE id IN ({wh_ph})",
                    tuple(warehouse_ids),
                )
                warehouse_name_map: Dict[int, str] = {r[0]: r[1] for r in cur.fetchall()}

                # 品类主名称
                cur.execute(
                    f"SELECT category_id, "
                    f"COALESCE(MAX(CASE WHEN is_main=1 THEN name END), MAX(name)) "
                    f"FROM dict_categories "
                    f"WHERE category_id IN ({cat_ph}) AND is_active=1 "
                    f"GROUP BY category_id",
                    tuple(category_ids),
                )
                cat_name_map: Dict[int, str] = {r[0]: r[1] for r in cur.fetchall()}

                # 冶炼厂名称
                cur.execute(
                    f"SELECT id, name FROM dict_factories WHERE id IN ({sm_ph})",
                    tuple(smelter_ids),
                )
                factory_name_map: Dict[int, str] = {r[0]: r[1] for r in cur.fetchall()}

                # 最新运费：每个(仓库, 冶炼厂)取最新日期，保留仓库维度
                cur.execute(
                    f"""
                    SELECT dw.id AS wid, dw.name AS wname,
                           df.id AS fid, df.name AS fname,
                           fr.price_per_ton
                    FROM freight_rates fr
                    JOIN dict_warehouses dw ON fr.warehouse_id = dw.id
                    JOIN dict_factories  df ON fr.factory_id  = df.id
                    WHERE dw.id IN ({wh_ph})
                      AND df.id IN ({sm_ph})
                      AND fr.effective_date = (
                          SELECT MAX(fr2.effective_date)
                          FROM freight_rates fr2
                          WHERE fr2.factory_id  = fr.factory_id
                            AND fr2.warehouse_id = fr.warehouse_id
                      )
                    """,
                    tuple(warehouse_ids) + tuple(smelter_ids),
                )
                # freight_map: {(warehouse_id, factory_id): freight}
                freight_map: Dict[tuple, float] = {
                    (r[0], r[2]): (float(r[4]) if r[4] is not None else 0.0) for r in cur.fetchall()
                }

                # 税率表
                cur.execute(
                    f"SELECT factory_id, tax_type, tax_rate "
                    f"FROM factory_tax_rates WHERE factory_id IN ({sm_ph})",
                    tuple(smelter_ids),
                )
                tax_rate_map: Dict[int, Dict[str, float]] = {}
                for fid, ttype, rate in cur.fetchall():
                    tax_rate_map.setdefault(fid, {})[ttype] = float(rate)

                # category_id → 品类名称列表
                cur.execute(
                    f"SELECT category_id, name FROM dict_categories "
                    f"WHERE category_id IN ({cat_ph}) AND is_active = 1",
                    tuple(category_ids),
                )
                cat_id_to_names: Dict[int, List[str]] = {}
                for cat_id, name in cur.fetchall():
                    n = str(name).strip()
                    if not n:
                        continue
                    lst = cat_id_to_names.setdefault(cat_id, [])
                    if n not in lst:
                        lst.append(n)

                if not cat_id_to_names:
                    return {"demand_rows": [], "raw": []}

                all_cat_names: List[str] = []
                _seen_cn: set = set()
                for names in cat_id_to_names.values():
                    for n in names:
                        if n not in _seen_cn:
                            _seen_cn.add(n)
                            all_cat_names.append(n)
                cn_ph = ",".join(["%s"] * len(all_cat_names))

                ref_day = _comparison_quote_calendar_date()
                cur.execute(
                    f"""
                    SELECT qd.factory_id, TRIM(qd.category_name) AS category_name,
                           qd.unit_price, qd.price_1pct_vat, qd.price_3pct_vat,
                           qd.price_13pct_vat,
                           qd.price_normal_invoice, qd.price_reverse_invoice
                    FROM quote_details qd
                    INNER JOIN (
                        SELECT id FROM (
                            SELECT id,
                                   ROW_NUMBER() OVER (
                                       PARTITION BY factory_id, TRIM(category_name)
                                       ORDER BY ABS(DATEDIFF(quote_date, %s)) ASC,
                                                created_at DESC,
                                                id DESC
                                   ) AS rn
                            FROM quote_details
                            WHERE factory_id IN ({sm_ph})
                              AND TRIM(category_name) IN ({cn_ph})
                        ) ranked
                        WHERE ranked.rn = 1
                    ) pick ON pick.id = qd.id
                    """,
                    (ref_day,) + tuple(smelter_ids) + tuple(all_cat_names),
                )
                col_names = ["unit_price", "price_1pct_vat", "price_3pct_vat",
                             "price_13pct_vat", "price_normal_invoice", "price_reverse_invoice"]
                raw_price_map: Dict[tuple, Dict[str, Optional[float]]] = {}
                for row in cur.fetchall():
                    fid_r, cat_name = row[0], row[1]
                    raw_price_map[(fid_r, cat_name)] = {
                        col: (float(v) if v is not None else None)
                        for col, v in zip(col_names, row[2:])
                    }

                cur.execute(
                    f"SELECT id, COALESCE(use_xunrongbao, 0) FROM dict_factories "
                    f"WHERE id IN ({sm_ph})",
                    tuple(smelter_ids),
                )
                xrb_fids_ps = {int(r[0]) for r in cur.fetchall() if int(r[1]) == 1}
                for map_key in list(raw_price_map.keys()):
                    fid_k, _cn = map_key
                    if fid_k not in xrb_fids_ps:
                        continue
                    merged = merge_factory_rates(tax_rate_map.get(fid_k, {}))
                    raw_price_map[map_key] = apply_per_ton_premium_to_quote_row(
                        raw_price_map[map_key],
                        merged,
                        XUNRONGBAO_SHIPPING_PREMIUM_PER_TON,
                    )

        # 价格反算逻辑
        COL_TO_TAX: Dict[str, str] = {
            "price_1pct_vat": "1pct",
            "price_3pct_vat": "3pct",
            "price_13pct_vat": "13pct",
        }

        def resolve_price(fid: int, cat_id: int) -> Optional[float]:
            cat_names = cat_id_to_names.get(cat_id, [])
            for cat_name in cat_names:
                prices = raw_price_map.get((fid, cat_name), {})
                if not prices:
                    continue

                rates = tax_rate_map.get(fid, {})
                merged = merge_factory_rates(rates)

                direct = prices.get(target_col)
                if direct is not None:
                    return direct

                if target_tax and prices.get("unit_price") is not None and target_tax in merged:
                    return inclusive_from_net(float(prices["unit_price"]), merged[target_tax])

                if target_col == "unit_price":
                    for col, src_tax in COL_TO_TAX.items():
                        known_price = prices.get(col)
                        if known_price is not None and src_tax in merged:
                            net = net_from_inclusive(float(known_price), merged[src_tax])
                            return round(net, 2)

                if target_tax and target_tax in merged:
                    for col, src_tax in COL_TO_TAX.items():
                        known_price = prices.get(col)
                        if known_price is not None and src_tax in merged:
                            net = net_from_inclusive(float(known_price), merged[src_tax])
                            return inclusive_from_net(net, merged[target_tax])

            return None

        # 构建 price_map: {(factory_id, category_id): price}
        price_map: Dict[tuple, Optional[float]] = {}
        for fid in smelter_ids:
            for cid in category_ids:
                price_map[(fid, cid)] = resolve_price(fid, cid)

        # 构造结构化数据：每条需求 × 全部冶炼厂，报价与各仓库运费对比
        # 与 get_comparison 一致：比价利润 = 报价×吨数 − 运费×吨数 → 元/吨档为 (报价 − 运费)
        demand_rows = []
        raw = []
        for d in demands:
            cid = d["category_id"]
            demand_tons = float(d["demand"])
            for fid in smelter_ids:
                fname = factory_name_map.get(fid, f"冶炼厂{fid}")
                cat_name = cat_name_map.get(cid, f"品类{cid}")
                price = price_map.get((fid, cid))

                warehouse_options = []
                for wid in warehouse_ids:
                    wname = warehouse_name_map.get(wid, f"仓库{wid}")
                    freight = freight_map.get((wid, fid))
                    margin_per_ton: Optional[float] = None
                    if price is not None and freight is not None:
                        margin_per_ton = round(float(price) - float(freight), 2)
                    profit_yuan: Optional[float] = None
                    if margin_per_ton is not None:
                        profit_yuan = round(margin_per_ton * demand_tons, 2)
                    warehouse_options.append({
                        "仓库": wname,
                        "运费(元/吨)": freight,
                        "比价利润元每吨": margin_per_ton,
                        "比价利润(元)": profit_yuan,
                    })
                    raw.append({
                        "冶炼厂": fname,
                        "品类": cat_name,
                        "需求吨数": demand_tons,
                        "报价(元/吨)": price,
                        "仓库": wname,
                        "运费(元/吨)": freight,
                        "比价利润元每吨": margin_per_ton,
                        "比价利润(元)": profit_yuan,
                    })

                demand_rows.append({
                    "冶炼厂": fname,
                    "品类": cat_name,
                    "需求吨数(吨)": demand_tons,
                    "报价(元/吨)": price,
                    "各仓库运费对比": warehouse_options,
                })

        # 构造 prompt，调用大模型（OpenAI 兼容协议）
        import json
        from openai import OpenAI
        from app import config as app_config

        if not (app_config.LLM_API_KEY or "").strip():
            raise ValueError(
                "未配置文本大模型密钥，无法生成采购建议。请设置 LLM_API_KEY；若与报价图识别共用阿里云百炼，"
                "也可只配 VLM_API_KEY（或 DASHSCOPE_API_KEY / QWEN_API_KEY），"
                "此时默认使用百炼兼容端点与 qwen-plus。其它厂商请显式配置 LLM_API_KEY、LLM_BASE_URL、LLM_MODEL。"
            )

        client = OpenAI(api_key=app_config.LLM_API_KEY, base_url=app_config.LLM_BASE_URL)
        data_str = json.dumps(demand_rows, ensure_ascii=False, indent=2)
        prompt = f"""以下是各需求的报价及各仓库运费数据：

{data_str}

请给出各仓库发车建议，要求：
1. 与系统比价一致：每条线路的「比价利润(元)」= 报价×吨数 − 运费×吨数（数据中已按此计算）；优先选比价利润更高（更优）的仓库
2. 同仓库不同品类可混装，尽量整车（20-30吨）
3. 按仓库分段输出：仓库名、装车方案（品类+吨数+冶炼厂+比价利润）、备注
4. 数据缺失的在备注注明
5. 纯文本，简洁"""

        try:
            resp = client.chat.completions.create(
                model=app_config.LLM_MODEL,
                max_tokens=2048,
                messages=[{"role": "user", "content": prompt}],
            )
            suggestion = resp.choices[0].message.content
        except Exception as exc:
            err_text = str(exc).lower()
            status = getattr(exc, "status_code", None)
            logger.exception("采购建议大模型调用失败")
            if status == 403 or ("403" in str(exc) and "forbidden" in err_text):
                raise PurchaseSuggestionLLMError(
                    "大模型服务端拒绝请求（HTTP 403）。常见原因：API Key 无效或无权访问该模型、"
                    "LLM_BASE_URL 与密钥不属于同一服务商、控制台 IP 白名单未放行当前服务器、"
                    "套餐/配额或地域策略限制。请在部署环境检查 LLM_API_KEY、LLM_BASE_URL、LLM_MODEL，"
                    "并登录模型服务商控制台核对权限与网络策略。"
                ) from exc
            raise PurchaseSuggestionLLMError(
                f"大模型调用失败，无法生成建议正文。原始错误：{exc}"
            ) from exc

        return {"code": 200, "data": {"suggestion": suggestion, "raw": raw}}


# ==================== 单例工厂 ====================

_tl_service: Optional[TLService] = None


def get_tl_service() -> TLService:
    global _tl_service
    if _tl_service is None:
        _tl_service = TLService()
    return _tl_service
